from __future__ import annotations

import argparse
import math
from datetime import datetime
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


WORK_DIR = Path(__file__).resolve().parent
DOCS_DIR = WORK_DIR / "docs"
ASSETS_DIR = DOCS_DIR / "generated_assets"
DEFAULT_OUTPUT = DOCS_DIR / "huong_dan_kpi_automation.pdf"
GITHUB_REPO_URL = "https://github.com/dungdq-maker/kpi-google-sheets-onedrive-automation-main"
GITHUB_ZIP_URL = "https://github.com/dungdq-maker/kpi-google-sheets-onedrive-automation-main/archive/refs/heads/main.zip"
PAYROLL_SOURCE_PATH = Path(r"C:\Users\admin\OneDrive\BCQT 2026\Vốn hóa chi phí nhân sự 2026.xlsx")

WINDOWS_FONT_CANDIDATES = [
    Path("C:/Windows/Fonts/arial.ttf"),
    Path("C:/Windows/Fonts/tahoma.ttf"),
    Path("C:/Windows/Fonts/segoeui.ttf"),
]
WINDOWS_BOLD_CANDIDATES = [
    Path("C:/Windows/Fonts/arialbd.ttf"),
    Path("C:/Windows/Fonts/tahomabd.ttf"),
    Path("C:/Windows/Fonts/segoeuib.ttf"),
]
WINDOWS_MONO_CANDIDATES = [
    Path("C:/Windows/Fonts/consola.ttf"),
    Path("C:/Windows/Fonts/cour.ttf"),
]


def latest_xlsx(folder: Path) -> Path | None:
    files = [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        return None
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def pick_font(candidates: Iterable[Path]) -> Path | None:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def register_reportlab_fonts() -> tuple[str, str, str]:
    regular = pick_font(WINDOWS_FONT_CANDIDATES)
    bold = pick_font(WINDOWS_BOLD_CANDIDATES)
    mono = pick_font(WINDOWS_MONO_CANDIDATES)

    regular_name = "Helvetica"
    bold_name = "Helvetica-Bold"
    mono_name = "Courier"

    if regular:
        regular_name = "GuideRegular"
        pdfmetrics.registerFont(TTFont(regular_name, str(regular)))
    if bold:
        bold_name = "GuideBold"
        pdfmetrics.registerFont(TTFont(bold_name, str(bold)))
    elif regular:
        bold_name = regular_name
    if mono:
        mono_name = "GuideMono"
        pdfmetrics.registerFont(TTFont(mono_name, str(mono)))

    return regular_name, bold_name, mono_name


def pil_font(size: int, bold: bool = False, mono: bool = False) -> ImageFont.FreeTypeFont:
    candidates = WINDOWS_BOLD_CANDIDATES if bold else WINDOWS_MONO_CANDIDATES if mono else WINDOWS_FONT_CANDIDATES
    font_path = pick_font(candidates)
    if font_path:
        return ImageFont.truetype(str(font_path), size=size)
    return ImageFont.load_default()


def safe_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def file_uri(path: Path) -> str:
    return path.resolve().as_uri()


def github_repo_html() -> str:
    return f'<link href="{GITHUB_REPO_URL}">{GITHUB_REPO_URL}</link>'


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    text = safe_text(text).replace("\r", " ").strip()
    if not text:
        return [""]

    words = text.split()
    if not words:
        return [""]

    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        bbox = draw.textbbox((0, 0), candidate, font=font)
        width = bbox[2] - bbox[0]
        if width <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def multiline_box_size(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int, line_gap: int = 4) -> tuple[int, int, list[str]]:
    lines = wrap_text(draw, text, font, max_width)
    widths = []
    line_height = 0
    for line in lines:
        bbox = draw.textbbox((0, 0), line or " ", font=font)
        widths.append(bbox[2] - bbox[0])
        line_height = max(line_height, bbox[3] - bbox[1])
    height = max(line_height, 1) * len(lines) + line_gap * max(len(lines) - 1, 0)
    return max(widths) if widths else 0, height, lines


def draw_wrapped_text(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill,
    max_width: int,
    line_gap: int = 4,
) -> int:
    x, y = xy
    _, height, lines = multiline_box_size(draw, text, font, max_width, line_gap=line_gap)
    cur_y = y
    bbox = draw.textbbox((0, 0), "Ag", font=font)
    line_height = bbox[3] - bbox[1]
    for line in lines:
        draw.text((x, cur_y), line, font=font, fill=fill)
        cur_y += line_height + line_gap
    return height


def create_canvas(width: int, height: int, bg: str = "#F8FAFC") -> tuple[Image.Image, ImageDraw.ImageDraw]:
    image = Image.new("RGB", (width, height), bg)
    return image, ImageDraw.Draw(image)


def rounded_rect(draw: ImageDraw.ImageDraw, box, fill, outline=None, radius: int = 18, width: int = 2) -> None:
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def save_image(image: Image.Image, name: str) -> Path:
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    path = ASSETS_DIR / name
    image.save(path)
    return path


def build_vscode_diagram() -> Path:
    img, d = create_canvas(1500, 780, "#0F172A")
    font_bold = pil_font(34, bold=True)
    font = pil_font(24)
    font_small = pil_font(20)
    font_mono = pil_font(24, mono=True)

    d.rectangle((0, 0, 1500, 72), fill="#1E293B")
    d.text((36, 20), "Visual Studio Code", font=font_bold, fill="#FFFFFF")
    rounded_rect(d, (1210, 14, 1298, 54), fill="#334155", outline="#475569", radius=12, width=2)
    d.text((1248, 24), "File", font=font_small, fill="#E2E8F0")
    rounded_rect(d, (1308, 14, 1460, 54), fill="#EA580C", outline="#FB923C", radius=12, width=2)
    d.text((1330, 24), "Terminal", font=font_small, fill="#FFFFFF")

    d.rectangle((0, 72, 190, 780), fill="#111827")
    d.text((30, 128), "Explorer", font=font_bold, fill="#F8FAFC")
    d.line((190, 72, 190, 780), fill="#334155", width=2)

    d.rectangle((230, 128, 1405, 540), fill="#F8FAFC", outline="#CBD5E1", width=2)
    d.rectangle((230, 72, 1405, 126), fill="#E2E8F0", outline="#CBD5E1", width=2)
    d.text((254, 84), "Mở thư mục dự án tại đây", font=font_small, fill="#334155")

    d.rectangle((254, 160, 1392, 224), fill="#EFF6FF", outline="#93C5FD", width=2)
    d.text((278, 182), "1. Mở project", font=font, fill="#1D4ED8")
    d.text((520, 182), "2. Mở terminal tích hợp", font=font, fill="#1D4ED8")
    d.text((860, 182), "3. Dán lệnh chạy automation", font=font, fill="#1D4ED8")

    d.rectangle((278, 260, 610, 332), fill="#FFFFFF", outline="#CBD5E1", width=2)
    d.text((300, 286), "Welcome to Visual Studio Code", font=font_small, fill="#0F172A")
    d.text((300, 294), "", font=font_small, fill="#0F172A")

    d.rectangle((930, 150, 1200, 220), fill="#FFF7ED", outline="#FB923C", width=2)
    d.text((952, 174), "New Terminal", font=font_small, fill="#9A3412")
    d.line((1200, 185, 1350, 185), fill="#EA580C", width=5)
    d.polygon([(1350, 185), (1328, 171), (1328, 199)], fill="#EA580C")
    d.text((275, 378), "Cấu trúc thường gặp:", font=font_bold, fill="#0F172A")
    d.text((275, 430), "• Chọn File → Open Folder", font=font, fill="#334155")
    d.text((275, 478), "• Chọn Terminal → New Terminal", font=font, fill="#334155")
    d.text((275, 526), "• Dán lệnh `py -3 automate_kpi.py --download --open`", font=font, fill="#334155")

    d.rectangle((260, 580, 1390, 720), fill="#0B1020", outline="#334155", width=2)
    d.text((288, 608), "PS C:\\project>", font=font_mono, fill="#E5E7EB")
    d.text((288, 652), "py -3 automate_kpi.py --download --open", font=font_mono, fill="#A7F3D0")
    d.text((288, 692), "Nếu `python` không chạy, dùng `py -3` trên Windows.", font=font_small, fill="#FDE68A")
    d.text((36, 738), "Hình minh họa: các nút cần bấm để mở project và chạy lệnh.", font=font_small, fill="#CBD5E1")
    return save_image(img, "vscode_diagram.png")


def build_terminal_diagram() -> Path:
    img, d = create_canvas(1500, 700, "#111827")
    font_bold = pil_font(30, bold=True)
    font = pil_font(22)
    font_small = pil_font(18)
    font_mono = pil_font(22, mono=True)

    d.rectangle((0, 0, 1500, 64), fill="#1F2937")
    d.text((28, 18), "Terminal", font=font_bold, fill="#FFFFFF")
    d.rectangle((24, 96, 1476, 626), fill="#0B1020", outline="#334155", width=2)
    d.text((52, 132), r"PS C:\Users\admin\Downloads\kpi-google-sheets-onedrive-automation-main>", font=font_mono, fill="#E5E7EB")
    d.text((52, 188), "1) Cài thư viện:", font=font, fill="#93C5FD")
    d.text((88, 232), "py -3 -m pip install -r requirements.txt", font=font_mono, fill="#A7F3D0")
    d.text((52, 300), "2) Chạy automation:", font=font, fill="#93C5FD")
    d.text((88, 344), "py -3 automate_kpi.py --download --open", font=font_mono, fill="#A7F3D0")
    d.text((52, 412), "3) Nếu đã có file raw sẵn:", font=font, fill="#93C5FD")
    d.text((88, 456), "py -3 automate_kpi.py", font=font_mono, fill="#A7F3D0")
    rounded_rect(d, (1050, 170, 1420, 270), fill="#1D4ED8", outline="#60A5FA", radius=18, width=2)
    d.text((1088, 206), "Dán lệnh vào đây", font=font_bold, fill="#FFFFFF")
    d.text((1088, 240), "rồi nhấn Enter", font=font_small, fill="#FFFFFF")
    d.line((1048, 220, 930, 312), fill="#60A5FA", width=5)
    d.polygon([(930, 312), (955, 300), (946, 327)], fill="#60A5FA")
    d.text((52, 540), "Nếu PowerShell báo không tìm thấy `python`, hãy dùng `py -3`.", font=font_small, fill="#FDE68A")
    d.text((52, 582), "Lệnh `--download` chỉ cần khi muốn tải lại dữ liệu mới nhất từ Google Sheets.", font=font_small, fill="#CBD5E1")
    return save_image(img, "terminal_diagram.png")


def build_output_diagram() -> Path:
    img, d = create_canvas(1500, 640, "#F8FAFC")
    font_bold = pil_font(28, bold=True)
    font = pil_font(22)
    font_small = pil_font(18)

    d.rectangle((24, 28, 480, 580), fill="#FFFFFF", outline="#CBD5E1", width=2)
    d.text((48, 50), "File Explorer", font=font_bold, fill="#0F172A")
    d.text((48, 124), "data/", font=font, fill="#475569")
    d.text((82, 170), "output/", font=font, fill="#475569")
    d.text((116, 216), "final/", font=font, fill="#475569")
    d.rectangle((148, 262, 432, 308), fill="#DCFCE7", outline="#22C55E", width=2)
    d.text((166, 273), "von_hoa_YYYYMMDD_HHMMSS.xlsx", font=font_small, fill="#166534")
    d.text((48, 344), "Chọn file mới nhất trong thư mục final.", font=font_small, fill="#334155")

    d.rectangle((540, 28, 1460, 580), fill="#0F172A", outline="#334155", width=2)
    d.text((572, 50), "Kết quả sau khi chạy", font=font_bold, fill="#FFFFFF")
    d.rectangle((580, 130, 1400, 186), fill="#1D4ED8", outline="#60A5FA", width=2)
    d.text((608, 147), "data/output/final/von_hoa_YYYYMMDD_HHMMSS.xlsx", font=font, fill="#FFFFFF")
    d.rectangle((580, 226, 1400, 282), fill="#0F766E", outline="#5EEAD4", width=2)
    d.text((608, 243), "data/output/staging/normalized_output_YYYYMMDD_HHMMSS.xlsx", font=font, fill="#FFFFFF")
    d.rectangle((580, 322, 1400, 378), fill="#F59E0B", outline="#FCD34D", width=2)
    d.text((608, 339), "logs/kpi_automation_YYYYMMDD_HHMMSS.log", font=font, fill="#111827")
    d.text((608, 420), "Mở workbook final để kiểm tra sheet kết quả và approval.", font=font, fill="#E2E8F0")
    d.text((608, 468), "Nếu file final chưa có, xem file staging trước.", font=font, fill="#E2E8F0")
    d.text((36, 620), "Hình minh họa: vị trí file output cần nhận sau khi automation hoàn tất.", font=font_small, fill="#475569")
    return save_image(img, "output_diagram.png")


def open_workbook(path: Path):
    return load_workbook(path, data_only=True)


def try_open_workbook(path: Path):
    if not path.exists():
        return None
    return load_workbook(path, data_only=True)


def sheet_values(ws, max_rows: int, max_cols: int) -> list[list[str]]:
    data: list[list[str]] = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        row = []
        for c in range(1, min(ws.max_column, max_cols) + 1):
            row.append(safe_text(ws.cell(r, c).value))
        data.append(row)
    return data


def preview_sheet_image(
    wb,
    sheet_name: str,
    out_name: str,
    max_rows: int = 8,
    max_cols: int = 6,
    note: str | None = None,
    highlight_header_contains: tuple[str, ...] = (),
    card_height: int = 900,
) -> Path:
    if sheet_name not in wb.sheetnames:
        return render_placeholder_card(
            title=sheet_name,
            message=f"Sheet '{sheet_name}' khong co trong workbook nay.",
            out_name=out_name,
            accent="#B91C1C",
            canvas_height=card_height,
        )
    ws = wb[sheet_name]
    data = sheet_values(ws, max_rows, max_cols)
    title = f"{sheet_name}"
    return render_table_card(
        title=title,
        subtitle=note or f"Preview từ file: {ws.parent.properties.title if hasattr(ws.parent, 'properties') else ''}",
        data=data,
        out_name=out_name,
        highlight_header_contains=highlight_header_contains,
        canvas_height=card_height,
    )


def render_placeholder_card(
    title: str,
    message: str,
    out_name: str,
    accent: str = "#1D4ED8",
    canvas_height: int = 900,
) -> Path:
    img, d = create_canvas(1500, canvas_height, "#FFFFFF")
    font_title = pil_font(32, bold=True)
    font_msg = pil_font(22)
    d.rounded_rectangle((24, 24, 1476, canvas_height - 24), radius=24, fill="#F8FAFC", outline=accent, width=4)
    d.rectangle((24, 24, 1476, 120), fill=accent)
    d.text((56, 48), title, font=font_title, fill="#FFFFFF")
    d.text((56, 200), message, font=font_msg, fill="#111827")
    d.text((56, 260), "Tài liệu vẫn tạo bình thường, nhưng sheet này không có trong workbook đang dùng.", font=font_msg, fill="#475569")
    return save_image(img, out_name)


def render_table_card(
    title: str,
    subtitle: str,
    data: list[list[str]],
    out_name: str,
    highlight_header_contains: tuple[str, ...] = (),
    canvas_height: int = 900,
) -> Path:
    img, d = create_canvas(1500, canvas_height, "#F8FAFC")
    font_title = pil_font(28, bold=True)
    font_sub = pil_font(18)
    font_head = pil_font(16, bold=True)
    font_cell = pil_font(15)

    d.rounded_rectangle((24, 24, 1476, canvas_height - 24), radius=24, fill="#FFFFFF", outline="#CBD5E1", width=2)
    d.text((56, 52), title, font=font_title, fill="#0F172A")
    d.text((56, 92), subtitle, font=font_sub, fill="#475569")

    if not data:
        d.text((56, 150), "Không có dữ liệu để hiển thị.", font=font_cell, fill="#B91C1C")
        return save_image(img, out_name)

    left = 56
    top = 150
    table_width = 1380
    table_height = canvas_height - 260
    cols = max(len(r) for r in data)
    normalized = [row + [""] * (cols - len(row)) for row in data]

    char_widths = [0] * cols
    for c in range(cols):
        longest = max(len(row[c]) for row in normalized)
        char_widths[c] = max(90, min(280, 18 * min(longest, 22) + 20))
    total = sum(char_widths)
    scale = min(1.0, table_width / total) if total else 1.0
    col_widths = [max(70, int(w * scale)) for w in char_widths]
    if sum(col_widths) > table_width:
        col_widths[-1] -= sum(col_widths) - table_width

    wrapped_rows: list[list[list[str]]] = []
    row_heights: list[int] = []
    for r_idx, row in enumerate(normalized):
        wrapped_row = []
        max_height = 0
        for c_idx, cell in enumerate(row):
            max_text_width = max(50, col_widths[c_idx] - 16)
            font = font_head if r_idx == 0 else font_cell
            lines = wrap_text(d, cell, font, max_text_width)
            wrapped_row.append(lines)
            height = len(lines) * (22 if r_idx == 0 else 20) + 12
            max_height = max(max_height, height)
        wrapped_rows.append(wrapped_row)
        row_heights.append(max_height)

    y = top
    for r_idx, row in enumerate(wrapped_rows):
        x = left
        row_h = row_heights[r_idx]
        header_fill = "#1D4ED8"
        body_fill = "#F8FAFC" if r_idx % 2 == 1 else "#FFFFFF"
        fill = header_fill if r_idx == 0 else body_fill
        for c_idx, lines in enumerate(row):
            cell_w = col_widths[c_idx]
            edge = "#CBD5E1"
            d.rectangle((x, y, x + cell_w, y + row_h), fill=fill, outline=edge, width=2)
            txt_color = "#FFFFFF" if r_idx == 0 else "#0F172A"
            if r_idx == 0 and highlight_header_contains:
                header_text = " ".join(lines).lower()
                if any(term.lower() in header_text for term in highlight_header_contains):
                    d.rectangle((x, y, x + cell_w, y + row_h), fill="#EA580C", outline=edge, width=2)
                    txt_color = "#FFFFFF"
            cur_y = y + 8
            for line in lines:
                d.text((x + 8, cur_y), line, font=font_head if r_idx == 0 else font_cell, fill=txt_color)
                cur_y += 20 if r_idx == 0 else 18
            x += cell_w
        y += row_h
        if y > top + table_height:
            break

    footer = "Ảnh được tạo trực tiếp từ workbook output mới nhất."
    d.text((56, canvas_height - 42), footer, font=font_sub, fill="#64748B")
    return save_image(img, out_name)


def stack_images(paths: list[Path], out_name: str, spacing: int = 24) -> Path:
    opened = [Image.open(p).convert("RGB") for p in paths]
    target_width = 1500
    scaled: list[Image.Image] = []
    total_height = spacing
    for img in opened:
        ratio = target_width / img.width
        resized = img.resize((target_width, int(img.height * ratio)), Image.LANCZOS)
        scaled.append(resized)
        total_height += resized.height + spacing
    canvas = Image.new("RGB", (target_width, total_height), "#F8FAFC")
    y = spacing
    for img in scaled:
        canvas.paste(img, (0, y))
        y += img.height + spacing
    return save_image(canvas, out_name)


def image_flowable(path: Path, width: float) -> RLImage:
    img = RLImage(str(path))
    ratio = width / img.imageWidth
    img.drawWidth = width
    img.drawHeight = img.imageHeight * ratio
    return img


def paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def rich_paragraph(html: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html, style)


def code_block(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def add_bullets(lines: list[str], style: ParagraphStyle) -> list[Paragraph]:
    return [paragraph(f"• {line}", style) for line in lines]


def make_title_table(left: str, right: str, styles: dict[str, ParagraphStyle]) -> Table:
    table = Table([[paragraph(left, styles["title"]), paragraph(right, styles["body"])]], colWidths=[110 * mm, 70 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FEF3C7")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#F59E0B")),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def build_story(styles: dict[str, ParagraphStyle], assets: dict[str, Path], latest_final: Path | None) -> list:
    story: list = []

    story.append(Spacer(1, 22 * mm))
    story.append(paragraph("HƯỚNG DẪN CHI TIẾT CHẠY AUTOMATION KPI", styles["cover_title"]))
    story.append(Spacer(1, 4 * mm))
    story.append(paragraph("Dành cho người dùng Windows, từ cài VS Code đến approval và kiểm tra sheet IT / Media", styles["cover_subtitle"]))
    story.append(Spacer(1, 8 * mm))
    story.append(make_title_table("Mục tiêu", "Tài liệu này mô tả từng bước để tải, chạy và kiểm tra file output mới nhất của automation.", styles))
    story.append(Spacer(1, 10 * mm))
    quick = Table(
        [
            ["Bạn sẽ dùng", "Mục đích"],
            ["VS Code", "Mở project và chạy lệnh"],
            ["PowerShell / Terminal", "Nhập lệnh cài đặt và chạy automation"],
            ["Workbook output", "Kiểm tra approval và dữ liệu IT / Media"],
            ["OneDrive", "Nhận file kết quả sau khi hoàn tất"],
        ],
        colWidths=[55 * mm, 115 * mm],
    )
    quick.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(quick)
    story.append(PageBreak())

    story.append(paragraph("0. Mở chương trình và mở mã nguồn", styles["h1"]))
    story.append(paragraph("Người nhận sẽ tải source từ GitHub, giải nén vào máy, rồi mở bằng VS Code.", styles["body"]))
    story.append(rich_paragraph(f'Repo GitHub: {github_repo_html()}', styles["link"]))
    story.append(rich_paragraph(f'Link tải ZIP trực tiếp: <link href="{GITHUB_ZIP_URL}">{GITHUB_ZIP_URL}</link>', styles["link"]))
    code_folder = WORK_DIR
    folder_link = file_uri(code_folder)
    story.append(
        rich_paragraph(
            f'<link href="{folder_link}">Bấm để mở thư mục code trên máy hiện tại</link>',
            styles["link"],
        )
    )
    story.append(
        Table(
            [
                ["File chính", "Mục đích"],
                [
                    rich_paragraph(f'<link href="{file_uri(WORK_DIR / "automate_kpi.py")}">automate_kpi.py</link>', styles["link"]),
                    "Chạy toàn bộ automation KPI",
                ],
                [
                    rich_paragraph(f'<link href="{file_uri(WORK_DIR / "run_kpi_automation.ps1")}">run_kpi_automation.ps1</link>', styles["link"]),
                    "Chạy bằng PowerShell trên Windows",
                ],
                [
                    rich_paragraph(f'<link href="{file_uri(WORK_DIR / "generate_huong_dan_pdf.py")}">generate_huong_dan_pdf.py</link>', styles["link"]),
                    "Tạo lại file hướng dẫn PDF",
                ],
            ],
            colWidths=[55 * mm, 115 * mm],
        )
    )
    story[-1].setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(Spacer(1, 4 * mm))
    story.append(paragraph("Nếu người dùng mở từ máy khác, bạn nên thay link trên bằng link OneDrive, SharePoint hoặc GitHub thực tế.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("1. Chuẩn bị trước khi bắt đầu", styles["h1"]))
    story.extend(
        add_bullets(
            [
                "Kiểm tra máy đang chạy Windows 10 hoặc Windows 11.",
                "Đảm bảo đã có Python 3 và Visual Studio Code.",
                "Có mạng Internet để tải source từ GitHub và dữ liệu từ Google Sheets.",
                "Nếu OneDrive đang đồng bộ, nên chờ đồng bộ xong trước khi mở file output.",
            ],
            styles["body"],
        )
    )
    story.append(paragraph("Nếu máy chưa có Python, cài Python 3 từ python.org và bật tùy chọn Add to PATH khi cài.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("2. Tải source từ GitHub và giải nén", styles["h1"]))
    story.append(paragraph("Bước này dành cho người dùng mở project trên máy khác. Dùng link GitHub hoặc file ZIP để tải source về.", styles["body"]))
    story.append(
        Table(
            [
                ["Cách tải", "Thao tác"],
                ["Tải từ GitHub", "Mở repo, bấm Code, chọn Download ZIP"],
                ["Tải ZIP trực tiếp", "Mở link ZIP và lưu file về máy"],
                ["Giải nén", "Chuột phải file ZIP -> Extract All..."],
            ],
            colWidths=[48 * mm, 122 * mm],
        )
    )
    story[-1].setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Sau khi tải ZIP, giải nén vào một thư mục dễ nhớ trên máy.",
                "Ví dụ: `C:\\Users\\<ten_user>\\Downloads\\kpi-google-sheets-onedrive-automation-main`.",
                "Không mở file `.zip` trực tiếp để chạy code.",
            ],
            styles["body"],
        )
    )
    story.append(PageBreak())

    story.append(paragraph("3. Tải và cài Visual Studio Code", styles["h1"]))
    story.append(paragraph("Bước này chỉ cần làm một lần. Nếu máy đã có VS Code, có thể bỏ qua phần tải.", styles["body"]))
    story.append(image_flowable(assets["vscode"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Mở trang tải VS Code.",
                "Bấm nút Download.",
                "Chạy file cài đặt vừa tải về.",
                "Chọn cấu hình mặc định nếu không có yêu cầu đặc biệt.",
                "Mở VS Code từ Start Menu sau khi cài xong.",
            ],
            styles["body"],
        )
    )
    story.append(paragraph("Mẹo: nếu Windows hỏi quyền cài đặt, chọn Yes hoặc Allow.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("4. Mở project và mở terminal", styles["h1"]))
    story.append(paragraph("Mở đúng thư mục gốc của dự án để VS Code nhìn thấy file automate_kpi.py và requirements.txt.", styles["body"]))
    story.append(image_flowable(assets["terminal"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Chọn File → Open Folder và trỏ đến thư mục project.",
                "Chọn Terminal → New Terminal.",
                "Dán lệnh cài thư viện hoặc lệnh chạy automation.",
                "Nếu `python` không chạy, dùng `py -3` trên Windows.",
            ],
            styles["body"],
        )
    )
    story.append(code_block("py -3 -m pip install -r requirements.txt", styles["code"]))
    story.append(Spacer(1, 2 * mm))
    story.append(code_block("py -3 automate_kpi.py --download --open", styles["code"]))
    story.append(PageBreak())

    story.append(paragraph("5. Cài thư viện và chạy automation", styles["h1"]))
    story.extend(
        add_bullets(
            [
                "Lệnh cài thư viện chỉ cần chạy một lần hoặc khi requirements.txt thay đổi.",
                "Lệnh `--download` sẽ tải lại dữ liệu nguồn từ Google Sheets trước khi xử lý.",
                "Lệnh `--open` sẽ mở file kết quả sau khi chạy xong.",
                "Nếu đã có sẵn file raw trong data/input/raw thì có thể chạy không cần `--download`.",
            ],
            styles["body"],
        )
    )
    run_table = Table(
        [
            ["Cách chạy", "Lệnh"],
            ["Tải dữ liệu mới nhất rồi chạy", "py -3 automate_kpi.py --download --open"],
            ["Chạy từ file raw đã có sẵn", "py -3 automate_kpi.py"],
            ["Chạy bằng file PowerShell", ".\\run_kpi_automation.ps1"],
        ],
        colWidths=[58 * mm, 112 * mm],
    )
    run_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(run_table)
    story.append(Spacer(1, 3 * mm))
    story.append(paragraph("Sau khi chạy xong, terminal sẽ in ra đường dẫn file output và số dòng đã ghi.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("6. Nhận file đầu ra", styles["h1"]))
    story.append(image_flowable(assets["output"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "File final thường nằm trong data/output/final.",
                "Tên file có dạng von_hoa_YYYYMMDD_HHMMSS.xlsx.",
                "File staging nằm trong data/output/staging để kiểm tra trung gian.",
                "Nếu file final chưa có, mở log để xem lý do và kiểm tra lại template.",
            ],
            styles["body"],
        )
    )
    if latest_final:
        story.append(paragraph(f"Workbook cuối cùng được dùng để tạo ảnh minh họa: {latest_final}", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("8. Bước approval cho IT và Media", styles["h1"]))
    story.append(paragraph("Quy trình approval đã được ghi ngay trong sheet Huong_dan_Approval của workbook output mới nhất. Sheet này cũng chứa hướng dẫn cho SX, Payroll và các bước duyệt kết quả.", styles["body"]))
    story.append(image_flowable(assets["approval_guide"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Mở workbook output mới nhất trong data/output/final.",
                "Mở sheet Huong_dan_Approval để đọc đúng quy trình và lệnh cần chạy, bao gồm cả SX.",
                "Nếu cần sửa dữ liệu, chỉ đánh YES ở cột Apply? cho dòng thật sự cần xử lý.",
                "Lưu workbook approval rồi chạy lại bằng tham số --approval-file. Nếu approval liên quan đến SX, bắt buộc ghi rõ --sx-year và --sx-month của kỳ đang xử lý.",
            ],
            styles["body"],
        )
    )
    story.append(code_block('py -3 automate_kpi.py --sx-year 2026 --sx-month 4 --approval-file "duong_dan_toi_file_approval.xlsx"', styles["code"]))
    story.append(paragraph("Lưu ý: đổi --sx-month 4 thành đúng tháng SX đang approve. Không nên bỏ --sx-month khi rerun approval SX vì chương trình có thể dùng tháng mặc định.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("9. Kiểm tra dữ liệu IT", styles["h1"]))
    story.append(paragraph("Mục tiêu là so sánh dữ liệu nguồn, dữ liệu đã xử lý và kết quả approval của IT.", styles["body"]))
    story.append(image_flowable(assets["check_it"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Xem Check_IT_CPNS để so sánh Input Timesheet IT với Processed Chi phí nhân sự IT.",
                "Nếu cột Diff khác 0, mở lại dòng tương ứng để kiểm tra dữ liệu nguồn hoặc công thức.",
                "Kiểm tra IT_New_Project_Master khi cần thêm dự án mới.",
                "Sau approval, xem IT_Approval_Result, Project_Master_Approval_Result và Downstream_Approval_Result.",
            ],
            styles["body"],
        )
    )
    story.append(Spacer(1, 2 * mm))
    story.append(
        Table(
            [
                ["Sheet", "Vai trò"],
                ["Timesheet IT", "Dữ liệu đầu vào IT để kiểm tra và tổng hợp"],
                ["Chi phí nhân sự IT", "Sheet xử lý chi phí nhân sự IT"],
                ["Check_IT_CPNS", "Bảng so sánh input với output và nơi duyệt sửa"],
                ["IT_Approval_Result", "Kết quả applied / skipped / failed của IT"],
                ["IT_New_Project_Master", "Duyệt thêm project master mới"],
                ["Check_IT_Downstream", "Duyệt catalog và capitalization downstream"],
            ],
            colWidths=[52 * mm, 118 * mm],
        )
    )
    story[-1].setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(PageBreak())

    story.append(paragraph("10. Kiểm tra dữ liệu Media", styles["h1"]))
    story.append(paragraph("Phần Media cần xem cả sheet nguồn và bảng so sánh để chắc chắn dữ liệu không lệch trước khi duyệt.", styles["body"]))
    story.append(image_flowable(assets["check_media"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.append(image_flowable(assets["sources"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "Xem Check_Media_Timesheet để so sánh Input Data media với Processed Timesheet Media.",
                "Mở Data media ACCA+CFA+CMA để kiểm tra các cột BU, Tên Dự Án, Nhiệm vụ, Số giờ hoàn thành công việc, Tháng, Người làm task, MNV.",
                "Kiểm tra Timesheet Media nếu cần đối chiếu theo nhân viên hoặc dự án.",
                "Nếu cần sửa, đánh YES ở các dòng được nêu trong Huong_dan_Approval và chạy lại --approval-file.",
            ],
            styles["body"],
        )
    )
    story.append(PageBreak())

    story.append(paragraph("9. Luồng SX: kiểm tra nguồn, checkpoint và approval", styles["h1"]))
    story.append(paragraph("SX có 2 checkpoint riêng. Một cái dùng để rà dữ liệu nguồn SX trước khi duyệt, một cái dùng để kiểm tra các điểm rơi ở downstream như SX_Allocation_Build, Timesheet SX, 4.1 Chi phí nhân sự SX và 3.Vốn hóa.", styles["body"]))
    sx_flow = Table(
        [
            [paragraph("Bước", styles["table"]), paragraph("Lệnh / thao tác", styles["table"]), paragraph("Mục đích", styles["table"])],
            [paragraph("1", styles["table"]), paragraph("py -3 merge_SX.py --check-sources", styles["table"]), paragraph("Kiểm tra file nguồn SX trước khi sinh dữ liệu staging.", styles["table"])],
            [paragraph("2", styles["table"]), paragraph("py -3 merge_SX.py --year 2026 --month 4", styles["table"]), paragraph("Sinh file staging SX cho tháng / năm cần chạy. Đổi year và month theo kỳ thực tế.", styles["table"])],
            [paragraph("3", styles["table"]), paragraph("py -3 automate_kpi.py --sx-year 2026 --sx-month 4", styles["table"]), paragraph("Chạy automation để tạo checkpoint data SX, SX_Allocation_Build và Check_SX_Downstream.", styles["table"])],
            [paragraph("4", styles["table"]), paragraph("Mở Huong_dan_Approval và đặt YES ở Apply?", styles["table"]), paragraph("Chỉ đánh YES cho các dòng thực sự muốn duyệt.", styles["table"])],
            [paragraph("5", styles["table"]), paragraph('py -3 automate_kpi.py --sx-year 2026 --sx-month 4 --approval-file "duong_dan_toi_file_approval.xlsx"', styles["table"]), paragraph("Áp dụng các dòng đã duyệt và tạo workbook result mới. Khi approve SX, luôn đổi --sx-month theo đúng tháng đang xử lý.", styles["table"])],
        ],
        colWidths=[16 * mm, 66 * mm, 88 * mm],
    )
    sx_flow.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 9.4),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(sx_flow)
    story.append(Spacer(1, 4 * mm))
    story.append(image_flowable(assets["sx_checkpoint"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.append(image_flowable(assets["sx_downstream"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "checkpoint data SX: nếu Details báo missing MNV, cập nhật sheet Mã nhân viên rồi chạy lại approval.",
                "Check_SX_Downstream: sửa các dòng được nêu ở Recommended action cho SX_Allocation_Build, Timesheet SX, 4.1 Chi phí nhân sự SX và 3.Vốn hóa.",
                "Sau khi sửa xong, đánh YES ở cột Apply? cho những dòng bạn đồng ý duyệt.",
                "Workbook đầu ra sau khi chạy lại sẽ có SX_Approval_Result và SX_Downstream_Approval_Result để xem trạng thái applied / skipped / failed.",
                "Ví dụ approve SX tháng 5: py -3 automate_kpi.py --sx-year 2026 --sx-month 5 --approval-file \"duong_dan_toi_file_approval.xlsx\".",
            ],
            styles["body"],
        )
    )
    story.append(image_flowable(assets["sx_results"], 175 * mm))
    story.append(PageBreak())

    story.append(paragraph("10. Kiểm tra kết quả sau approval", styles["h1"]))
    story.append(paragraph("Sau khi chạy lại bằng --approval-file, hãy mở các sheet kết quả để xem status.", styles["body"]))
    story.append(image_flowable(assets["results"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                "IT_Approval_Result cho các thao tác sửa trong IT.",
                "Project_Master_Approval_Result cho việc thêm project master mới.",
                "Downstream_Approval_Result cho các thao tác catalog và capitalization.",
                "Media_Approval_Result cho các thao tác sửa Media.",
                "SX_Approval_Result cho checkpoint data SX.",
                "SX_Downstream_Approval_Result cho các thao tác downstream của SX.",
            ],
            styles["body"],
        )
    )
    story.append(Paragraph("Nếu trạng thái là applied thì thay đổi đã được ghi vào workbook kết quả mới.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("11. Nhập liệu từ bảng lương và kiểm tra Check_Payroll", styles["h1"]))
    story.append(paragraph("Sau khi cập nhật file bảng lương trên OneDrive, chạy lại automation để đồng bộ dữ liệu lương vào workbook output.", styles["body"]))
    story.append(image_flowable(assets["payroll"], 175 * mm))
    story.append(Spacer(1, 3 * mm))
    story.extend(
        add_bullets(
            [
                f"File bảng lương nguồn đang trỏ tới: {PAYROLL_SOURCE_PATH}",
                "Nếu file nguồn đổi vị trí, hãy cập nhật `PAYROLL_SOURCE_PATH` trong code trước khi chạy lại.",
                "Chạy lại bằng `py -3 automate_kpi.py --download --open` nếu cần đồng bộ cả nguồn Google Sheets và bảng lương.",
                "Nếu chỉ muốn xử lý từ dữ liệu raw đã có, chạy `py -3 automate_kpi.py --open`.",
                "Mở sheet Check_Payroll để xem Source exists, Source sheet exists, Output sheet exists, Rows copied và Status.",
            ],
            styles["body"],
        )
    )
    story.append(
        Table(
            [
                ["Cột trong Check_Payroll", "Ý nghĩa"],
                ["Source path", "Đường dẫn file bảng lương gốc"],
                ["Source modified", "Thời điểm file nguồn được sửa gần nhất"],
                ["Rows copied", "Số dòng đã đồng bộ sang workbook output"],
                ["Duplicate Month+MNV", "Số dòng trùng tháng + MNV"],
                ["Status", "OK hoặc cảnh báo / lỗi"],
            ],
            colWidths=[58 * mm, 112 * mm],
        )
    )
    story[-1].setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(Spacer(1, 3 * mm))
    story.append(paragraph("Nếu Status không phải OK, kiểm tra lại file nguồn bảng lương, tên sheet, hoặc số dòng trùng tháng + MNV.", styles["note"]))
    story.append(PageBreak())

    story.append(paragraph("12. Luồng GitHub cho người duy trì code và người dùng cuối", styles["h1"]))
    github_flow = Table(
        [
            [paragraph("Nhóm", styles["table"]), paragraph("Cách làm", styles["table"])],
            [paragraph("Người duy trì code", styles["table"]), paragraph("git clone repo -> sửa code -> git add . -> git commit -> git push lên GitHub.", styles["table"])],
            [paragraph("Người dùng cuối", styles["table"]), paragraph("Mở repo GitHub -> tải ZIP -> giải nén -> mở bằng VS Code -> chạy automation.", styles["table"])],
        ],
        colWidths=[42 * mm, 118 * mm],
    )
    github_flow.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(github_flow)
    story.append(Spacer(1, 4 * mm))
    story.extend(
        add_bullets(
            [
                "GitHub là nguồn phát hành source chính; file ZIP chỉ là cách tải nhanh cho người dùng cuối.",
                "Nếu nội dung code thay đổi, người duy trì cần cập nhật GitHub trước rồi mới phát PDF hoặc gửi cho người dùng.",
                "Người dùng cuối không cần dùng git, chỉ cần link repo và file ZIP.",
            ],
            styles["body"],
        )
    )
    story.append(PageBreak())

    story.append(paragraph("13. Điểm yếu cố hữu còn tồn tại", styles["h1"]))
    weak_table = Table(
        [
            [paragraph("Điểm yếu", styles["table"]), paragraph("Tác động thực tế", styles["table"])],
            [paragraph("Đường dẫn bảng lương đang hard-code", styles["table"]), paragraph("Phải sửa `PAYROLL_SOURCE_PATH` nếu file OneDrive đổi vị trí hoặc đổi tên.", styles["table"])],
            [paragraph("Phụ thuộc tên sheet cố định", styles["table"]), paragraph("Nếu sheet bị đổi tên như `Timesheet IT`, `Timesheet Media`, `Check_Payroll` thì script có thể không chạy đúng.", styles["table"])],
            [paragraph("Approval vẫn cần thao tác thủ công", styles["table"]), paragraph("Người dùng phải mở workbook, đánh YES và chạy lại `--approval-file`. Với SX, lệnh approval phải có cả `--sx-year` và `--sx-month` đúng kỳ.", styles["table"])],
            [paragraph("Phụ thuộc workbook output mới nhất", styles["table"]), paragraph("Nếu thư mục final có nhiều file, người dùng cần tự chọn đúng file gần nhất để tránh nhầm version.", styles["table"])],
            [paragraph("Công thức và sheet nguồn phụ thuộc cấu trúc Excel hiện tại", styles["table"]), paragraph("Nếu workbook nguồn thay layout, mapping và công thức XLOOKUP có thể sai hoặc báo lỗi.", styles["table"])],
        ],
        colWidths=[60 * mm, 110 * mm],
    )
    weak_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C2D12")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF7ED")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(weak_table)
    story.append(Spacer(1, 4 * mm))
    story.extend(
        add_bullets(
            [
                "Hiện tại script tối ưu cho môi trường Windows nội bộ, không phải luồng cross-platform.",
                "Nếu dữ liệu nguồn thay đổi mạnh về layout, cần cập nhật code và sinh lại tài liệu hướng dẫn.",
                "Bản PDF này mô tả trạng thái hiện tại của automation, không phải cam kết rằng mọi workbook biến thể đều chạy được ngay.",
            ],
            styles["body"],
        )
    )
    story.append(PageBreak())

    story.append(paragraph("14. Lỗi thường gặp", styles["h1"]))
    trouble = Table(
        [
            ["Vấn đề", "Cách xử lý"],
            ["Không chạy được python", "Dùng py -3 hoặc cài lại Python và bật Add to PATH"],
            ["Không thấy file output", "Kiểm tra log, chạy lại với --download, xem template có bị khóa không"],
            ["Google Sheet không tải được", "Kiểm tra link trong config/sources.json và quyền chia sẻ"],
            ["File Excel bị khóa", "Đóng file đang mở rồi chạy lại"],
            ["PDF lỗi font", "Chạy lại script mới, script đã dùng font Unicode và ảnh PNG tự tạo"],
        ],
        colWidths=[54 * mm, 116 * mm],
    )
    trouble.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C2D12")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("FONTNAME", (0, 0), (-1, -1), styles["table"].fontName),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF7ED")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(trouble)
    return story


def build_pdf(output_path: Path) -> Path:
    latest_final = latest_xlsx(WORK_DIR / "data" / "output" / "final")
    latest_staging = latest_xlsx(WORK_DIR / "data" / "output" / "staging")
    approval_workbook = latest_final or latest_staging
    if approval_workbook is None:
        raise FileNotFoundError("Khong tim thay workbook output .xlsx de tao anh minh hoa.")

    wb = open_workbook(approval_workbook)
    payroll_wb = try_open_workbook(PAYROLL_SOURCE_PATH)

    assets = {
        "vscode": build_vscode_diagram(),
        "terminal": build_terminal_diagram(),
        "output": build_output_diagram(),
        "approval_guide": preview_sheet_image(
            wb,
            "Huong_dan_Approval",
            "approval_guide.png",
            max_rows=12,
            max_cols=5,
            note=f"Trich tu workbook: {approval_workbook.name}",
        ),
        "check_it": stack_images(
            [
                preview_sheet_image(
                    wb,
                    "Check_IT_CPNS",
                    "check_it_cpns.png",
                    max_rows=8,
                    max_cols=5,
                    note=f"Workbook: {approval_workbook.name}",
                    highlight_header_contains=("diff", "apply"),
                ),
                preview_sheet_image(
                    wb,
                    "IT_Approval_Result",
                    "it_approval_result.png",
                    max_rows=6,
                    max_cols=6,
                    note="Kết quả IT approval sau khi chạy.",
                ),
            ],
            "check_it_composite.png",
        ),
        "check_media": stack_images(
            [
                preview_sheet_image(
                    wb,
                    "Check_Media_Timesheet",
                    "check_media_timesheet.png",
                    max_rows=8,
                    max_cols=5,
                    note=f"Workbook: {approval_workbook.name}",
                    highlight_header_contains=("diff", "apply"),
                ),
                preview_sheet_image(
                    wb,
                    "Data media ACCA+CFA+CMA",
                    "data_media_sheet.png",
                    max_rows=8,
                    max_cols=8,
                    note="Sheet nguồn Media để đối chiếu cột BU / project / task / hours / month / employee / MNV.",
                ),
            ],
            "check_media_composite.png",
        ),
        "sources": stack_images(
            [
                preview_sheet_image(
                    wb,
                    "Timesheet IT",
                    "timesheet_it.png",
                    max_rows=7,
                    max_cols=8,
                    note="Sheet nguồn IT để kiểm tra dữ liệu trước khi approval.",
                ),
                preview_sheet_image(
                    wb,
                    "Timesheet Media",
                    "timesheet_media.png",
                    max_rows=7,
                    max_cols=8,
                    note="Sheet nguồn Media để kiểm tra dữ liệu trước khi approval.",
                ),
            ],
            "sources_composite.png",
        ),
        "payroll": preview_sheet_image(
            wb,
            "Check_Payroll",
            "check_payroll.png",
            max_rows=6,
            max_cols=11,
            note="Checkpoint sau khi sync bang luong tu OneDrive.",
            card_height=650,
        ),
        "results": stack_images(
            [
                preview_sheet_image(
                    wb,
                    "Downstream_Approval_Result",
                    "downstream_approval_result.png",
                    max_rows=5,
                    max_cols=6,
                    note="Kết quả downstream approval.",
                    card_height=650,
                ),
                preview_sheet_image(
                    wb,
                    "Project_Master_Approval_Result",
                    "project_master_approval_result.png",
                    max_rows=5,
                    max_cols=6,
                    note="Kết quả project master approval.",
                    card_height=650,
                ),
                preview_sheet_image(
                    wb,
                    "Media_Approval_Result",
                    "media_approval_result.png",
                    max_rows=5,
                    max_cols=6,
                    note="Kết quả media approval.",
                    card_height=650,
                ),
            ],
            "results_composite.png",
        ),
        "sx_checkpoint": preview_sheet_image(
            wb,
            "checkpoint data SX",
            "sx_checkpoint.png",
            max_rows=8,
            max_cols=8,
            note=f"Checkpoint SX từ workbook: {approval_workbook.name}",
            card_height=650,
        ),
        "sx_downstream": preview_sheet_image(
            wb,
            "Check_SX_Downstream",
            "sx_downstream.png",
            max_rows=8,
            max_cols=8,
            note=f"Checkpoint downstream SX từ workbook: {approval_workbook.name}",
            card_height=650,
        ),
        "sx_results": stack_images(
            [
                preview_sheet_image(
                    wb,
                    "SX_Approval_Result",
                    "sx_approval_result.png",
                    max_rows=5,
                    max_cols=6,
                    note="Kết quả SX approval sau khi chạy.",
                    card_height=650,
                ),
                preview_sheet_image(
                    wb,
                    "SX_Downstream_Approval_Result",
                    "sx_downstream_approval_result.png",
                    max_rows=5,
                    max_cols=6,
                    note="Kết quả SX downstream approval sau khi chạy.",
                    card_height=650,
                ),
            ],
            "sx_results_composite.png",
        ),
    }

    regular_font, bold_font, mono_font = register_reportlab_fonts()
    base = getSampleStyleSheet()
    styles = {
        "cover_title": ParagraphStyle(
            "cover_title",
            parent=base["Title"],
            fontName=bold_font,
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#0F172A"),
            alignment=TA_CENTER,
        ),
        "cover_subtitle": ParagraphStyle(
            "cover_subtitle",
            parent=base["Heading2"],
            fontName=regular_font,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor("#334155"),
            alignment=TA_CENTER,
        ),
        "h1": ParagraphStyle(
            "h1",
            parent=base["Heading1"],
            fontName=bold_font,
            fontSize=17,
            leading=21,
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "body",
            parent=base["BodyText"],
            fontName=regular_font,
            fontSize=10.5,
            leading=15,
            textColor=colors.HexColor("#111827"),
            spaceAfter=6,
        ),
        "note": ParagraphStyle(
            "note",
            parent=base["BodyText"],
            fontName=regular_font,
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor("#7C2D12"),
            backColor=colors.HexColor("#FFFBEB"),
            borderPadding=6,
            borderColor=colors.HexColor("#FBBF24"),
            borderWidth=0.5,
            borderRadius=4,
            spaceAfter=6,
        ),
        "link": ParagraphStyle(
            "link",
            parent=base["BodyText"],
            fontName=bold_font,
            fontSize=10.5,
            leading=15,
            textColor=colors.HexColor("#1D4ED8"),
            spaceAfter=8,
        ),
        "code": ParagraphStyle(
            "code",
            parent=base["Code"],
            fontName=mono_font,
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#0F172A"),
            backColor=colors.HexColor("#E2E8F0"),
        ),
        "table": ParagraphStyle(
            "table",
            parent=base["BodyText"],
            fontName=regular_font,
            fontSize=10,
            leading=13,
        ),
    }
    styles["title"] = styles["body"]

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title="Huong dan automation KPI",
        author="Codex",
    )

    story = build_story(styles, assets, approval_workbook)

    def on_page(canvas, _doc):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#F7F4EF"))
        canvas.rect(0, 0, A4[0], A4[1], stroke=0, fill=1)
        canvas.setFillColor(colors.HexColor("#1D4ED8"))
        canvas.rect(0, A4[1] - 14 * mm, A4[0], 14 * mm, stroke=0, fill=1)
        canvas.setFillColor(colors.white)
        canvas.setFont(bold_font, 9)
        canvas.drawString(15 * mm, A4[1] - 9 * mm, "Huong dan automation KPI")
        canvas.setFont(regular_font, 8)
        canvas.setFillColor(colors.HexColor("#4B5563"))
        canvas.drawRightString(A4[0] - 15 * mm, 8 * mm, f"Trang {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Sinh PDF huong dan chi tiet cho automation KPI")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Duong dan file PDF can tao")
    args = parser.parse_args()
    out = build_pdf(args.output)
    print(f"Da tao PDF: {out}")


if __name__ == "__main__":
    main()
