from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shutil
import sys
import zipfile
from collections import defaultdict
import urllib.request
import unicodedata
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter

import merge_SX as sx_merge

WORK_DIR = Path(__file__).parent

DIRS = {
    "raw": WORK_DIR / "data" / "input" / "raw",
    "template": WORK_DIR / "data" / "input" / "template",
    "staging": WORK_DIR / "data" / "output" / "staging",
    "final": WORK_DIR / "data" / "output" / "final",
    "logs": WORK_DIR / "logs",
}

DEFAULT_SOURCES = {
    "IT": {
        "url": "https://docs.google.com/spreadsheets/d/1xk2Bd95uXoOJAeBqkOSzV1wEo9peOF-e/edit?usp=sharing&ouid=111275165352399420864&rtpof=true&sd=true",
        "file": "IT.xlsx",
        "type": "matrix",
    },
    "MEDIA": {
        "url": "https://docs.google.com/spreadsheets/d/17Eaj28S4nRgR1Iy6yKzkB7wo9XzS68aP/edit?gid=646344289#gid=646344289",
        "file": "MEDIA.xlsx",
        "type": "task_table",
    },
}

TEMPLATE_CANDIDATES = [
    "von_hoa_template.xlsx",
    "Vốn hóa chi phí nhân sự 2026.xlsx",
    "Vốn hóa chi phí nhân sự 2026 (1).xlsx",
    "Vốn hóa chi phí nhân sự 2026 (2).xlsx",
]

SHEET_IT = "Timesheet IT"
SHEET_MEDIA = "Data media ACCA+CFA+CMA"
SHEET_TIMESHEET_MEDIA = "Timesheet Media"
SHEET_IT_COST = "Chi phí nhân sự IT"
SHEET_SALARY_FULLTIME = "Lương nhân viên full time"
SHEET_SALARY_PARTTIME = "Lương nhân viên part time"
SHEET_PROJECT_CATALOG = "1.Danh mục dự án"
SHEET_IT_CHECKING = "Checking Vốn hóa IT"
SHEET_CAPITALIZATION = "3.Vốn hóa"
SHEET_EMPLOYEE = "Mã nhân viên"
SHEET_SX_TARGET = "Data SX ACCA+CMA"
SHEET_SX_ALLOCATION = "SX_Allocation_Build"
SHEET_SX_CHECKPOINT = "checkpoint data SX"
SHEET_SX_DOWNSTREAM_CHECKPOINT = "Check_SX_Downstream"
SHEET_SX_DOWNSTREAM_RESULT = "SX_Downstream_Approval_Result"

SHEET_GROUP_COLORS = {
    "IT": "FF1D4ED8",
    "MEDIA": "FF0F766E",
    "SX": "FFEA580C",
    "COMMON": "FF64748B",
}

PAYROLL_SOURCE_PATH = Path(r"C:\Users\admin\OneDrive\BCQT 2026\Vốn hóa chi phí nhân sự 2026.xlsx")
PAYROLL_SHEETS = [SHEET_SALARY_FULLTIME, SHEET_SALARY_PARTTIME]

IT_HEADER_ROW = 13
IT_TEMPLATE_ROW = 13
MEDIA_HEADER_ROW = 5
MEDIA_TEMPLATE_ROW = 6
MEDIA_AUDIT_FIELDS = [
    "source_sheet",
    "source_row",
    "bu",
    "project_name",
    "task_name",
    "hours",
    "month",
    "employee",
    "mnv",
]
APPROVAL_YES_VALUES = {"yes", "y", "true", "1", "apply", "approved"}
SX_TEMPLATE_ROW = 5
SX_APPEND_BUFFER_ROWS = 500
SX_ALLOWED_PROGRAMS = {"ACCA", "CMA", "CFA"}
SX_PROGRAM_SOURCE_FILES = {
    "ACCA": "ACCA.xlsx",
    "CMA": "CMA.xlsx",
    "CFA": "CFA.xlsx",
}
SX_TEMPLATE_COLUMN_MAP = {
    1: 1,   # month
    2: 2,   # program
    3: 3,   # position
    4: 4,   # employee
    6: 6,   # product
    7: 8,   # product type
    8: 9,   # project
    9: 10,  # department
    10: 11, # feature
    11: 12, # deliverable
    12: 15, # unit
    13: 16, # actual
    14: 17, # KPI standard
}
SX_ALLOCATION_SHARE_START_COL = 24  # X


if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        v = re.sub(r"\s+", " ", value).strip()
        return v or None
    return value


def norm(value: Any) -> str:
    value = clean(value)
    if value is None:
        return ""
    text = str(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_num(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def setup_dirs() -> None:
    for path in DIRS.values():
        path.mkdir(parents=True, exist_ok=True)


def load_source_config() -> dict[str, dict[str, Any]]:
    cfg_path = WORK_DIR / "config" / "sources.json"
    if not cfg_path.exists():
        return DEFAULT_SOURCES.copy()

    with cfg_path.open("r", encoding="utf-8") as f:
        loaded = json.load(f)

    result: dict[str, dict[str, Any]] = {}
    for key, fallback in DEFAULT_SOURCES.items():
        item = dict(fallback)
        if key in loaded and isinstance(loaded[key], dict):
            item.update({k: v for k, v in loaded[key].items() if v is not None})
        result[key] = item
    return result


def find_template() -> Path | None:
    for name in TEMPLATE_CANDIDATES:
        p = DIRS["template"] / name
        if p.exists():
            return p

    candidates: list[Path] = []
    for folder in [DIRS["template"], WORK_DIR]:
        if folder.exists():
            candidates.extend(folder.glob("*.xlsx"))

    for p in candidates:
        low = p.name.lower()
        if "von_hoa" in low or "vốn hóa" in low or "von hoa" in low:
            return p
    return None


def extract_spreadsheet_id(url: str) -> str | None:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    return None


def extract_gid(url: str) -> str | None:
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    if "gid" in query and query["gid"]:
        return query["gid"][0]
    frag = parsed.fragment
    if frag.startswith("gid="):
        return frag.split("=", 1)[1]
    return None


def build_export_url(url: str, full_workbook: bool = True) -> str:
    sheet_id = extract_spreadsheet_id(url)
    if not sheet_id:
        return url

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    gid = extract_gid(url)
    if gid and not full_workbook:
        export_url += f"&gid={gid}"
    return export_url


def download_source(url: str, out_path: Path, full_workbook: bool = True) -> None:
    export_url = build_export_url(url, full_workbook=full_workbook)
    tmp = out_path.with_name(f".{out_path.stem}.download.xlsx")
    req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})

    try:
        with urllib.request.urlopen(req) as resp, tmp.open("wb") as f:
            shutil.copyfileobj(resp, f)

        try:
            tmp.replace(out_path)
        except PermissionError:
            shutil.copy2(tmp, out_path)
            tmp.unlink(missing_ok=True)
    except Exception as exc:
        tmp.unlink(missing_ok=True)
        if out_path.exists():
            print(f"    download skipped, using existing local file: {exc}")
            return
        raise


def download_all(sources: dict[str, dict[str, Any]]) -> None:
    for key, meta in sources.items():
        out = DIRS["raw"] / meta["file"]
        print(f"  {key} -> {out.name}")
        try:
            download_source(meta["url"], out, full_workbook=True)
        except Exception as exc:
            print(f"    FAILED: {exc}")


def ensure_row_style(ws, template_row: int, target_row: int, max_col: int | None = None) -> None:
    max_col = max_col or ws.max_column
    if target_row <= template_row:
        return

    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(template_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def row_has_any_value(ws, row: int, max_col: int | None = None) -> bool:
    max_col = max_col or ws.max_column
    # Use the internal cell map so probing empty rows does not create new cell objects.
    for c in range(1, max_col + 1):
        cell = ws._cells.get((row, c))
        if cell is not None and clean(cell.value) is not None:
            return True
    return False


def copy_row_style_and_formulas_only(ws, source_row: int, target_row: int, max_col: int | None = None) -> None:
    max_col = max_col or ws.max_column
    if target_row <= source_row:
        return

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = None


def find_sx_append_start_row(ws, template_row: int = SX_TEMPLATE_ROW, buffer_rows: int = SX_APPEND_BUFFER_ROWS) -> int:
    last_data_row = template_row
    for r in range(template_row, ws.max_row + 1):
        if row_has_any_value(ws, r, 18):
            last_data_row = r

    candidate = last_data_row + 1
    while candidate <= ws.max_row + buffer_rows + 1:
        if all(not row_has_any_value(ws, r, 18) for r in range(candidate, candidate + buffer_rows)):
            return candidate
        candidate += 1
    return last_data_row + 1


def load_sx_staging_rows(sx_output: Path) -> list[dict[str, Any]]:
    wb = load_workbook(sx_output, data_only=True)
    ws = wb.active

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        month = clean(ws.cell(r, 2).value)
        program = clean(ws.cell(r, 3).value)
        position = clean(ws.cell(r, 4).value)
        employee = clean(ws.cell(r, 5).value)
        product = clean(ws.cell(r, 6).value)
        if not any([month, program, position, employee, product, ws.cell(r, 13).value, ws.cell(r, 14).value]):
            continue
        rows.append(
            {
                "year": clean(ws.cell(r, 1).value),
                "month": month,
                "program": program,
                "position": position,
                "employee": employee,
                "product": product,
                "product_type": clean(ws.cell(r, 7).value),
                "project": clean(ws.cell(r, 8).value),
                "department": clean(ws.cell(r, 9).value),
                "feature": clean(ws.cell(r, 10).value),
                "deliverable": clean(ws.cell(r, 11).value),
                "unit": clean(ws.cell(r, 12).value),
                "actual": ws.cell(r, 13).value,
                "kpi_standard": ws.cell(r, 14).value,
                "_source_row": r,
            }
        )

    return rows


def sx_row_value(row: dict[str, Any], *labels: str) -> Any:
    wanted = {norm(label) for label in labels if label}
    for key, value in row.items():
        if norm(key) in wanted:
            return value
    return None


def normalize_sx_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "year": sx_row_value(row, "Năm", "Year"),
        "month": sx_row_value(row, "Tháng", "Month"),
        "program": sx_row_value(row, "Chương trình", "Program"),
        "position": sx_row_value(row, "Vị trí", "Position"),
        "employee": sx_row_value(row, "Tên nhân viên", "Employee"),
        "product": sx_row_value(row, "Tên sản phẩm", "Product"),
        "product_type": sx_row_value(row, "Sản phẩm mới/sản phẩm cũ", "Product type"),
        "project": sx_row_value(row, "Tên dự án", "Project"),
        "department": sx_row_value(row, "Bộ môn", "Department"),
        "feature": sx_row_value(row, "Đặc tính sản phẩm", "Feature"),
        "deliverable": sx_row_value(row, "Sản phẩm bàn giao", "Deliverable"),
        "unit": sx_row_value(row, "Đơn vị tính", "Unit"),
        "actual": sx_row_value(row, "Số lượng actual", "Actual"),
        "kpi_standard": sx_row_value(row, "KPI standard (h)/1 ĐV quy đổi", "KPI standard", "KPI standard (h)"),
        "_source_file": row.get("_source_file"),
        "_source_sheet": row.get("_source_sheet"),
        "_source_row": row.get("_source_row"),
    }


def source_file_for_program(program: Any) -> str | None:
    key = norm(program).upper()
    return SX_PROGRAM_SOURCE_FILES.get(key)


def build_sx_checkpoint_data(
    sx_rows: list[dict[str, Any]],
    employee_lookup: dict[str, str],
    include_cfa: bool,
    append_start_row: int | None = None,
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], int, int]:
    summary: list[list[Any]] = []
    details: list[list[Any]] = []
    missing_mnv_detail: list[list[Any]] = []

    kept_rows = 0
    skipped_rows = 0
    missing_employee_rows = 0
    missing_mnv_rows = 0
    missing_mnv_employees: set[str] = set()
    next_append_row = append_start_row

    def recommendation_for(issues: list[str]) -> tuple[str, str]:
        if not issues:
            return "OK", "YES"
        if "missing MNV" in issues and "missing employee" in issues:
            return "Update source employee name and then refresh 'Mã nhân viên' lookup", "NO"
        if "missing MNV" in issues:
            return "Add or fix the employee in sheet 'Mã nhân viên' or correct the source name", "NO"
        if "missing employee" in issues:
            return "Fill the employee name in the source row", "NO"
        return "Review source row", "NO"

    for row in sx_rows:
        program = clean(row.get("program"))
        program_key = norm(program).upper()
        if program_key not in SX_ALLOWED_PROGRAMS:
            skipped_rows += 1
            details.append([
                row.get("year"),
                row.get("month"),
                program,
                row.get("_source_file") or source_file_for_program(program),
                row.get("_source_sheet"),
                row.get("_source_row"),
                None,
                row.get("employee"),
                row.get("position"),
                None,
                row.get("product"),
                row.get("project"),
                "unsupported program",
                "Verify source program and keep only ACCA/CMA/CFA rows",
                "NO",
                None,
            ])
            continue
        if program_key == "CFA" and not include_cfa:
            skipped_rows += 1
            details.append([
                row.get("year"),
                row.get("month"),
                program,
                row.get("_source_file") or source_file_for_program(program),
                row.get("_source_sheet"),
                row.get("_source_row"),
                None,
                row.get("employee"),
                row.get("position"),
                None,
                row.get("product"),
                row.get("project"),
                "CFA excluded before 05/2026",
                "Keep on Data SX CFA sheet for months before 05/2026",
                "NO",
                None,
            ])
            continue

        employee = clean(row.get("employee"))
        mnv = lookup_employee_code(employee_lookup, employee)
        issues = []
        if not employee:
            issues.append("missing employee")
            missing_employee_rows += 1
        if not mnv:
            issues.append("missing MNV")
            missing_mnv_rows += 1
            if employee:
                missing_mnv_employees.add(str(employee))

        kept_rows += 1
        recommended_action, apply_flag = recommendation_for(issues)
        current_append_row = next_append_row if next_append_row is not None else None
        if next_append_row is not None:
            next_append_row += 1
        if issues:
            details.append([
                row.get("year"),
                row.get("month"),
                program,
                row.get("_source_file") or source_file_for_program(program),
                row.get("_source_sheet"),
                row.get("_source_row"),
                current_append_row,
                employee,
                row.get("position"),
                mnv,
                row.get("product"),
                row.get("project"),
                "; ".join(issues) if issues else None,
                recommended_action,
                apply_flag,
                None,
            ])
        if not mnv:
            missing_mnv_detail.append([
                row.get("year"),
                row.get("month"),
                program,
                row.get("_source_file") or source_file_for_program(program),
                row.get("_source_sheet"),
                row.get("_source_row"),
                employee,
                clean(row.get("position")),
                row.get("product"),
                row.get("project"),
                row.get("feature"),
                row.get("unit"),
                row.get("actual"),
                row.get("kpi_standard"),
                "missing MNV",
                "Add or fix the employee in sheet 'Mã nhân viên' or correct the source name",
                "NO",
                None,
            ])

    summary.extend([
        ["Rows staged", len(sx_rows)],
        ["Rows appended", kept_rows],
        ["Rows skipped", skipped_rows],
        ["Rows missing employee", missing_employee_rows],
        ["Rows missing MNV", missing_mnv_rows],
        ["Distinct employees missing MNV", len(missing_mnv_employees)],
    ])
    return summary, details, missing_mnv_detail, kept_rows, skipped_rows


def write_sx_checkpoint_sheet(
    wb,
    summary_rows: list[list[Any]],
    detail_rows: list[list[Any]],
    missing_mnv_rows: list[list[Any]],
) -> None:
    ws = reset_sheet(wb, SHEET_SX_CHECKPOINT)
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Value"],
        summary_rows,
    )
    next_row = append_table(
        ws,
        next_row,
        "Details",
        [
            "Year",
            "Month",
            "Program",
            "Source file",
            "Source sheet",
            "Source row",
            "Append row",
            "Employee",
            "Position",
            "MNV lookup",
            "Product",
            "Project",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        detail_rows,
    )
    ws.freeze_panes = "A3"
    for col, width in {
        "A": 10,
        "B": 10,
        "C": 12,
        "D": 16,
        "E": 16,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 24,
        "J": 16,
        "K": 28,
        "L": 28,
        "M": 28,
        "N": 36,
        "O": 10,
    }.items():
        ws.column_dimensions[col].width = width


def read_sx_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, SHEET_SX_CHECKPOINT)
    if ws is None:
        return []

    approvals: list[dict[str, Any]] = []

    def parse_section(section_name: str) -> None:
        header_row = find_section_header(ws, section_name)
        if header_row is None:
            return
        headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
        if "Apply?" not in headers or "Recommended action" not in headers:
            return
        for r in range(header_row + 1, ws.max_row + 1):
            marker = ws.cell(r, 1).value
            if marker in {"Summary", "Details"}:
                break
            if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
                continue
            approvals.append(
                {
                    "section": section_name,
                    "approval_row": r,
                    "year": ws.cell(r, headers.get("Year", 0)).value if headers.get("Year") else None,
                    "month": ws.cell(r, headers.get("Month", 0)).value if headers.get("Month") else None,
                    "program": ws.cell(r, headers.get("Program", 0)).value if headers.get("Program") else None,
                    "source_file": ws.cell(r, headers.get("Source file", 0)).value if headers.get("Source file") else None,
                    "source_sheet": ws.cell(r, headers.get("Source sheet", 0)).value if headers.get("Source sheet") else None,
                    "source_row": ws.cell(r, headers.get("Source row", 0)).value if headers.get("Source row") else None,
                    "append_row": ws.cell(r, headers.get("Append row", 0)).value if headers.get("Append row") else None,
                    "position": ws.cell(r, headers.get("Position", 0)).value if headers.get("Position") else None,
                    "employee": ws.cell(r, headers.get("Employee", 0)).value if headers.get("Employee") else None,
                    "mnv_lookup": ws.cell(r, headers.get("MNV lookup", 0)).value if headers.get("MNV lookup") else None,
                    "product": ws.cell(r, headers.get("Product", 0)).value if headers.get("Product") else None,
                    "project": ws.cell(r, headers.get("Project", 0)).value if headers.get("Project") else None,
                    "feature": ws.cell(r, headers.get("Feature", 0)).value if headers.get("Feature") else None,
                    "unit": ws.cell(r, headers.get("Unit", 0)).value if headers.get("Unit") else None,
                    "actual": ws.cell(r, headers.get("Actual", 0)).value if headers.get("Actual") else None,
                    "kpi_standard": ws.cell(r, headers.get("KPI standard", 0)).value if headers.get("KPI standard") else None,
                    "issue": ws.cell(r, headers.get("Issue", 0)).value if headers.get("Issue") else None,
                    "recommended_action": ws.cell(r, headers["Recommended action"]).value,
                    "apply": ws.cell(r, headers["Apply?"]).value,
                    "approval_notes": ws.cell(r, headers.get("Approval notes", 0)).value if headers.get("Approval notes") else None,
                }
            )

    parse_section("Details")
    return approvals


def write_sx_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "SX_Approval_Result")
    headers = [
        "Approval row",
        "Section",
        "Year",
        "Month",
        "Program",
        "Employee",
        "MNV lookup",
        "Issue",
        "Recommended action",
        "Apply?",
        "Resolved?",
        "Resolution note",
        "Status",
    ]
    append_table(ws, 1, "SX Approval Result", headers, results)
    for col, width in {"A": 14, "B": 16, "C": 10, "D": 10, "E": 12, "F": 24, "G": 16, "H": 28, "I": 42, "J": 10, "K": 12, "L": 40, "M": 80}.items():
        ws.column_dimensions[col].width = width


def apply_sx_approvals(wb, approval_file: Path | None, employee_lookup: dict[str, str]) -> None:
    if approval_file is None:
        return
    approvals = read_sx_approvals(approval_file)
    if not approvals:
        return

    results: list[list[Any]] = []
    for approval in approvals:
        is_yes = approval_value_is_yes(approval.get("apply"))
        mnv_lookup = clean(approval.get("mnv_lookup"))
        status = "skipped: Apply? not yes"
        resolved = "NO"
        resolution_note = None
        if is_yes:
            if "missing MNV" in str(approval.get("issue") or "").lower():
                if mnv_lookup or lookup_employee_code(employee_lookup, approval.get("employee")):
                    status = "applied: resolved by updated MNV lookup"
                    resolved = "YES"
                    resolution_note = "Missing MNV resolved in approval workbook"
                else:
                    status = "failed: still missing MNV"
                    resolution_note = "Add/update MNV in the approval workbook and rerun"
            else:
                status = "applied: retained row for regenerated SX output"
                resolved = "YES"
                resolution_note = "Row kept in regenerated SX output"

        results.append(
            [
                approval.get("approval_row"),
                approval.get("section"),
                approval.get("year"),
                approval.get("month"),
                approval.get("program"),
                approval.get("employee"),
                mnv_lookup,
                approval.get("issue"),
                approval.get("recommended_action"),
                approval.get("apply"),
                resolved,
                resolution_note,
                status,
            ]
        )

    write_sx_approval_result_sheet(wb, results)


def find_month_label_col(ws, month: int, label_row: int = 1) -> int | None:
    target = norm(f"Tháng {month}")
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(label_row, c).value) == target:
            return c
    return None


def find_row_with_exact_match(ws, start_row: int, matchers: list[tuple[int, Any]]) -> int | None:
    targets = [(col, norm(value)) for col, value in matchers if norm(value)]
    if not targets:
        return None
    for r in range(start_row, ws.max_row + 1):
        ok = True
        for col, target in targets:
            if norm(ws.cell(r, col).value) != target:
                ok = False
                break
        if ok:
            return r
    return None


def find_row_by_column_value(ws, start_row: int, col: int, value: Any) -> int | None:
    target = norm(value)
    if not target:
        return None
    for r in range(start_row, ws.max_row + 1):
        if norm(ws.cell(r, col).value) == target:
            return r
    return None


def sx_downstream_month_block_start(month: int) -> int:
    return 11 + (month - 1) * 4


def build_sx_downstream_checkpoint_data(
    wb,
    sx_rows: list[dict[str, Any]],
    year: int,
    month: int,
    employee_lookup: dict[str, str],
    capital_values_wb=None,
) -> tuple[list[list[Any]], list[list[Any]]]:
    include_cfa = (year, month) >= (2026, 5)

    filtered_rows: list[dict[str, Any]] = []
    skipped_rows = 0
    missing_mnv_rows = 0
    for row in sx_rows:
        program = norm(row.get("program")).upper()
        if program not in SX_ALLOWED_PROGRAMS:
            skipped_rows += 1
            continue
        if program == "CFA" and not include_cfa:
            skipped_rows += 1
            continue
        # Keep all valid SX rows and let downstream approval decide whether the
        # project should be added to 1.Danh mục dự án / 3.Vốn hóa.
        filtered_rows.append(row)
        if not lookup_employee_code(employee_lookup, row.get("employee")):
            missing_mnv_rows += 1

    details: list[list[Any]] = []
    summary: list[list[Any]] = []

    allocation_ws = find_sheet(wb, SHEET_SX_ALLOCATION)
    timesheet_ws = find_sheet(wb, "Timesheet SX")
    cost_ws = find_sheet(wb, "4.1 Chi phí nhân sự SX")
    capital_ws = find_sheet(wb, SHEET_CAPITALIZATION)

    if allocation_ws is None or timesheet_ws is None or cost_ws is None or capital_ws is None:
        summary.extend([
            ["Rows staged from Data SX ACCA+CMA", len(filtered_rows)],
            ["Rows missing in downstream", len(filtered_rows)],
            ["Rows with missing MNV", missing_mnv_rows],
            ["Rows excluded by filter", skipped_rows],
        ])
        details.append([
                "sheet structure",
                year,
                month,
                None,
                None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
                "missing downstream sheet(s)",
                "Restore missing downstream sheet(s) in template",
                "NO",
                None,
                None,
                None,
            "NO",
        ])
        return summary, details

    catalog_projects: dict[str, str] = {}
    capital_projects: dict[str, str] = {}
    project_codes: dict[str, str] = {}
    capital_source_wb = capital_values_wb or wb
    catalog_ws = find_sheet(capital_source_wb, SHEET_PROJECT_CATALOG)
    if catalog_ws is not None:
        for r in range(2, catalog_ws.max_row + 1):
            project_name = clean(catalog_ws.cell(r, 3).value)
            project_code = clean(catalog_ws.cell(r, 2).value)
            if project_name:
                catalog_projects[norm(project_name)] = project_name
            if project_name and project_code and norm(project_name) not in project_codes:
                project_codes[norm(project_name)] = str(project_code)
    if capital_ws is not None:
        for r in range(3, capital_ws.max_row + 1):
            project_name = clean(capital_ws.cell(r, 3).value)
            project_code = clean(capital_ws.cell(r, 2).value)
            if project_name:
                capital_projects[norm(project_name)] = project_name
            if project_name and project_code and norm(project_name) not in project_codes:
                project_codes[norm(project_name)] = str(project_code)

    eligible_rows: list[dict[str, Any]] = []
    for row in filtered_rows:
        program = clean(row.get("program"))
        project_raw = clean(row.get("project"))
        employee = clean(row.get("employee"))
        employee_code = lookup_employee_code(employee_lookup, employee) or "Khong tim thay"
        ts_row = find_row_with_exact_match(
            timesheet_ws,
            4,
            [(1, year), (2, program), (4, project_raw), (5, "SX"), (6, employee), (7, employee_code)],
        )
        capital_project = clean(timesheet_ws.cell(ts_row, 3).value) if ts_row is not None else None
        enriched = dict(row)
        enriched["_capital_project"] = capital_project
        enriched["_timesheet_row"] = ts_row
        enriched["_employee_code"] = employee_code
        enriched["_project_code"] = project_codes.get(norm(capital_project or project_raw))
        eligible_rows.append(enriched)

    pairs: dict[tuple[str, str], dict[str, Any]] = {}
    projects: dict[str, dict[str, Any]] = {}
    for row in eligible_rows:
        employee = clean(row.get("employee"))
        project = clean(row.get("_capital_project") or row.get("project"))
        pair_key = (norm(project), norm(employee))
        if pair_key not in pairs:
            pairs[pair_key] = row
        if project and norm(project) not in projects:
            projects[norm(project)] = row

    allocation_month_share_col = sx_allocation_month_share_col(month)
    timesheet_month_col = 7 + month
    cost_month_start_col = sx_downstream_month_block_start(month)
    capital_month_col = find_month_label_col(capital_ws, month, label_row=1)
    timesheet_month_header_col = find_month_label_col(timesheet_ws, month, label_row=2)
    cost_month_header_col = find_month_label_col(cost_ws, month, label_row=1)

    matched_consol = 0
    matched_timesheet = 0
    matched_cost = 0
    matched_capital = 0
    missing_consol = 0
    missing_timesheet = 0
    missing_cost = 0
    missing_capital = 0
    exact_match_mismatch = 0

    def add_issue(
        stage: str,
        row: dict[str, Any],
        issue: str,
        recommended_action: str,
        expected_target: Any = None,
        actual_target: Any = None,
        matched: str = "NO",
        append_row: int | None = None,
    ) -> None:
        details.append([
            stage,
            row.get("year"),
            row.get("month"),
            row.get("program"),
            row.get("_source_file") or source_file_for_program(row.get("program")),
            row.get("_source_sheet"),
            row.get("_source_row"),
            append_row,
            row.get("employee"),
            row.get("position"),
            lookup_employee_code(employee_lookup, row.get("employee")),
            row.get("_project_code"),
            row.get("project"),
            issue,
            recommended_action,
            "NO",
            None,
            expected_target,
            actual_target,
            matched,
        ])

    if timesheet_month_header_col is None:
        add_issue(
            "Timesheet SX",
            filtered_rows[0] if filtered_rows else {"year": year, "month": month, "program": "SX"},
            f"missing month {month} column",
            "Copy the month block from the previous month and refresh formulas",
            expected_target=f"Tháng {month}",
            actual_target=None,
        )
    if cost_month_header_col is None:
        add_issue(
            "4.1 Chi phí nhân sự SX",
            filtered_rows[0] if filtered_rows else {"year": year, "month": month, "program": "SX"},
            f"missing month {month} block",
            "Copy the month block from the previous month and refresh formulas",
            expected_target=f"Tháng {month}",
            actual_target=None,
        )
    if capital_month_col is None:
        add_issue(
            SHEET_CAPITALIZATION,
            filtered_rows[0] if filtered_rows else {"year": year, "month": month, "program": "SX"},
            f"missing month {month} block",
            "Copy the month block from the previous month and refresh formulas",
            expected_target=f"Tháng {month}",
            actual_target=None,
        )

    projects_needing_approval: set[str] = set()
    project_approval_rows = 0
    missing_catalog_projects = 0
    missing_capital_projects = 0
    for project_key, row in projects.items():
        project_name = clean(row.get("_capital_project") or row.get("project"))
        project_code = clean(row.get("_project_code")) or project_codes.get(project_key)
        catalog_row = find_row_by_clean_value(catalog_ws, 3, project_name, start_row=2) if catalog_ws is not None else None
        capital_row = find_row_by_clean_value(capital_ws, 3, project_name, start_row=3) if capital_ws is not None else None

        catalog_exists = catalog_row is not None
        capital_exists = capital_row is not None
        if not catalog_exists:
            missing_catalog_projects += 1
        if not capital_exists:
            missing_capital_projects += 1
        if catalog_exists and capital_exists:
            continue

        projects_needing_approval.add(project_key)
        sample_row = dict(row)
        sample_row["_project_code"] = project_code
        sample_row["employee"] = None
        sample_row["position"] = None
        sample_row["_employee_code"] = None

        issue_parts: list[str] = []
        recommended_action: list[str] = []
        if not catalog_exists:
            issue_parts.append("missing in 1.Danh mục dự án")
            recommended_action.append("ADD_TO_PROJECT_CATALOG")
        if not capital_exists:
            issue_parts.append("missing in 3.Vốn hóa")
            recommended_action.append("ADD_TO_CAPITALIZATION")

        add_issue(
            "1.Danh mục dự án" if not catalog_exists else SHEET_CAPITALIZATION,
            sample_row,
            "; ".join(issue_parts),
            "; ".join(recommended_action),
            expected_target=project_name,
            actual_target=project_code,
            matched="NO",
        )
        project_approval_rows += 1

    for (project_key, employee_key), row in pairs.items():
        if project_key in projects_needing_approval:
            continue
        program = clean(row.get("program"))
        project = clean(row.get("project"))
        capital_project = clean(row.get("_capital_project") or row.get("project"))
        employee = clean(row.get("employee"))
        append_row = None
        # SX_Allocation_Build
        alloc_row = find_row_with_exact_match(
            allocation_ws,
            2,
            [(1, year), (2, program), (3, project), (4, employee), (5, row.get("_employee_code") or lookup_employee_code(employee_lookup, employee) or "Khong tim thay"), (6, "SX")],
        )
        if alloc_row is None or norm(allocation_ws.cell(alloc_row, allocation_month_share_col).value) in {"", "none", "0", "0.0"}:
            missing_consol += 1
            add_issue(
                "SX_Allocation_Build",
                row,
                "missing in SX_Allocation_Build",
                "Rebuild SX_Allocation_Build from raw SX rows and verify project/employee mapping",
                expected_target=f"{project} / {employee}",
                actual_target=None if alloc_row is None else allocation_ws.cell(alloc_row, allocation_month_share_col).value,
                matched="NO" if alloc_row is None else "PARTIAL",
                append_row=append_row,
            )
        else:
            matched_consol += 1

        # Timesheet SX
        ts_row = row.get("_timesheet_row") or find_row_with_exact_match(
            timesheet_ws,
            4,
            [
                (1, year),
                (2, program),
                (4, project),
                (5, "SX"),
                (6, employee),
                (7, row.get("_employee_code") or lookup_employee_code(employee_lookup, employee) or "Khong tim thay"),
            ],
        )
        if ts_row is None or norm(timesheet_ws.cell(ts_row, timesheet_month_col).value) in {"", "none"}:
            missing_timesheet += 1
            add_issue(
                "Timesheet SX",
                row,
                "missing in Timesheet SX",
                "Check project/employee mapping in SX_Allocation_Build and refresh Timesheet SX formulas",
                expected_target=f"{project} / {employee}",
                actual_target=None if ts_row is None else timesheet_ws.cell(ts_row, timesheet_month_col).value,
                matched="NO" if ts_row is None else "PARTIAL",
                append_row=append_row,
            )
        else:
            matched_timesheet += 1

        # 4.1 Chi phí nhân sự SX
        cost_row = find_row_with_exact_match(cost_ws, 3, [(1, capital_project), (7, employee)])
        if cost_row is None or all(norm(cost_ws.cell(cost_row, c).value) in {"", "none"} for c in range(cost_month_start_col, min(cost_month_start_col + 3, cost_ws.max_column + 1))):
            missing_cost += 1
            add_issue(
                "4.1 Chi phí nhân sự SX",
                row,
                "missing in 4.1 Chi phí nhân sự SX",
                "Refresh cost formulas and verify the project/employee row exists",
                expected_target=f"{capital_project} / {employee}",
                actual_target=None if cost_row is None else cost_ws.cell(cost_row, cost_month_start_col).value,
                matched="NO" if cost_row is None else "PARTIAL",
                append_row=append_row,
            )
        else:
            matched_cost += 1

    for project_key, row in projects.items():
        if project_key in projects_needing_approval:
            continue
        project = clean(row.get("_capital_project") or row.get("project"))
        project_code = clean(row.get("_project_code")) or project_codes.get(norm(project))
        capital_row = None
        if project_code:
            capital_row = find_row_with_exact_match(capital_ws, 3, [(2, project_code)])
        if capital_row is None:
            missing_capital += 1
            add_issue(
                SHEET_CAPITALIZATION,
                row,
                "missing in 3.Vốn hóa",
                "Add or verify the project row in the capitalization sheet and catalog",
                expected_target=project,
                actual_target=None,
                matched="NO",
                append_row=None,
            )
        else:
            matched_capital += 1

    summary.extend([
        ["Rows staged from Data SX ACCA+CMA", len(filtered_rows)],
        ["Projects needing downstream approval", project_approval_rows],
        ["Projects missing in 1.Danh mục dự án", missing_catalog_projects],
        ["Projects missing in 3.Vốn hóa", missing_capital_projects],
        ["Rows matched in SX_Allocation_Build", matched_consol],
        ["Rows missing in SX_Allocation_Build", missing_consol],
        ["Rows matched in Timesheet SX", matched_timesheet],
        ["Rows missing in Timesheet SX", missing_timesheet],
        ["Rows matched in 4.1 Chi phí nhân sự SX", matched_cost],
        ["Rows missing in 4.1 Chi phí nhân sự SX", missing_cost],
        ["Rows reaching 3.Vốn hóa", matched_capital],
        ["Rows missing in 3.Vốn hóa", missing_capital],
        ["Rows with missing MNV", missing_mnv_rows],
        ["Rows excluded by filter", skipped_rows],
        ["Rows with exact-match mismatch", exact_match_mismatch],
    ])
    return summary, details


def write_sx_downstream_checkpoint_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, SHEET_SX_DOWNSTREAM_CHECKPOINT)
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Value"],
        summary_rows,
    )
    append_table(
        ws,
        next_row,
        "Details",
        [
            "Stage",
            "Year",
            "Month",
            "Program",
            "Source file",
            "Source sheet",
            "Source row",
            "Append row",
            "Employee",
            "Position",
            "MNV lookup",
            "Project code",
            "Project",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
            "Expected target",
            "Actual target",
            "Matched?",
        ],
        detail_rows,
    )
    ws.freeze_panes = "A3"
    for col, width in {
        "A": 18,
        "B": 10,
        "C": 10,
        "D": 12,
        "E": 18,
        "F": 18,
        "G": 12,
        "H": 12,
        "I": 24,
        "J": 16,
        "K": 14,
        "L": 18,
        "M": 30,
        "N": 26,
        "O": 40,
        "P": 10,
        "Q": 32,
        "R": 24,
        "S": 24,
        "T": 10,
    }.items():
        ws.column_dimensions[col].width = width


def read_sx_downstream_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, SHEET_SX_DOWNSTREAM_CHECKPOINT)
    if ws is None:
        return []

    header_row = find_section_header(ws, "Details")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Stage", "Project", "Issue", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in Check_SX_Downstream: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker in {"Summary", "Details"}:
            break
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue
        approvals.append(
            {
                "approval_row": r,
                "stage": ws.cell(r, headers["Stage"]).value,
                "year": ws.cell(r, headers.get("Year", 0)).value if headers.get("Year") else None,
                "month": month_number(ws.cell(r, headers.get("Month", 0)).value) if headers.get("Month") else None,
                "program": ws.cell(r, headers.get("Program", 0)).value if headers.get("Program") else None,
                "source_file": ws.cell(r, headers.get("Source file", 0)).value if headers.get("Source file") else None,
                "source_sheet": ws.cell(r, headers.get("Source sheet", 0)).value if headers.get("Source sheet") else None,
                "source_row": ws.cell(r, headers.get("Source row", 0)).value if headers.get("Source row") else None,
                "append_row": ws.cell(r, headers.get("Append row", 0)).value if headers.get("Append row") else None,
                "employee": ws.cell(r, headers.get("Employee", 0)).value if headers.get("Employee") else None,
                "position": ws.cell(r, headers.get("Position", 0)).value if headers.get("Position") else None,
                "mnv_lookup": ws.cell(r, headers.get("MNV lookup", 0)).value if headers.get("MNV lookup") else None,
                "project_code": ws.cell(r, headers.get("Project code", 0)).value if headers.get("Project code") else None,
                "project": ws.cell(r, headers["Project"]).value,
                "project_name": ws.cell(r, headers["Project"]).value,
                "issue": ws.cell(r, headers["Issue"]).value,
                "recommended_action": ws.cell(r, headers["Recommended action"]).value,
                "apply": ws.cell(r, headers["Apply?"]).value,
                "approval_notes": ws.cell(r, headers.get("Approval notes", 0)).value if headers.get("Approval notes") else None,
                "expected_target": ws.cell(r, headers.get("Expected target", 0)).value if headers.get("Expected target") else None,
                "actual_target": ws.cell(r, headers.get("Actual target", 0)).value if headers.get("Actual target") else None,
                "matched": ws.cell(r, headers.get("Matched?", 0)).value if headers.get("Matched?") else None,
            }
        )
    return approvals


def write_sx_downstream_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, SHEET_SX_DOWNSTREAM_RESULT)
    headers = [
        "Approval row",
        "Stage",
        "Year",
        "Month",
        "Program",
        "Employee",
        "Project",
        "Issue",
        "Recommended action",
        "Apply?",
        "Resolved?",
        "Resolution note",
        "Status",
    ]
    append_table(ws, 1, "SX Downstream Approval Result", headers, results)
    for col, width in {"A": 14, "B": 20, "C": 10, "D": 10, "E": 12, "F": 24, "G": 28, "H": 28, "I": 42, "J": 10, "K": 12, "L": 40, "M": 80}.items():
        ws.column_dimensions[col].width = width


def apply_sx_downstream_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    approvals = read_sx_downstream_approvals(approval_file)
    if not approvals:
        return

    results: list[list[Any]] = []
    for approval in approvals:
        is_yes = approval_value_is_yes(approval.get("apply"))
        issue = str(approval.get("issue") or "")
        recommended_action = str(approval.get("recommended_action") or "")
        stage = str(approval.get("stage") or "")
        status = "skipped: Apply? not yes"
        resolved = "NO"
        resolution_note = None
        if is_yes:
            actions = [clean(action) for action in recommended_action.split(";") if clean(action)]
            action_statuses: list[str] = []
            for action in actions:
                try:
                    if action == "ADD_TO_PROJECT_CATALOG":
                        action_statuses.append(apply_add_to_project_catalog(wb[SHEET_PROJECT_CATALOG], approval))
                    elif action == "ADD_TO_CAPITALIZATION":
                        action_statuses.append(apply_add_to_capitalization(wb[SHEET_CAPITALIZATION], approval))
                    elif "missing month" in norm(issue) and stage in {"4.1 Chi phí nhân sự SX", SHEET_CAPITALIZATION, "Timesheet SX"}:
                        action_statuses.append("applied: downstream month block retained for regenerated workbook")
                    else:
                        action_statuses.append(f"skipped: unsupported action {action}")
                except Exception as exc:
                    action_statuses.append(f"failed {action}: {exc}")
            if action_statuses:
                status = "; ".join(action_statuses)
            else:
                status = "applied: downstream issue reviewed"
            resolved = "YES"
            resolution_note = "Issue processed from approval file"

        results.append(
            [
                approval.get("approval_row"),
                stage,
                approval.get("year"),
                approval.get("month"),
                approval.get("program"),
                approval.get("employee"),
                approval.get("project"),
                approval.get("issue"),
                approval.get("recommended_action"),
                approval.get("apply"),
                resolved,
                resolution_note,
                status,
            ]
        )

    write_sx_downstream_approval_result_sheet(wb, results)


def append_sx_to_template(
    wb,
    sx_rows: list[dict[str, Any]],
    employee_lookup: dict[str, str],
    year: int,
    month: int,
) -> tuple[int, int]:
    ws = find_sheet(wb, SHEET_SX_TARGET)
    if ws is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_SX_TARGET}.")

    include_cfa = (year, month) >= (2026, 5)
    template_row = SX_TEMPLATE_ROW
    append_row = find_sx_append_start_row(ws, template_row=template_row)
    summary_rows, detail_rows, missing_mnv_rows, kept_rows, skipped_rows = build_sx_checkpoint_data(
        sx_rows,
        employee_lookup,
        include_cfa,
        append_start_row=append_row,
    )

    current_row = append_row

    for row in sx_rows:
        program = norm(row.get("program")).upper()
        if program not in SX_ALLOWED_PROGRAMS:
            continue
        if program == "CFA" and not include_cfa:
            continue

        if current_row == append_row:
            copy_row_style_and_formulas_only(ws, template_row, current_row, 18)
        else:
            copy_row_style_and_formulas_only(ws, current_row - 1, current_row, 18)

        ws.cell(current_row, 1).value = month
        ws.cell(current_row, 2).value = clean(row.get("program"))
        ws.cell(current_row, 3).value = clean(row.get("position"))
        ws.cell(current_row, 4).value = clean(row.get("employee"))
        ws.cell(current_row, 6).value = clean(row.get("product"))
        ws.cell(current_row, 7).value = None
        ws.cell(current_row, 8).value = clean(row.get("product_type"))
        ws.cell(current_row, 9).value = clean(row.get("project"))
        ws.cell(current_row, 10).value = clean(row.get("department"))
        ws.cell(current_row, 11).value = clean(row.get("feature"))
        ws.cell(current_row, 12).value = clean(row.get("deliverable"))
        ws.cell(current_row, 13).value = None
        ws.cell(current_row, 14).value = None
        ws.cell(current_row, 15).value = clean(row.get("unit"))
        ws.cell(current_row, 16).value = row.get("actual")
        ws.cell(current_row, 17).value = row.get("kpi_standard")
        ws.cell(current_row, 18).value = None

        for detail in detail_rows:
            if detail[0] == row.get("year") and detail[1] == row.get("month") and norm(detail[2]).upper() == program and detail[5] == row.get("_source_row"):
                detail[4] = row.get("_source_sheet")
                detail[6] = current_row
                detail[7] = clean(row.get("employee"))
                detail[8] = clean(row.get("position"))
                detail[9] = lookup_employee_code(employee_lookup, row.get("employee"))
                break

        current_row += 1

    summary_rows[:0] = [
        ["Year", year],
        ["Month", month],
        ["Include CFA", "YES" if include_cfa else "NO"],
        ["Append start row", append_row],
        ["Append end row", current_row - 1 if current_row > append_row else None],
    ]
    if current_row > append_row:
        # Keep the filter/table range aligned with the newly appended SX rows.
        ws.auto_filter.ref = f"A2:V{current_row - 1}"
    write_sx_checkpoint_sheet(wb, summary_rows, detail_rows, missing_mnv_rows)
    return kept_rows, skipped_rows


def clear_values(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).value = None


def find_header_row(ws, required_terms: list[str], max_scan_row: int = 20) -> tuple[int | None, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row or 0, max_scan_row) + 1):
        headers: dict[str, int] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            h = norm(ws.cell(row_idx, col).value)
            if h:
                headers[h] = col
        if sum(1 for term in required_terms if any(term in h for h in headers)) >= max(2, len(required_terms) // 2):
            return row_idx, headers
    return None, {}


def find_section_header(ws, section_title: str, max_scan_row: int = 120) -> int | None:
    target = norm(section_title)
    if not target:
        return None
    for row_idx in range(1, min(ws.max_row or 0, max_scan_row) + 1):
        first_cell = norm(ws.cell(row_idx, 1).value)
        if first_cell == target:
            return row_idx + 1
    return None


def get_header_col(headers: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        alias_n = norm(alias)
        if alias_n in headers:
            return headers[alias_n]
        for header, col in headers.items():
            if alias_n in header:
                return col
    return None


def cell_value(ws, row: int, headers: dict[str, int], *aliases: str) -> Any:
    col = get_header_col(headers, *aliases)
    if col is None:
        return None
    return clean(ws.cell(row, col).value)


def parse_month_value(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.month
    if is_num(value):
        n = int(float(value))
        return n if 1 <= n <= 12 else None
    text = str(value).strip()
    if not text:
        return None
    compact_prefix = re.match(r"^(0[1-9]|1[0-2])\d{2}(?:\D|$)", text)
    if compact_prefix:
        return int(compact_prefix.group(1))
    m = re.search(r"\b(1[0-2]|0?[1-9])\b", text)
    if m:
        return int(m.group(1))
    for name, num in [
        ("jan", 1), ("feb", 2), ("mar", 3), ("apr", 4), ("may", 5), ("jun", 6),
        ("jul", 7), ("aug", 8), ("sep", 9), ("oct", 10), ("nov", 11), ("dec", 12),
    ]:
        if name in text.lower():
            return num
    return None


def month_number(value: Any) -> int | None:
    return parse_month_value(value)


def parse_hours(value: Any) -> float | None:
    if value is None:
        return None
    if is_num(value):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", text)
    if m:
        return float(m.group(1))
    return None


def numeric(value: Any) -> float | None:
    return parse_hours(value)


def build_employee_lookup(template_wb) -> dict[str, str]:
    if SHEET_EMPLOYEE not in template_wb.sheetnames:
        return {}

    ws = template_wb[SHEET_EMPLOYEE]
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(1, c).value)
        if h:
            headers[h] = c

    name_col = get_header_col(headers, "ten nhan vien", "ho ten", "employee")
    code_col = get_header_col(headers, "ma nhan vien", "mnv", "employee code")
    if name_col is None or code_col is None:
        return {}

    lookup: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, name_col).value)
        code = clean(ws.cell(r, code_col).value)
        if name and code:
            lookup[norm(name)] = str(code)
    return lookup


def lookup_employee_code(lookup: dict[str, str], name: Any) -> str | None:
    if not name:
        return None
    return lookup.get(norm(name))


def resolve_sx_employee_code(raw_code: Any, employee: Any, employee_lookup: dict[str, str]) -> str:
    code_text = clean(raw_code)
    if code_text and not code_text.startswith("="):
        normalized = code_text.upper().replace(" ", "")
        if normalized.startswith("MNV"):
            return code_text

    looked_up = lookup_employee_code(employee_lookup, employee)
    if looked_up:
        return looked_up

    if code_text and not code_text.startswith("="):
        return code_text
    return "Khong tim thay"


def sx_allocation_month_share_col(month: int) -> int:
    return SX_ALLOCATION_SHARE_START_COL + month - 1


def find_sheet(wb, sheet_name: str):
    target = str(sheet_name).strip()
    if not target:
        return None
    for ws in wb.worksheets:
        if str(ws.title).strip() == target:
            return ws
    for ws in wb.worksheets:
        if str(ws.title).strip().casefold() == target.casefold():
            return ws
    return None


def find_it_sheet(wb):
    for ws in wb.worksheets:
        if str(ws.title).isdigit():
            return ws
    return wb.worksheets[0]


def read_it_sheet(path: Path) -> tuple[str, list[list[Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = find_it_sheet(wb)
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]
    return ws.title, rows


def write_it_sheet(ws, rows: list[list[Any]], source_path: Path) -> None:
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= IT_HEADER_ROW:
            ws.unmerge_cells(str(mc))

    for r in range(IT_HEADER_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    needed_max_row = IT_HEADER_ROW + max(0, len(rows) - IT_HEADER_ROW)
    if needed_max_row > ws.max_row:
        for r in range(ws.max_row + 1, needed_max_row + 1):
            ensure_row_style(ws, IT_TEMPLATE_ROW, r, ws.max_column)

    for offset, row in enumerate(rows[IT_HEADER_ROW - 1 :]):
        r = IT_HEADER_ROW + offset
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = clean(value)

    # Fill down cột "Tên dự án" theo từng block merge gốc của source IT.
    # Nếu chỉ copy nguyên trạng, các dòng tháng 2..12 sẽ bị blank ở cột B
    # và SUMIFS phía dưới không khớp criteria project name.
    src_wb = load_workbook(source_path, data_only=False)
    src_ws = find_it_sheet(src_wb)
    for mc in src_ws.merged_cells.ranges:
        if mc.min_col != 2 or mc.max_col != 2:
            continue
        if mc.min_row < IT_HEADER_ROW + 2:
            continue
        project_name = clean(src_ws.cell(mc.min_row, 2).value)
        if project_name is None:
            continue
        top_style = ws.cell(mc.min_row, 2)
        for r in range(mc.min_row, mc.max_row + 1):
            cell = ws.cell(r, 2)
            cell.value = project_name
            if r != mc.min_row and top_style.has_style:
                cell._style = copy(top_style._style)
                cell.number_format = top_style.number_format
                cell.font = copy(top_style.font)
                cell.fill = copy(top_style.fill)
                cell.border = copy(top_style.border)
                cell.alignment = copy(top_style.alignment)
                cell.protection = copy(top_style.protection)


def sheet_bu_hint(sheet_name: str) -> str | None:
    name = norm(sheet_name)
    if "cfa" in name:
        return "CFA"
    if "cma" in name:
        return "CMA"
    if "acca" in name:
        return "ACCA"
    return None


def extract_media_rows(path: Path, employee_lookup: dict[str, str]) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)

    rows: list[dict[str, Any]] = []
    required = ["ngay tao", "ten nhiem vu", "nguoi lam task"]

    for ws in wb.worksheets:
        header_row, headers = find_header_row(ws, required_terms=required, max_scan_row=8)
        if header_row is None:
            continue

        bu_from_sheet = sheet_bu_hint(ws.title)
        for r in range(header_row + 1, ws.max_row + 1):
            raw = {
                "bu": cell_value(ws, r, headers, "bu", "phong ban giao nhiem vu", "bo phan") or bu_from_sheet,
                "project_name": cell_value(
                    ws, r, headers,
                    "ten du an sx",
                    "ten du an",
                    "ten du an sx ( diền theo list ở sheet \"ten du an\")",
                    "ten du an sx ( diền theo list ở sheet 'ten du an')",
                ),
                "task_name": cell_value(ws, r, headers, "ten nhiem vu", "nhiem vu", "ten task", "task"),
                "hours": parse_hours(
                    cell_value(
                        ws,
                        r,
                        headers,
                        "so gio hoan thanh cong viec",
                        "thoi gian can de hoan thanh cong viec",
                        "so gio lam viec thuc te",
                        "thoi gian leader phan bo",
                        "thoi gian hoan thanh task",
                        "thoi gian",
                        "total gio",
                    )
                ),
                "month": parse_month_value(cell_value(ws, r, headers, "thang", "ngay tao"))
                or parse_month_value(cell_value(ws, r, headers, "ten nhiem vu", "nhiem vu", "ten task", "task")),
                "employee": cell_value(
                    ws,
                    r,
                    headers,
                    "nguoi lam task (dien ho va ten)",
                    "nguoi lam task",
                    "giao cho",
                    "nguoi tao",
                    "nguoi thuc hien",
                ),
                "mnv": cell_value(ws, r, headers, "ma nhan vien", "mnv"),
            }

            if not any([
                raw["bu"],
                raw["project_name"],
                raw["task_name"],
                raw["hours"],
                raw["month"],
                raw["employee"],
                raw["mnv"],
            ]):
                continue

            mnv = raw["mnv"] or lookup_employee_code(employee_lookup, raw["employee"])
            rows.append(
                {
                    "source_sheet": ws.title,
                    "source_row": r,
                    "bu": raw["bu"],
                    "project_name": raw["project_name"],
                    "task_name": raw["task_name"],
                    "hours": raw["hours"],
                    "month": raw["month"],
                    "employee": raw["employee"],
                    "mnv": mnv,
                }
            )

    return rows


def derive_project_name(task_name: Any, fallback: Any = None) -> Any:
    if fallback:
        return fallback
    if not task_name:
        return None
    text = str(task_name).strip()
    text = re.sub(r"^\d{3,4}[_-]", "", text)
    text = re.sub(r"^(ACCA|CFA|CMA|IT|MEDIA)[_-]", "", text, flags=re.I)
    if "_" in text:
        candidate = text.split("_", 1)[0].strip()
        if candidate and len(candidate) <= 40:
            return candidate
    if " - " in text:
        candidate = text.split(" - ", 1)[0].strip()
        if candidate and len(candidate) <= 40:
            return candidate
    return None


def ensure_media_rows(ws, count: int) -> None:
    needed_max_row = MEDIA_TEMPLATE_ROW + max(0, count - 1)
    if needed_max_row > ws.max_row:
        for r in range(ws.max_row + 1, needed_max_row + 1):
            ensure_row_style(ws, MEDIA_TEMPLATE_ROW, r, ws.max_column)


def media_formula_h(row: int) -> str:
    return (
        f"=_xlfn.XLOOKUP(1, "
        f"('Lương nhân viên full time'!$A:$A='Data media ACCA+CFA+CMA'!E{row})*"
        f"('Lương nhân viên full time'!$E:$E='Data media ACCA+CFA+CMA'!G{row}), "
        f"'Lương nhân viên full time'!$AG:$AG)"
    )


def write_media_sheet(ws, rows: list[dict[str, Any]]) -> None:
    clear_values(ws, MEDIA_TEMPLATE_ROW, ws.max_row, 1, min(ws.max_column, 9))
    ensure_media_rows(ws, len(rows))

    for idx, row in enumerate(rows):
        r = MEDIA_TEMPLATE_ROW + idx
        project_name = derive_project_name(row["task_name"], row["project_name"])
        ws.cell(r, 1).value = row["bu"]
        ws.cell(r, 2).value = project_name
        ws.cell(r, 3).value = row["task_name"]
        ws.cell(r, 4).value = row["hours"]
        ws.cell(r, 5).value = row["month"]
        ws.cell(r, 6).value = row["employee"]
        ws.cell(r, 7).value = row["mnv"]
        ws.cell(r, 8).value = media_formula_h(r)
        ws.cell(r, 9).value = f"=D{r}/H{r}"


def write_media_audit_sheet(wb, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.append(MEDIA_AUDIT_FIELDS)
    for row in rows:
        ws.append([row.get(field) for field in MEDIA_AUDIT_FIELDS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def reset_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
        return ws
    return wb.create_sheet(title=sheet_name)


def append_table(ws, start_row: int, title: str, headers: list[str], rows: list[list[Any]]) -> int:
    ws.cell(start_row, 1).value = title
    header_row = start_row + 1
    for c, value in enumerate(headers, start=1):
        ws.cell(header_row, c).value = value
    for r_offset, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(header_row + r_offset, c).value = value
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
    return header_row + len(rows) + 2


def clear_worksheet(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def copy_sheet_content(src_ws, dst_ws) -> None:
    clear_worksheet(dst_ws)

    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(src_cell.row, src_cell.column)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format
            if src_cell.font:
                dst_cell.font = copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border = copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy(src_cell.alignment)
            if src_cell.protection:
                dst_cell.protection = copy(src_cell.protection)

    for col_key, col_dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_key]
        dst_dim.width = col_dim.width
        dst_dim.hidden = col_dim.hidden
        dst_dim.outlineLevel = col_dim.outlineLevel

    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[row_idx]
        dst_dim.height = row_dim.height
        dst_dim.hidden = row_dim.hidden
        dst_dim.outlineLevel = row_dim.outlineLevel

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter and src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref


def patch_pivot_refresh_flags(xlsx_path: Path) -> None:
    pivot_cache_pattern = re.compile(r"(<pivotCacheDefinition\b[^>]*)(/?>)", re.IGNORECASE)

    tmp_path = xlsx_path.with_suffix(".refreshing.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as src_zip, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as dst_zip:
        for info in src_zip.infolist():
            data = src_zip.read(info.filename)
            if info.filename.startswith("xl/pivotCache/pivotCacheDefinition") and info.filename.endswith(".xml"):
                text = data.decode("utf-8")
                if "refreshOnLoad" not in text:
                    text = pivot_cache_pattern.sub(lambda m: f"{m.group(1)} refreshOnLoad=\"1\"{m.group(2)}", text, count=1)
                else:
                    text = re.sub(r'refreshOnLoad=\"[^\"]*\"', 'refreshOnLoad=\"1\"', text, count=1)
                data = text.encode("utf-8")
            dst_zip.writestr(info, data)

    xlsx_path.unlink(missing_ok=True)
    tmp_path.replace(xlsx_path)


def recalculate_workbook_with_excel(xlsx_path: Path) -> bool:
    if not xlsx_path.exists():
        return False

    # Excel is the only reliable way here to materialize formula caches in the saved .xlsx.
    workbook_path = str(xlsx_path.resolve())
    powershell_script = rf"""
$ErrorActionPreference = 'Stop'
$path = '{workbook_path}'
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.EnableEvents = $false
    $excel.ScreenUpdating = $false
    try {{
        $excel.Calculation = -4105
    }} catch {{}}
    $workbook = $excel.Workbooks.Open($path, 0, $false)
    try {{
        $excel.CalculateFullRebuild() | Out-Null
    }} catch {{
        $excel.CalculateFull() | Out-Null
    }}
    $workbook.Save()
    $workbook.Close($true)
    $excel.Quit()
    exit 0
}} catch {{
    if ($workbook) {{
        try {{ $workbook.Close($false) }} catch {{}}
    }}
    if ($excel) {{
        try {{ $excel.Quit() }} catch {{}}
    }}
    Write-Error $_
    exit 1
}} finally {{
    if ($workbook) {{
        try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null }} catch {{}}
    }}
    if ($excel) {{
        try {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }} catch {{}}
    }}
}}
"""

    try:
        import subprocess

        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", powershell_script],
            capture_output=True,
            text=True,
            timeout=300,
        )
        if result.returncode != 0:
            print(f"  Excel recalc skipped: {result.stderr.strip() or result.stdout.strip() or 'unknown error'}")
            return False
        return True
    except Exception as exc:
        print(f"  Excel recalc skipped: {exc}")
        return False


def rebuild_sx_consol_sheet(wb, employee_lookup: dict[str, str]) -> None:
    raw_ws = find_sheet(wb, SHEET_SX_TARGET)
    consol_ws = find_sheet(wb, "Data SX consol")
    if raw_ws is None or consol_ws is None:
        return

    # Collect month totals per (project, employee, employee-code) and per employee/month.
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    employee_month_totals: dict[tuple[str, int], float] = {}

    for r in range(4, raw_ws.max_row + 1):
        month = month_number(raw_ws.cell(r, 1).value)
        project = clean(raw_ws.cell(r, 9).value)
        employee = clean(raw_ws.cell(r, 4).value)
        if month is None or not project or not employee:
            continue

        code = resolve_sx_employee_code(raw_ws.cell(r, 5).value, employee, employee_lookup)
        actual = parse_hours(raw_ws.cell(r, 16).value) or 0.0
        kpi_standard = parse_hours(raw_ws.cell(r, 17).value) or 0.0
        total_kpi = parse_hours(raw_ws.cell(r, 18).value)
        if total_kpi is None:
            total_kpi = actual * kpi_standard
        if total_kpi == 0:
            continue

        key = (project, employee, code)
        bucket = grouped.setdefault(
            key,
            {
                "project": project,
                "employee": employee,
                "code": code,
                "first_row": r,
                "months": {m: 0.0 for m in range(1, 13)},
            },
        )
        bucket["months"][month] += float(total_kpi)
        employee_month_totals[(employee, month)] = employee_month_totals.get((employee, month), 0.0) + float(total_kpi)

    rows = sorted(
        grouped.values(),
        key=lambda item: (
            norm(item["project"]),
            norm(item["employee"]),
            norm(item["code"]),
            int(item["first_row"]),
        ),
    )

    # Clear the old consolidated zones so stale spill values do not remain visible.
    clear_values(consol_ws, 6, max(consol_ws.max_row, 200), 1, 27)
    clear_values(consol_ws, 32, max(consol_ws.max_row, 200), 30, 44)

    summary_start_row = 6
    compact_start_row = 32
    for idx, item in enumerate(rows):
        summary_row = summary_start_row + idx
        compact_row = compact_start_row + idx
        month_totals = [float(item["months"][m]) for m in range(1, 13)]
        month_shares = []
        for month in range(1, 13):
            denom = employee_month_totals.get((item["employee"], month), 0.0)
            month_shares.append((month_totals[month - 1] / denom) if denom else 0.0)

        summary_values = [
            item["project"],
            item["employee"],
            item["code"],
            *month_totals,
            *month_shares,
        ]
        for col_idx, value in enumerate(summary_values, start=1):
            consol_ws.cell(summary_row, col_idx).value = value

        compact_values = [
            item["project"],
            item["employee"],
            item["code"],
            *month_shares,
        ]
        for col_idx, value in enumerate(compact_values, start=30):  # AD
            consol_ws.cell(compact_row, col_idx).value = value

    # Keep the old filter range aligned with the rebuilt summary.
    if rows:
        consol_ws.auto_filter.ref = f"A5:AA{summary_start_row + len(rows) - 1}"


def remove_worksheet_if_exists(wb, sheet_name: str) -> bool:
    ws = find_sheet(wb, sheet_name)
    if ws is None:
        return False
    wb.remove(ws)
    return True


def set_sheet_tab_color(ws, color_argb: str | None) -> None:
    if color_argb is None:
        return
    try:
        ws.sheet_properties.tabColor = Color(rgb=color_argb)
    except Exception:
        pass


def arrange_workbook_sections(wb) -> None:
    section_orders = {
        "COMMON": [
            "Hướng dẫn",
            "1.Danh mục dự án",
            "2.Nhân sự thực hiện",
            "3.Vốn hóa",
            "Lương nhân viên full time",
            "Lương nhân viên part time",
            "Mã nhân viên",
            "Check mã nhân viên",
            "Check_Payroll",
            "Huong_dan_Approval",
            "Khu_vuc_xu_ly_chung",
            "Check_IT_Cost_Month_Block",
            "Check_Media_Timesheet",
            "Check_SX_Downstream",
            "Check_Vonhoa_Month_Block",
        ],
        "IT": [
            "Timesheet IT",
            "Chi phí nhân sự IT",
            "Sổ kế toán SC",
            "242",
            "242 T8-9",
            "Bút toán điều chỉnh",
            "Checking Vốn hóa IT",
            "IT_Approval_Result",
            "Check_IT_CPNS",
            "IT_New_Project_Master",
            "Check_IT_Downstream",
            "IT_Cost_Block_Approval_Result",
        ],
        "MEDIA": [
            "Timesheet Media",
            "Data media ACCA+CFA+CMA",
            "Media_CFA",
            "Media_CMA",
            "Media_ACCA",
            "Media_Approval_Result",
        ],
        "SX": [
            "Data SX ACCA+CMA",
            "Data SX CFA",
            "Data SX+Media SC",
            "SX_Allocation_Build",
            "Timesheet SX",
            "4.1 Chi phí nhân sự SX",
            "checkpoint data SX",
            "SX_Approval_Result",
            "Check_SX_Downstream",
            "SX_Downstream_Approval_Result",
            "Vonhoa_Block_Approval_Result",
        ],
    }

    group_colors = {
        "COMMON": SHEET_GROUP_COLORS["COMMON"],
        "IT": SHEET_GROUP_COLORS["IT"],
        "MEDIA": SHEET_GROUP_COLORS["MEDIA"],
        "SX": SHEET_GROUP_COLORS["SX"],
    }

    ordered_names: list[str] = []
    seen: set[str] = set()
    for group in ["COMMON", "IT", "MEDIA", "SX"]:
        for sheet_name in section_orders[group]:
            if sheet_name in wb.sheetnames and sheet_name not in seen:
                ordered_names.append(sheet_name)
                seen.add(sheet_name)

    for sheet_name in wb.sheetnames:
        if sheet_name not in seen:
            ordered_names.append(sheet_name)
            seen.add(sheet_name)

    wb._sheets = [wb[sheet_name] for sheet_name in ordered_names]

    for sheet_name in section_orders["COMMON"]:
        if sheet_name in wb.sheetnames:
            set_sheet_tab_color(wb[sheet_name], group_colors["COMMON"])
    for sheet_name in section_orders["IT"]:
        if sheet_name in wb.sheetnames:
            set_sheet_tab_color(wb[sheet_name], group_colors["IT"])
    for sheet_name in section_orders["MEDIA"]:
        if sheet_name in wb.sheetnames:
            set_sheet_tab_color(wb[sheet_name], group_colors["MEDIA"])
    for sheet_name in section_orders["SX"]:
        if sheet_name in wb.sheetnames:
            set_sheet_tab_color(wb[sheet_name], group_colors["SX"])


def sanitize_sx_raw_formulas(wb) -> int:
    raw_ws = find_sheet(wb, SHEET_SX_TARGET)
    if raw_ws is None:
        return 0

    replaced = 0
    for row in raw_ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("=") and "KPIs Standard (final)" in value:
                cell.value = None
                replaced += 1
    return replaced


def rebuild_sx_allocation_sheet(wb, employee_lookup: dict[str, str], year: int) -> int:
    raw_ws = find_sheet(wb, SHEET_SX_TARGET)
    allocation_ws = find_sheet(wb, SHEET_SX_ALLOCATION)
    if raw_ws is None:
        return 0
    if allocation_ws is None:
        allocation_ws = wb.create_sheet(title=SHEET_SX_ALLOCATION)
    else:
        clear_worksheet(allocation_ws)

    headers = [
        "year",
        "program",
        "project_raw",
        "employee_name",
        "employee_code",
        "department",
        "position",
        "source_file",
        "source_sheet",
        "source_row_min",
        "source_row_max",
    ]
    headers.extend([f"month_{month}_kpi" for month in range(1, 13)])
    headers.extend([f"month_{month}_share" for month in range(1, 13)])
    headers.extend(["allocation_status", "issue", "recommended_action"])

    for c, value in enumerate(headers, start=1):
        allocation_ws.cell(1, c).value = value

    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    employee_month_totals: dict[tuple[Any, ...], float] = {}

    for r in range(4, raw_ws.max_row + 1):
        month = month_number(raw_ws.cell(r, 1).value)
        program = clean(raw_ws.cell(r, 2).value)
        position = clean(raw_ws.cell(r, 3).value)
        employee = clean(raw_ws.cell(r, 4).value)
        project = clean(raw_ws.cell(r, 9).value)
        source_department = clean(raw_ws.cell(r, 10).value)
        department = "SX"
        if month is None or not program or not employee or not project:
            continue

        program_key = norm(program).upper()
        if program_key not in SX_ALLOWED_PROGRAMS:
            continue
        if program_key == "CFA" and (year, month) < (2026, 5):
            continue

        actual = parse_hours(raw_ws.cell(r, 16).value) or 0.0
        kpi_standard = parse_hours(raw_ws.cell(r, 17).value) or 0.0
        total_kpi = parse_hours(raw_ws.cell(r, 18).value)
        if total_kpi is None:
            total_kpi = actual * kpi_standard
        if not total_kpi:
            continue

        employee_code = lookup_employee_code(employee_lookup, employee) or "Khong tim thay"
        grouping_code = employee_code if employee_code != "Khong tim thay" else str(employee)
        key = (year, program_key, project, str(employee), employee_code, department)
        bucket = grouped.setdefault(
            key,
            {
                "year": year,
                "program": program_key,
                "project_raw": project,
                "employee_name": employee,
                "employee_code": employee_code,
                "department": department,
                "position": position,
                "source_file": source_file_for_program(program_key),
                "source_sheet": raw_ws.title,
                "source_row_min": r,
                "source_row_max": r,
                "source_department": source_department,
                "months": {m: 0.0 for m in range(1, 13)},
            },
        )
        bucket["source_row_min"] = min(bucket["source_row_min"], r)
        bucket["source_row_max"] = max(bucket["source_row_max"], r)
        bucket["months"][month] += float(total_kpi)
        employee_month_totals[(year, grouping_code, month)] = employee_month_totals.get((year, grouping_code, month), 0.0) + float(total_kpi)

    rows = sorted(
        grouped.values(),
        key=lambda item: (
            norm(item["program"]),
            norm(item["project_raw"]),
            norm(item["employee_name"]),
            norm(item["employee_code"]),
            norm(item["department"]),
            int(item["source_row_min"]),
        ),
    )

    for idx, item in enumerate(rows, start=2):
        month_totals = [float(item["months"][m]) for m in range(1, 13)]
        grouping_code = item["employee_code"] if item["employee_code"] != "Khong tim thay" else str(item["employee_name"])
        month_shares = []
        for month in range(1, 13):
            denom = employee_month_totals.get((year, grouping_code, month), 0.0)
            month_shares.append((month_totals[month - 1] / denom) if denom else 0.0)

        allocation_status = "OK"
        issue = None
        recommended_action = None
        if item["employee_code"] == "Khong tim thay":
            allocation_status = "missing MNV"
            issue = "missing MNV"
            recommended_action = "Add or fix the employee in sheet 'Mã nhân viên' or correct the source name"

        values = [
            item["year"],
            item["program"],
            item["project_raw"],
            item["employee_name"],
            item["employee_code"],
            item["department"],
            item["position"],
            item["source_file"],
            item["source_sheet"],
            item["source_row_min"],
            item["source_row_max"],
            *month_totals,
            *month_shares,
            allocation_status,
            issue,
            recommended_action,
        ]
        for c, value in enumerate(values, start=1):
            allocation_ws.cell(idx, c).value = value

    allocation_ws.freeze_panes = "A2"
    allocation_ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    for col, width in {
        "A": 10,
        "B": 12,
        "C": 34,
        "D": 24,
        "E": 14,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 12,
        "T": 12,
        "U": 12,
        "V": 12,
        "W": 12,
        "X": 12,
        "Y": 12,
        "Z": 12,
        "AA": 12,
        "AB": 12,
        "AC": 12,
        "AD": 12,
        "AE": 12,
        "AF": 12,
        "AG": 12,
        "AH": 12,
        "AI": 12,
        "AJ": 16,
        "AK": 24,
        "AL": 42,
    }.items():
        allocation_ws.column_dimensions[col].width = width
    return len(rows)


def rewrite_timesheet_sx_from_allocation(wb) -> int:
    timesheet_ws = find_sheet(wb, "Timesheet SX")
    allocation_ws = find_sheet(wb, SHEET_SX_ALLOCATION)
    if timesheet_ws is None or allocation_ws is None:
        return 0

    rewritten = 0
    for r in range(4, timesheet_ws.max_row + 1):
        if not any(clean(timesheet_ws.cell(r, c).value) is not None for c in range(1, 8)):
            continue

        for month_col in range(8, 20):
            header_month = month_col - 7
            share_col = sx_allocation_month_share_col(header_month)
            share_letter = get_column_letter(share_col)
            formula = (
                f"=SUMIFS('{SHEET_SX_ALLOCATION}'!${share_letter}:${share_letter},"
                f"'{SHEET_SX_ALLOCATION}'!$A:$A,$A{r},"
                f"'{SHEET_SX_ALLOCATION}'!$B:$B,$B{r},"
                f"'{SHEET_SX_ALLOCATION}'!$C:$C,$D{r},"
                f"'{SHEET_SX_ALLOCATION}'!$D:$D,$F{r},"
                f"'{SHEET_SX_ALLOCATION}'!$E:$E,$G{r},"
                f"'{SHEET_SX_ALLOCATION}'!$F:$F,$E{r})"
            )
            timesheet_ws.cell(r, month_col).value = formula
            rewritten += 1
    return rewritten


def copy_column_block_with_translation(ws, source_start_col: int, target_start_col: int, width: int) -> None:
    for offset in range(width):
        src_col = source_start_col + offset
        dst_col = target_start_col + offset
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_col)
        src_dim = ws.column_dimensions[src_letter]
        dst_dim = ws.column_dimensions[dst_letter]
        dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        dst_dim.outlineLevel = src_dim.outlineLevel

    for r in range(1, ws.max_row + 1):
        for offset in range(width):
            src = ws.cell(r, source_start_col + offset)
            dst = ws.cell(r, target_start_col + offset)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.protection:
                dst.protection = copy(src.protection)
            if isinstance(src.value, str) and src.value.startswith("="):
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            else:
                dst.value = src.value


def remap_month_block_external_refs(formula: str, sheet_name: str, month_columns: list[str]) -> str:
    if sheet_name not in formula:
        return formula

    pattern = re.compile(
        rf"('{re.escape(sheet_name)}'!)(\$(?:[A-Z]{{1,3}}))(:(\$(?:[A-Z]{{1,3}})))?"
    )

    def shift_column(col_ref: str) -> str:
        col = col_ref.replace("$", "")
        if col not in month_columns:
            return col_ref
        idx = month_columns.index(col)
        if idx >= len(month_columns) - 1:
            return col_ref
        return f"${month_columns[idx + 1]}"

    def repl(match: re.Match[str]) -> str:
        prefix = match.group(1)
        start = match.group(2)
        end = match.group(4)
        if end is not None:
            shifted_start = shift_column(start)
            shifted_end = shift_column(end)
            return f"{prefix}{shifted_start}:{shifted_end}"
        return f"{prefix}{shift_column(start)}"

    return pattern.sub(repl, formula)


def remap_capitalization_month_block_formulas(ws, start_col: int, width: int) -> None:
    # The copied month block inherits formulas from the previous month block.
    # Absolute references to month-based cost columns need to move forward as well,
    # otherwise month 5 still points at month 4 cost columns.
    month_ref_maps = {
        "Chi phí nhân sự IT": ["M", "Q", "W", "AI", "AO", "AU", "BA", "BG", "BM", "BS", "BY", "CE", "CK"],
        "4.1 Chi phí nhân sự SX": ["M", "Q", "U", "Y", "AC", "AG", "AK", "AO", "AS", "AW", "BA", "BE", "BI"],
    }

    end_col = start_col + width - 1
    for r in range(1, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            formula = cell.value
            formula = remap_month_block_external_refs(formula, "Chi phí nhân sự IT", month_ref_maps["Chi phí nhân sự IT"])
            formula = remap_month_block_external_refs(formula, "4.1 Chi phí nhân sự SX", month_ref_maps["4.1 Chi phí nhân sự SX"])
            cell.value = formula


def copy_capitalization_month_block(capitalization, month: int) -> str:
    month_col = find_month_label_col(capitalization, month, label_row=1)
    if month_col is not None:
        return "skipped: month block already exists"

    prev_month = month - 1
    source_col = find_month_label_col(capitalization, prev_month, label_row=1)
    if source_col is None:
        return f"failed: no template month block found for month {prev_month}"

    block_width = 8
    target_col = source_col + block_width
    copy_column_block_with_translation(capitalization, source_col, target_col, block_width)
    capitalization.cell(1, target_col).value = f"Tháng {month}"
    remap_capitalization_month_block_formulas(capitalization, target_col, block_width)
    return f"applied: copied month {prev_month} block to month {month}"


def build_capitalization_month_checkpoint(capitalization, month: int) -> tuple[list[list[Any]], list[list[Any]]]:
    summary: list[list[Any]] = []
    details: list[list[Any]] = []

    requested_month = month
    month_col = find_month_label_col(capitalization, requested_month, label_row=1)
    target_month = requested_month
    if month_col is not None:
        target_month = None
        for candidate in range(requested_month + 1, 13):
            if find_month_label_col(capitalization, candidate, label_row=1) is None:
                target_month = candidate
                break
    prev_month = target_month - 1 if target_month is not None else None
    prev_col = find_month_label_col(capitalization, prev_month, label_row=1) if prev_month is not None else None

    status = "OK" if target_month is None else "MISSING"
    summary.extend([
        ["Requested month", requested_month],
        ["Target month", target_month],
        ["Month block status", status],
        ["Previous month available", "YES" if prev_col is not None else "NO"],
        ["Recommended action", "Copy previous month block" if target_month is not None and prev_col is not None else "No action required"],
    ])

    if target_month is not None:
        details.append([
            SHEET_CAPITALIZATION,
            target_month,
            f"missing month {target_month} block",
            f"Copy month {prev_month} block to month {target_month}",
            "NO",
            None,
            prev_month,
            target_month,
            None,
        ])

    return summary, details


def write_capitalization_month_checkpoint_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Check_Vonhoa_Month_Block")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Value"],
        summary_rows,
    )
    append_table(
        ws,
        next_row,
        "Details",
        ["Sheet", "Month", "Issue", "Recommended action", "Apply?", "Approval notes", "Source month", "Target month", "Status"],
        detail_rows,
    )
    ws.freeze_panes = "A3"
    for col, width in {"A": 20, "B": 10, "C": 24, "D": 36, "E": 10, "F": 24, "G": 12, "H": 12, "I": 14}.items():
        ws.column_dimensions[col].width = width


def read_capitalization_month_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")
    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "Check_Vonhoa_Month_Block")
    if ws is None:
        return []

    header_row = find_section_header(ws, "Details")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Sheet", "Month", "Issue", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in Check_Vonhoa_Month_Block: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker in {"Summary", "Details"}:
            break
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue
        approvals.append(
            {
                "approval_row": r,
                "sheet": ws.cell(r, headers["Sheet"]).value,
                "month": month_number(ws.cell(r, headers["Month"]).value),
                "issue": ws.cell(r, headers["Issue"]).value,
                "recommended_action": ws.cell(r, headers["Recommended action"]).value,
                "apply": ws.cell(r, headers["Apply?"]).value,
                "approval_notes": ws.cell(r, headers.get("Approval notes", 0)).value if headers.get("Approval notes") else None,
                "source_month": month_number(ws.cell(r, headers.get("Source month", 0)).value) if headers.get("Source month") else None,
                "target_month": month_number(ws.cell(r, headers.get("Target month", 0)).value) if headers.get("Target month") else None,
                "status": ws.cell(r, headers.get("Status", 0)).value if headers.get("Status") else None,
            }
        )
    return approvals


def write_capitalization_month_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Vonhoa_Block_Approval_Result")
    headers = [
        "Approval row",
        "Sheet",
        "Month",
        "Issue",
        "Recommended action",
        "Apply?",
        "Resolved?",
        "Resolution note",
        "Status",
    ]
    append_table(ws, 1, "Von hoa Month Block Approval Result", headers, results)
    for col, width in {"A": 14, "B": 20, "C": 10, "D": 24, "E": 36, "F": 10, "G": 12, "H": 40, "I": 80}.items():
        ws.column_dimensions[col].width = width


def apply_capitalization_month_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    approvals = read_capitalization_month_approvals(approval_file)
    if not approvals:
        return

    capital = find_sheet(wb, SHEET_CAPITALIZATION)
    if capital is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_CAPITALIZATION}.")

    results: list[list[Any]] = []
    for approval in approvals:
        month = approval.get("month")
        issue = norm(approval.get("issue"))
        is_yes = approval_value_is_yes(approval.get("apply"))
        status = "skipped: Apply? not yes"
        resolved = "NO"
        resolution_note = None
        if is_yes:
            if month is None:
                status = "failed: missing month"
                resolution_note = "Set the target month in the approval workbook"
            elif "missing month" in issue:
                status = copy_capitalization_month_block(capital, month)
                resolved = "YES" if status.startswith("applied:") or status.startswith("skipped:") else "NO"
                resolution_note = f"Month block for {month} handled"
            else:
                status = "skipped: unsupported issue"
                resolution_note = "Only missing month block issues are supported"
        results.append([
            approval.get("approval_row"),
            approval.get("sheet"),
            month,
            approval.get("issue"),
            approval.get("recommended_action"),
            approval.get("apply"),
            resolved,
            resolution_note,
            status,
        ])

    write_capitalization_month_result_sheet(wb, results)


def build_it_cost_month_checkpoint(cost, month: int) -> tuple[list[list[Any]], list[list[Any]]]:
    summary: list[list[Any]] = []
    details: list[list[Any]] = []

    requested_month = month
    month_col = find_month_label_col(cost, requested_month, label_row=1)
    target_month = requested_month
    if month_col is not None and it_cost_month_block_has_formulas(cost, requested_month):
        target_month = None
        for candidate in range(requested_month + 1, 13):
            if not it_cost_month_block_has_formulas(cost, candidate):
                target_month = candidate
                break

    prev_month = target_month - 1 if target_month is not None else None
    prev_col = find_month_label_col(cost, prev_month, label_row=1) if prev_month is not None else None

    status = "OK" if target_month is None else "MISSING"
    summary.extend([
        ["Requested month", requested_month],
        ["Target month", target_month],
        ["Month block status", status],
        ["Previous month available", "YES" if prev_col is not None else "NO"],
        ["Recommended action", "Copy previous month block" if target_month is not None and prev_col is not None else "No action required"],
    ])

    if target_month is not None:
        issue = "missing formulas in month block" if month_col is not None else f"missing month {target_month} block"
        details.append([
            SHEET_IT_COST,
            target_month,
            issue,
            f"Copy month {prev_month} block to month {target_month}",
            "NO",
            None,
            prev_month,
            target_month,
            None,
        ])

    return summary, details


def write_it_cost_month_checkpoint_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Check_IT_Cost_Month_Block")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Value"],
        summary_rows,
    )
    append_table(
        ws,
        next_row,
        "Details",
        ["Sheet", "Month", "Issue", "Recommended action", "Apply?", "Approval notes", "Source month", "Target month", "Status"],
        detail_rows,
    )
    ws.freeze_panes = "A3"
    for col, width in {"A": 22, "B": 10, "C": 24, "D": 36, "E": 10, "F": 24, "G": 12, "H": 12, "I": 14}.items():
        ws.column_dimensions[col].width = width


def read_it_cost_month_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "Check_IT_Cost_Month_Block")
    if ws is None:
        return []

    header_row = find_section_header(ws, "Details")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Sheet", "Month", "Issue", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in Check_IT_Cost_Month_Block: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker in {"Summary", "Details"}:
            break
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue
        approvals.append(
            {
                "approval_row": r,
                "sheet": ws.cell(r, headers["Sheet"]).value,
                "month": month_number(ws.cell(r, headers["Month"]).value),
                "issue": ws.cell(r, headers["Issue"]).value,
                "recommended_action": ws.cell(r, headers["Recommended action"]).value,
                "apply": ws.cell(r, headers["Apply?"]).value,
                "approval_notes": ws.cell(r, headers.get("Approval notes", 0)).value if headers.get("Approval notes") else None,
                "source_month": month_number(ws.cell(r, headers.get("Source month", 0)).value) if headers.get("Source month") else None,
                "target_month": month_number(ws.cell(r, headers.get("Target month", 0)).value) if headers.get("Target month") else None,
                "status": ws.cell(r, headers.get("Status", 0)).value if headers.get("Status") else None,
            }
        )
    return approvals


def write_it_cost_month_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "IT_Cost_Block_Approval_Result")
    headers = ["Approval row", "Sheet", "Month", "Issue", "Recommended action", "Apply?", "Resolved?", "Resolution note", "Status"]
    append_table(ws, 1, "IT Cost Month Block Approval Result", headers, results)
    for col, width in {"A": 14, "B": 22, "C": 10, "D": 24, "E": 36, "F": 10, "G": 12, "H": 40, "I": 80}.items():
        ws.column_dimensions[col].width = width


def it_cost_month_block_has_formulas(cost, month: int) -> bool:
    start_col = find_month_label_col(cost, month, label_row=1)
    if start_col is None:
        return False

    data_cols = range(start_col + 1, min(start_col + 4, cost.max_column + 1))
    for r in range(3, min(cost.max_row, 80) + 1):
        for c in data_cols:
            value = cost.cell(r, c).value
            if isinstance(value, str) and value.startswith("="):
                return True
    return False


def copy_it_cost_month_block(cost, month: int) -> str:
    prev_month = month - 1
    source_col = find_month_label_col(cost, prev_month, label_row=1)
    if source_col is None:
        return f"failed: no template month block found for month {prev_month}"

    target_col = find_month_label_col(cost, month, label_row=1)
    if target_col is not None and it_cost_month_block_has_formulas(cost, month):
        return "skipped: month block already exists"
    if target_col is None:
        target_col = source_col + 6

    block_width = 6
    copy_column_block_with_translation(cost, source_col, target_col, block_width)
    cost.cell(1, target_col).value = f"Tháng {month}"
    cost.cell(1, target_col + 1).value = month
    return f"applied: copied month {prev_month} block to month {month}"


def apply_it_cost_month_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    cost = find_sheet(wb, SHEET_IT_COST)
    if cost is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_IT_COST}.")

    approvals = read_it_cost_month_approvals(approval_file)
    results: list[list[Any]] = []
    for approval in approvals:
        month = approval.get("month")
        issue = norm(approval.get("issue"))
        is_yes = approval_value_is_yes(approval.get("apply"))
        status = "skipped: Apply? not yes"
        resolved = "NO"
        resolution_note = None
        if is_yes:
            if month is None:
                status = "failed: missing month"
                resolution_note = "Set the target month in the approval workbook"
            elif "missing month" in issue or "missing formulas" in issue or "incomplete" in issue:
                status = copy_it_cost_month_block(cost, month)
                resolved = "YES" if status.startswith("applied:") or status.startswith("skipped:") else "NO"
                resolution_note = f"Month block for {month} handled"
            else:
                status = "skipped: unsupported issue"
                resolution_note = "Only missing month block issues are supported"
        results.append([
            approval.get("approval_row"),
            approval.get("sheet"),
            month,
            approval.get("issue"),
            approval.get("recommended_action"),
            approval.get("apply"),
            resolved,
            resolution_note,
            status,
        ])

    write_it_cost_month_result_sheet(wb, results)


def restore_carry_forward_sheets_from_approval_file(wb, approval_file: Path | None) -> None:
    if approval_file is None or not approval_file.exists():
        return

    try:
        source_wb = load_workbook(approval_file, data_only=False)
    except Exception:
        return

    sheet_names = [
        # Only carry forward the sheets that are meant to be edited/approved by the user.
        # Generated source/output sheets are rebuilt from the latest inputs on every run.
        "Check_IT_CPNS",
        "Check_IT_Cost_Month_Block",
        "IT_New_Project_Master",
        "Check_IT_Downstream",
        "Check_Media_Timesheet",
        "checkpoint data SX",
        "Check_SX_Downstream",
        "Check_Vonhoa_Month_Block",
    ]

    for sheet_name in sheet_names:
        source_ws = find_sheet(source_wb, sheet_name)
        if source_ws is None:
            continue
        target_ws = find_sheet(wb, sheet_name)
        if target_ws is None:
            target_ws = wb.create_sheet(title=sheet_name)
        copy_sheet_content(source_ws, target_ws)
        if sheet_name == "Check_IT_CPNS":
            # Keep the summary and header row fixed, but allow the details table to scroll normally.
            target_ws.freeze_panes = "A18"


def restore_missing_checkpoint_sheets_from_latest_output(
    wb,
    output_path: Path,
    sheet_names: list[str],
) -> list[str]:
    restored: list[str] = []
    missing_names = [name for name in sheet_names if name not in wb.sheetnames]
    if not missing_names:
        return restored

    final_dir = DIRS["final"]
    candidate_paths = [
        path
        for path in sorted(final_dir.glob("von_hoa_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if path.resolve() != output_path.resolve()
    ]

    for source_path in candidate_paths:
        try:
            source_wb = load_workbook(source_path, data_only=False)
        except Exception:
            continue
        for sheet_name in list(missing_names):
            if sheet_name not in source_wb.sheetnames:
                continue
            source_ws = source_wb[sheet_name]
            target_ws = wb.create_sheet(title=sheet_name)
            copy_sheet_content(source_ws, target_ws)
            if sheet_name == "Check_IT_CPNS":
                target_ws.freeze_panes = "A18"
            restored.append(sheet_name)
            missing_names.remove(sheet_name)
        if not missing_names:
            break

    return restored


def find_payroll_header_row(ws) -> int | None:
    for r in range(1, min(ws.max_row or 0, 20) + 1):
        values = {norm(ws.cell(r, c).value) for c in range(1, (ws.max_column or 0) + 1)}
        if "thang" in values and ("ma nhan vien" in values or "ma nhan su" in values):
            return r
    return None


def payroll_duplicate_keys(ws, header_row: int | None) -> int:
    if header_row is None:
        return 0
    headers = {norm(ws.cell(header_row, c).value): c for c in range(1, (ws.max_column or 0) + 1)}
    month_col = headers.get("thang")
    mnv_col = headers.get("ma nhan vien") or headers.get("ma nhan su")
    if not month_col or not mnv_col:
        return 0

    seen: set[tuple[Any, Any]] = set()
    duplicates = 0
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        month = clean(ws.cell(r, month_col).value)
        mnv = clean(ws.cell(r, mnv_col).value)
        if month is None and mnv is None:
            continue
        key = (month, mnv)
        if key in seen:
            duplicates += 1
        else:
            seen.add(key)
    return duplicates


def write_payroll_checkpoint_sheet(wb, rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Check_Payroll")
    append_table(
        ws,
        1,
        "Payroll Sync Check",
        [
            "Sheet",
            "Source path",
            "Source modified",
            "Source exists",
            "Source sheet exists",
            "Output sheet exists",
            "Rows copied",
            "Columns copied",
            "Header row",
            "Duplicate Month+MNV",
            "Status",
        ],
        rows,
    )
    ws.freeze_panes = "A3"
    for col, width in {"A": 32, "B": 90, "C": 22, "K": 60}.items():
        ws.column_dimensions[col].width = width


def sync_payroll_from_onedrive(wb, payroll_path: Path = PAYROLL_SOURCE_PATH) -> list[list[Any]]:
    source_exists = payroll_path.exists()
    modified = datetime.fromtimestamp(payroll_path.stat().st_mtime) if source_exists else None
    rows: list[list[Any]] = []

    if not source_exists:
        for sheet_name in PAYROLL_SHEETS:
            rows.append([sheet_name, str(payroll_path), None, False, False, find_sheet(wb, sheet_name) is not None, 0, 0, None, 0, "failed: payroll source file not found"])
        write_payroll_checkpoint_sheet(wb, rows)
        return rows

    temp_path = DIRS["staging"] / f".payroll_sync_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    try:
        shutil.copy2(payroll_path, temp_path)
        source_wb = load_workbook(temp_path, data_only=False)
    except PermissionError:
        for sheet_name in PAYROLL_SHEETS:
            rows.append([sheet_name, str(payroll_path), modified, True, False, find_sheet(wb, sheet_name) is not None, 0, 0, None, 0, "failed: payroll source file is locked or in use"])
        write_payroll_checkpoint_sheet(wb, rows)
        return rows
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        for sheet_name in PAYROLL_SHEETS:
            rows.append([sheet_name, str(payroll_path), modified, True, False, find_sheet(wb, sheet_name) is not None, 0, 0, None, 0, f"failed: payroll temp copy/open error: {exc}"])
        write_payroll_checkpoint_sheet(wb, rows)
        return rows

    try:
        for sheet_name in PAYROLL_SHEETS:
            source_ws = find_sheet(source_wb, sheet_name)
            target_ws = find_sheet(wb, sheet_name)
            source_sheet_exists = source_ws is not None
            output_sheet_exists = target_ws is not None
            copied_rows = 0
            copied_cols = 0
            header_row = None
            duplicate_count = 0
            status = "OK"

            if source_ws is None:
                status = "failed: source sheet missing"
            elif target_ws is None:
                status = "failed: output sheet missing"
            else:
                copy_sheet_content(source_ws, target_ws)
                copied_rows = source_ws.max_row or 0
                copied_cols = source_ws.max_column or 0
                header_row = find_payroll_header_row(source_ws)
                duplicate_count = payroll_duplicate_keys(source_ws, header_row)
                if header_row is None:
                    status = "warning: header row not detected"
                elif duplicate_count:
                    status = f"warning: {duplicate_count} duplicate Month+MNV rows"

            rows.append([sheet_name, str(payroll_path), modified, source_exists, source_sheet_exists, output_sheet_exists, copied_rows, copied_cols, header_row, duplicate_count, status])
    finally:
        temp_path.unlink(missing_ok=True)

    write_payroll_checkpoint_sheet(wb, rows)
    return rows


def write_approval_guide_sheet(wb, output_path: Path, approval_file: Path | None = None) -> None:
    ws = reset_sheet(wb, "Huong_dan_Approval")
    rows = [
        ["Bước", "Workbook approval", "Lệnh / Giá trị", "Hướng dẫn"],
        [1, str(output_path), "", "Mở workbook này và kiểm tra khu vực chung Khu_vuc_xu_ly_chung cùng các sheet Check_Payroll, Check_IT_CPNS, Check_IT_Cost_Month_Block, Check_Media_Timesheet, checkpoint data SX, Check_SX_Downstream và Check_Vonhoa_Month_Block. Bạn có thể tick YES trực tiếp ngay ở Khu_vuc_xu_ly_chung; khi rerun bằng --approval-file, các sheet approval/checkpoint do người dùng chỉnh sẽ được carry forward và đồng bộ về sheet gốc trước khi apply."],
        [2, str(output_path), "YES / Y / TRUE / 1 / APPLY / APPROVED", "Nhập một trong các giá trị này vào cột Apply? cho những dòng bạn muốn duyệt. Để trống hoặc nhập NO nếu không duyệt."],
        [3, str(output_path), "checkpoint data SX", "Hành động SX source: kiểm tra một sheet duy nhất gồm Summary / Details. Các dòng thiếu MNV sẽ nằm ngay trong Details với Issue = missing MNV; cập nhật sheet Mã nhân viên nếu cần, ghi chú ở cột Approval notes nếu cần, rồi đặt Apply? cho các dòng muốn mang qua file mới."],
        [4, str(output_path), "Check_SX_Downstream", "Hành động SX downstream: kiểm tra Summary / Details để soi dữ liệu rơi ở SX_Allocation_Build, Timesheet SX, 4.1 Chi phí nhân sự SX và 3.Vốn hóa. Nếu project chưa có trong 1.Danh mục dự án hoặc 3.Vốn hóa thì review theo Recommended action và đặt Apply? để cho phép tạo project mới / thêm vào vốn hóa khi cần."],
        [5, str(output_path), "Khu_vuc_xu_ly_chung", "Hành động chung: mở khu vực màu xám để jump nhanh sang Check_IT_CPNS, Check_IT_Cost_Month_Block, Check_Media_Timesheet, Check_SX_Downstream và Check_Vonhoa_Month_Block. Có thể duyệt trực tiếp tại đây, rồi rerun để script đẩy Apply? sang sheet gốc tương ứng."],
        [6, str(output_path), f"py -3 automate_kpi.py --sx-year 2026 --sx-month 4 --approval-file \"{output_path}\"", "Sau khi đánh dấu các dòng cần duyệt trong workbook này, chạy lệnh này và trỏ --approval-file về đúng workbook đã chỉnh. Nếu có duyệt SX, luôn truyền đúng --sx-month của kỳ đang xử lý; không bỏ trống để tránh dùng nhầm tháng mặc định."],
        [7, str(approval_file) if approval_file else "(none)", "", "Nếu lần chạy này có dùng approval file, đường dẫn sẽ hiện ở đây. Muốn đổi ý thì sửa file đó hoặc chạy lại không kèm --approval-file. Không dùng output cũ ngẫu nhiên làm nguồn carry forward nữa."],
        [8, str(output_path), "py -3 automate_kpi.py", "Chạy lại không có --approval-file để bỏ qua toàn bộ phê duyệt và tạo workbook mới từ template / nguồn dữ liệu. Đây là chế độ rebuild sạch, không carry forward state approved."],
        [9, str(output_path), "IT_Approval_Result / Project_Master_Approval_Result / Downstream_Approval_Result / Media_Approval_Result / IT_Cost_Block_Approval_Result / SX_Approval_Result / SX_Downstream_Approval_Result / Vonhoa_Block_Approval_Result", "Workbook đầu ra tiếp theo sẽ có các sheet kết quả này để ghi trạng thái applied / skipped / failed."],
    ]
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c).value = value

    command_start = len(rows) + 3
    command_rows = [
        ["Bước", "Khi nào dùng", "Lệnh", "Chức năng", "File cần sửa / dùng"],
        [
            1,
            "Tạo output từ các file raw local hiện có",
            "py -3 automate_kpi.py",
            "Dùng data/input/raw/IT.xlsx và data/input/raw/MEDIA.xlsx, ghi Timesheet IT / Data media, tạo các sheet checkpoint, rồi mở workbook output.",
            "Không cần approval file. Output mới nhất cũng được lưu trong data/output/final.",
        ],
        [
            2,
            "Tải file nguồn mới nhất rồi mới chạy",
            "py -3 automate_kpi.py --download",
            "Tải file IT / MEDIA từ config/sources.json, sau đó tạo output và checkpoint.",
            "Dùng khi Google Sheets / file nguồn đã đổi và cần làm mới raw local.",
        ],
        [
            3,
            "Xem checkpoint và chọn dòng cần sửa",
            "(sửa workbook, không cần lệnh)",
            "Mở Check_Payroll, checkpoint data SX và Check_SX_Downstream. Phần SX_Allocation_Build và Timesheet SX sẽ được rebuild tự động khi run. Chỉ đặt YES ở Apply? cho các dòng bạn đồng ý duyệt.",
            f"Sửa trực tiếp workbook approval ở cột B, thường là {output_path}.",
        ],
        [
            4,
            "Áp dụng các dòng đã duyệt",
            f'py -3 automate_kpi.py --sx-year 2026 --sx-month 4 --approval-file "{output_path}"',
            "Đọc các dòng YES từ workbook approval, áp dụng các thay đổi hợp lệ, rồi tạo workbook output mới.",
            "Đường dẫn --approval-file phải trỏ đúng workbook bạn đã chỉnh YES. Với SX, đổi --sx-month 4 thành đúng tháng đang approve.",
        ],
        [
            5,
            "Tải nguồn mới và áp dụng phê duyệt trong cùng một lượt",
            f'py -3 automate_kpi.py --download --sx-year 2026 --sx-month 4 --approval-file "{output_path}"',
            "Làm mới raw IT / MEDIA trước, sau đó áp dụng các phê duyệt từ workbook đã chọn.",
            "Dùng cẩn thận: approval lấy từ workbook cũ, còn dữ liệu được refresh từ nguồn mới nhất. Với SX, phải đổi --sx-month theo đúng kỳ approve.",
        ],
        [
            6,
            "Đổi ý và bỏ qua toàn bộ phê duyệt",
            "py -3 automate_kpi.py",
            "Tạo output mới mà không đọc bất kỳ dòng YES nào.",
            "Không truyền --approval-file.",
        ],
        [
            7,
            "Đổi ý với một số dòng cụ thể",
            "(sửa workbook, rồi chạy lại lệnh ở bước 4)",
            "Đổi Apply? từ YES sang NO hoặc để trống cho những dòng bạn không còn muốn duyệt, rồi chạy lại với --approval-file.",
            "Sửa lại chính workbook approval đó trước khi chạy lại.",
        ],
        [
            8,
            "Kiểm tra những gì đã được áp dụng",
            "(mở workbook output)",
            "Xem IT_Approval_Result, Project_Master_Approval_Result, Downstream_Approval_Result, Media_Approval_Result, SX_Approval_Result, SX_Downstream_Approval_Result và Vonhoa_Block_Approval_Result để kiểm tra trạng thái applied / skipped / failed.",
            "Dùng workbook output mới nhất được tạo sau khi chạy --approval-file.",
        ],
        [
            9,
            "Chạy mà không tự mở output",
            "py -3 automate_kpi.py --no-open",
            "Tạo output từ raw local nhưng không mở workbook kết quả.",
            "Dùng cho chạy nền / theo lịch.",
        ],
        [
            10,
            "Chủ động mở output sau khi chạy",
            "py -3 automate_kpi.py --open",
            "Tạo output từ raw local và mở workbook kết quả. Đây là hành vi mặc định.",
            "Không cần approval file trừ khi bạn muốn chạy kèm --approval-file.",
        ],
    ]
    for r_offset, row in enumerate(command_rows, start=command_start):
        for c, value in enumerate(row, start=1):
            ws.cell(r_offset, c).value = value

    process_start = command_start + len(command_rows) + 3
    process_rows = [
        ["Quy trình", "Sheet checkpoint", "Hành động cần duyệt", "Người dùng cần điền", "Sheet kết quả"],
        [
            "Payroll sync",
            "Check_Payroll",
            "Không có hành động duyệt",
            "Script đọc file OneDrive local và thay thế Lương nhân viên full time / Lương nhân viên part time trong output. Kiểm tra đường dẫn nguồn, thời gian chỉnh sửa, số dòng và cảnh báo trùng Month+MNV.",
            "Check_Payroll",
        ],
        [
            "SX thiếu MNV / duyệt source",
            "checkpoint data SX",
            "Giải quyết MNV và duyệt theo từng dòng",
            "Cập nhật sheet Mã nhân viên nếu thiếu MNV, đặt Apply? cho các dòng muốn mang sang file mới, và ghi chú nếu cần. Không có bảng Missing MNV riêng; mọi dòng nằm trong Details.",
            "SX_Approval_Result",
        ],
        [
            "SX downstream / duyệt downstream",
            "Check_SX_Downstream",
            "Giải quyết lỗi xuống dòng và duyệt theo từng dòng",
            "Sửa các điểm rơi ở SX_Allocation_Build, Timesheet SX, 4.1 Chi phí nhân sự SX và 3.Vốn hóa theo Recommended action. Nếu project chưa có trong danh mục hoặc vốn hóa thì review và approve để tạo project mới hoặc bổ sung vào vốn hóa.",
            "SX_Downstream_Approval_Result",
        ],
        [
            "Khu vực chung",
            "Khu_vuc_xu_ly_chung",
            "Gom các issue cần xử lý của IT / SX / Media / Vốn hóa",
            "Dùng sheet này để mở nhanh tất cả checkpoint quan trọng, đặc biệt là các block tháng và các dòng sửa whitespace / mapping / missing row đang nằm ở IT_CPNS, IT cost, Media và SX.",
            "Khu_vuc_xu_ly_chung",
        ],
        [
            "Vốn hóa / duyệt block tháng",
            "Check_Vonhoa_Month_Block",
            "Giải quyết block tháng bị thiếu trên 3.Vốn hóa",
            "Nếu tháng mới chưa có block, đặt Apply? = YES để copy block tháng trước sang. Đây là checkpoint cấu trúc, cần duyệt trước khi downstream chạy đúng.",
            "Vonhoa_Block_Approval_Result",
        ],
    ]
    for r_offset, row in enumerate(process_rows, start=process_start):
        for c, value in enumerate(row, start=1):
            ws.cell(r_offset, c).value = value

    ws.freeze_panes = "A3"
    for col, width in {"A": 12, "B": 54, "C": 28, "D": 72, "E": 60}.items():
        ws.column_dimensions[col].width = width


def write_common_processing_hub(wb) -> None:
    ws = reset_sheet(wb, "Khu_vuc_xu_ly_chung")
    summary_rows, queue_rows = build_common_processing_queue(wb)
    next_row = append_table(
        ws,
        1,
        "Tổng quan",
        ["Khu vực", "Sheet", "Mục đích", "Số issue", "Trạng thái", "Mở sheet"],
        summary_rows,
    )
    detail_title_row = next_row + 2
    append_table(
        ws,
        detail_title_row,
        "Danh sách xử lý",
        [
            "Khu vực",
            "Sheet nguồn",
            "Section",
            "Dòng nguồn",
            "Tháng",
            "Project / Stage",
            "Employee / MNV",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
            "Status",
            "Open",
        ],
        queue_rows,
    )
    for r in range(3, 3 + len(summary_rows)):
        sheet_name = ws.cell(r, 2).value
        if sheet_name in wb.sheetnames:
            ws.cell(r, 6).hyperlink = f"#'{sheet_name}'!A1"
            ws.cell(r, 6).value = "Mở"
    for r in range(detail_title_row + 2, detail_title_row + 2 + len(queue_rows)):
        sheet_name = ws.cell(r, 2).value
        if sheet_name in wb.sheetnames:
            ws.cell(r, 13).hyperlink = f"#'{sheet_name}'!A1"
            ws.cell(r, 13).value = "Mở"
    ws.freeze_panes = "A3"
    for col, width in {
        "A": 14,
        "B": 24,
        "C": 24,
        "D": 12,
        "E": 12,
        "F": 10,
        "G": 34,
        "H": 28,
        "I": 34,
        "J": 36,
        "K": 10,
        "L": 24,
        "M": 12,
    }.items():
        ws.column_dimensions[col].width = width


def build_common_processing_queue(wb) -> tuple[list[list[Any]], list[list[Any]]]:
    summary_rows: list[list[Any]] = []
    queue_rows: list[list[Any]] = []

    def add_link_row(area: str, sheet_name: str, purpose: str, issue_count: int) -> None:
        if sheet_name in wb.sheetnames:
            summary_rows.append([area, sheet_name, purpose, issue_count, "OPEN" if issue_count else "OK", "Mở"])
        else:
            summary_rows.append([area, sheet_name, purpose, issue_count, "MISSING SHEET", "Thiếu sheet"])

    def add_issue_rows(
        area: str,
        sheet_name: str,
        section_name: str,
        purpose: str,
        stop_markers: set[str],
    ) -> None:
        ws = find_sheet(wb, sheet_name)
        if ws is None:
            add_link_row(area, sheet_name, purpose, 0)
            return

        header_row = find_section_header(ws, section_name)
        if header_row is None:
            add_link_row(area, sheet_name, purpose, 0)
            return

        headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
        issue_col = headers.get("Issue")
        action_col = headers.get("Recommended action")
        apply_col = headers.get("Apply?")
        notes_col = headers.get("Approval notes")
        month_col = headers.get("Month")
        stage_col = headers.get("Stage")
        project_col = (
            headers.get("Project")
            or headers.get("Project name")
            or headers.get("Project code")
            or headers.get("Sheet")
        )
        employee_col = headers.get("Employee")
        mnv_col = headers.get("MNV")
        source_row_col = headers.get("Source row")

        issue_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            marker = ws.cell(r, 1).value
            if marker in stop_markers:
                break
            if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
                continue

            issue_value = clean(ws.cell(r, issue_col).value) if issue_col else None
            action_value = clean(ws.cell(r, action_col).value) if action_col else None
            if not issue_value and not action_value:
                continue

            issue_count += 1
            values = {
                "section": section_name,
                "source_row": ws.cell(r, source_row_col).value if source_row_col else r,
                "month": ws.cell(r, month_col).value if month_col else None,
                "project": ws.cell(r, project_col).value if project_col else None,
                "stage": ws.cell(r, stage_col).value if stage_col else None,
                "employee": ws.cell(r, employee_col).value if employee_col else None,
                "mnv": ws.cell(r, mnv_col).value if mnv_col else None,
                "issue": issue_value,
                "action": action_value,
                "apply": ws.cell(r, apply_col).value if apply_col else None,
                "notes": ws.cell(r, notes_col).value if notes_col else None,
            }

            project_text = values["project"] or values["stage"] or "-"
            employee_text = values["employee"] or values["mnv"] or "-"
            if values["section"] == "Source Issues":
                project_text = values["project"] or "-"

            detail_row = [
                area,
                sheet_name,
                values["section"],
                values["source_row"],
                values["month"],
                project_text,
                employee_text,
                values["issue"],
                values["action"],
                values["apply"],
                values["notes"],
                "READY" if approval_value_is_yes(values["apply"]) else "PENDING",
                "Mở",
            ]
            queue_rows.append(detail_row)

        add_link_row(area, sheet_name, purpose, issue_count)

    add_issue_rows(
        "IT",
        "Check_IT_CPNS",
        "Details",
        "Whitespace / mismatch giữa Timesheet IT và Chi phí nhân sự IT",
        {"Employee Month Over 100%", "Summary", "Details"},
    )
    add_issue_rows(
        "IT",
        "Check_IT_Cost_Month_Block",
        "Details",
        "Copy block tháng của Chi phí nhân sự IT",
        {"Summary", "Details"},
    )
    add_issue_rows(
        "IT",
        "Check_IT_Downstream",
        "Details",
        "Duyệt catalog / vốn hóa IT",
        {"Summary", "Details"},
    )
    add_issue_rows(
        "IT",
        "IT_New_Project_Master",
        "New Project Master Required",
        "Tạo project master và bổ sung cost row cho dự án IT mới",
        {"Summary", "New Project Master Required"},
    )
    add_issue_rows(
        "MEDIA",
        "Check_Media_Timesheet",
        "Details",
        "Kiểm tra dữ liệu Media theo tháng",
        {"Summary", "Details", "Source Issues", "Employee Month Over 100%"},
    )
    add_issue_rows(
        "MEDIA",
        "Check_Media_Timesheet",
        "Source Issues",
        "Sửa dòng nguồn Media bị thiếu công thức / weight",
        {"Summary", "Details", "Source Issues", "Employee Month Over 100%"},
    )
    add_issue_rows(
        "MEDIA",
        "Check_Media_Timesheet",
        "Employee Month Over 100%",
        "Sửa Media khi total work vượt 100%",
        {"Summary", "Details", "Source Issues", "Employee Month Over 100%"},
    )
    add_issue_rows(
        "SX",
        "Check_SX_Downstream",
        "Details",
        "Kiểm tra downstream SX và block tháng",
        {"Summary", "Details"},
    )
    add_issue_rows(
        "COMMON",
        "Check_Vonhoa_Month_Block",
        "Details",
        "Copy block tháng cho 3.Vốn hóa",
        {"Summary", "Details"},
    )

    return summary_rows, queue_rows


def sync_common_processing_hub_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None or not approval_file.exists():
        return

    try:
        source_wb = load_workbook(approval_file, data_only=False)
    except Exception:
        return

    source_ws = find_sheet(source_wb, "Khu_vuc_xu_ly_chung")
    if source_ws is None:
        return

    header_row = find_section_header(source_ws, "Danh sách xử lý")
    if header_row is None:
        return

    headers = {str(source_ws.cell(header_row, c).value): c for c in range(1, source_ws.max_column + 1) if source_ws.cell(header_row, c).value}
    required = ["Sheet nguồn", "Section", "Dòng nguồn", "Apply?"]
    if any(name not in headers for name in required):
        return

    for r in range(header_row + 1, source_ws.max_row + 1):
        marker = source_ws.cell(r, 1).value
        if marker in {"Summary", "Danh sách xử lý"}:
            break
        if not any(source_ws.cell(r, c).value is not None for c in range(1, source_ws.max_column + 1)):
            continue

        target_sheet_name = clean(source_ws.cell(r, headers["Sheet nguồn"]).value)
        section_name = clean(source_ws.cell(r, headers["Section"]).value)
        source_row = source_ws.cell(r, headers["Dòng nguồn"]).value
        if not target_sheet_name or source_row is None:
            continue

        try:
            source_row_int = int(source_row)
        except Exception:
            continue

        target_ws = find_sheet(wb, target_sheet_name)
        if target_ws is None:
            continue
        target_header_row = find_section_header(target_ws, section_name) if section_name else None
        if target_header_row is None:
            continue
        target_headers = {str(target_ws.cell(target_header_row, c).value): c for c in range(1, target_ws.max_column + 1) if target_ws.cell(target_header_row, c).value}
        apply_col = target_headers.get("Apply?")
        notes_col = target_headers.get("Approval notes")
        if apply_col is not None:
            target_ws.cell(source_row_int, apply_col).value = source_ws.cell(r, headers["Apply?"]).value
        if notes_col is not None and "Approval notes" in headers:
            target_ws.cell(source_row_int, notes_col).value = source_ws.cell(r, headers["Approval notes"]).value


def prepare_synced_approval_file(approval_file: Path | None) -> Path | None:
    if approval_file is None or not approval_file.exists():
        return approval_file

    try:
        wb = load_workbook(approval_file)
    except Exception:
        return approval_file

    source_ws = find_sheet(wb, "Khu_vuc_xu_ly_chung")
    if source_ws is None:
        return approval_file

    header_row = find_section_header(source_ws, "Danh sách xử lý")
    if header_row is None:
        return approval_file

    headers = {str(source_ws.cell(header_row, c).value): c for c in range(1, source_ws.max_column + 1) if source_ws.cell(header_row, c).value}
    required = ["Sheet nguồn", "Section", "Dòng nguồn", "Apply?"]
    if any(name not in headers for name in required):
        return approval_file

    for r in range(header_row + 1, source_ws.max_row + 1):
        marker = source_ws.cell(r, 1).value
        if marker in {"Summary", "Danh sách xử lý"}:
            break
        if not any(source_ws.cell(r, c).value is not None for c in range(1, source_ws.max_column + 1)):
            continue

        target_sheet_name = clean(source_ws.cell(r, headers["Sheet nguồn"]).value)
        section_name = clean(source_ws.cell(r, headers["Section"]).value)
        source_row = source_ws.cell(r, headers["Dòng nguồn"]).value
        if not target_sheet_name or source_row is None:
            continue
        try:
            source_row_int = int(source_row)
        except Exception:
            continue

        target_ws = find_sheet(wb, target_sheet_name)
        if target_ws is None:
            continue
        target_header_row = find_section_header(target_ws, section_name) if section_name else None
        if target_header_row is None:
            continue
        target_headers = {str(target_ws.cell(target_header_row, c).value): c for c in range(1, target_ws.max_column + 1) if target_ws.cell(target_header_row, c).value}
        apply_col = target_headers.get("Apply?")
        notes_col = target_headers.get("Approval notes")
        if apply_col is not None:
            target_ws.cell(source_row_int, apply_col).value = source_ws.cell(r, headers["Apply?"]).value
        if notes_col is not None and "Approval notes" in headers:
            target_ws.cell(source_row_int, notes_col).value = source_ws.cell(r, headers["Approval notes"]).value

    synced_path = approval_file.with_name(f".{approval_file.stem}.synced{approval_file.suffix}")
    try:
        wb.save(synced_path)
        return synced_path
    except Exception:
        return approval_file


def read_it_mapping_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "Check_IT_CPNS")
    if ws is None:
        return []

    header_row = find_section_header(ws, "Details")
    if header_row is None:
        raise SystemExit("Approval file does not contain Details section in Check_IT_CPNS.")

    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Month", "Project", "Employee", "Employee column", "Timesheet rows", "Issue", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in Details: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker in {"Employee Month Over 100%", "Summary", "Details"}:
            break
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue

        approvals.append(
            {
                "approval_row": r,
                "month": month_number(ws.cell(r, headers["Month"]).value),
                "project": ws.cell(r, headers["Project"]).value,
                "employee": ws.cell(r, headers["Employee"]).value,
                "employee_col": ws.cell(r, headers["Employee column"]).value,
                "timesheet_rows": ws.cell(r, headers["Timesheet rows"]).value,
                "cost_rows_after_clean": ws.cell(r, headers.get("Cost rows matched after clean", 0)).value if headers.get("Cost rows matched after clean") else None,
                "issue": ws.cell(r, headers["Issue"]).value,
                "action": ws.cell(r, headers["Recommended action"]).value,
            }
        )
    return approvals


def read_media_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "Check_Media_Timesheet")
    if ws is None:
        return []

    approvals: list[dict[str, Any]] = []
    for section in ["Details", "Source Issues"]:
        header_row = find_section_header(ws, section)
        if header_row is None:
            continue
        headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
        if "Apply?" not in headers or "Recommended action" not in headers:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            marker = ws.cell(r, 1).value
            if marker in {"Summary", "Details", "Source Issues", "Employee Month Over 100%"}:
                break
            if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
                continue
            if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
                continue

            approvals.append(
                {
                    "section": section,
                    "approval_row": r,
                    "data_row": ws.cell(r, headers.get("Data media row", 0)).value if headers.get("Data media row") else None,
                    "project": ws.cell(r, headers.get("Project", 0)).value if headers.get("Project") else None,
                    "month": month_number(ws.cell(r, headers.get("Month", 0)).value) if headers.get("Month") else None,
                    "employee": ws.cell(r, headers.get("Employee", 0)).value if headers.get("Employee") else None,
                    "mnv": ws.cell(r, headers.get("MNV", 0)).value if headers.get("MNV") else None,
                    "action": ws.cell(r, headers["Recommended action"]).value,
                }
            )
    return approvals


def read_downstream_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "Check_IT_Downstream")
    if ws is None:
        return []

    header_row = find_section_header(ws, "Details")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Project code", "Project name", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in Check_IT_Downstream: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker in {"Summary", "Details"}:
            break
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue
        approvals.append(
            {
                "approval_row": r,
                "project_code": ws.cell(r, headers["Project code"]).value,
                "project_name": ws.cell(r, headers["Project name"]).value,
                "system": ws.cell(r, headers.get("System", 0)).value if headers.get("System") else None,
                "catalog_year": ws.cell(r, headers.get("Catalog year", 0)).value if headers.get("Catalog year") else 2026,
                "catalog_bu": ws.cell(r, headers.get("Catalog BU", 0)).value if headers.get("Catalog BU") else None,
                "catalog_classification": ws.cell(r, headers.get("Catalog classification", 0)).value if headers.get("Catalog classification") else None,
                "catalog_start": ws.cell(r, headers.get("Catalog start date", 0)).value if headers.get("Catalog start date") else None,
                "catalog_end": ws.cell(r, headers.get("Catalog end date", 0)).value if headers.get("Catalog end date") else None,
                "catalog_capitalization": ws.cell(r, headers.get("Catalog capitalization flag", 0)).value if headers.get("Catalog capitalization flag") else True,
                "action": ws.cell(r, headers["Recommended action"]).value,
            }
        )
    return approvals


def read_project_master_approvals(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False)
    ws = find_sheet(wb, "IT_New_Project_Master")
    if ws is None:
        return []
    if ws is None:
        return []

    header_row = find_section_header(ws, "New Project Master Required")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Project name", "Suggested project code", "Suggested BU", "Recommended action", "Apply?"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise SystemExit(f"Approval file is missing required columns in IT_New_Project_Master: {missing}")

    approvals: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            continue
        if not approval_value_is_yes(ws.cell(r, headers["Apply?"]).value):
            continue
        approvals.append(
            {
                "approval_row": r,
                "project": ws.cell(r, headers["Project name"]).value,
                "system": ws.cell(r, headers.get("System", 0)).value if headers.get("System") else None,
                "project_code": ws.cell(r, headers["Suggested project code"]).value,
                "bu": ws.cell(r, headers["Suggested BU"]).value,
                "start_month": month_number(ws.cell(r, headers.get("Start month", 0)).value) if headers.get("Start month") else None,
                "end_month": month_number(ws.cell(r, headers.get("End month", 0)).value) if headers.get("End month") else None,
                "timesheet_rows": ws.cell(r, headers.get("Timesheet rows", 0)).value if headers.get("Timesheet rows") else None,
                "action": ws.cell(r, headers["Recommended action"]).value,
            }
        )
    return approvals


def read_project_master_result_projects(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise SystemExit(f"Approval file not found: {path}")

    wb = load_workbook(path, data_only=False, read_only=True)
    ws = find_sheet(wb, "Project_Master_Approval_Result")
    if ws is None:
        return []

    header_row = find_section_header(ws, "Project Master Approval Result")
    if header_row is None:
        return []
    headers = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    required = ["Action", "Project code", "Project name", "Status"]
    if any(name not in headers for name in required):
        return []

    projects: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for r in range(header_row + 1, ws.max_row + 1):
        action = clean(ws.cell(r, headers["Action"]).value)
        code = clean(ws.cell(r, headers["Project code"]).value)
        project = clean(ws.cell(r, headers["Project name"]).value)
        status = clean(ws.cell(r, headers["Status"]).value)
        if action != "ADD_PROJECT_MASTER_FIRST" or not code or not project:
            continue
        if not status or not str(status).startswith("applied:"):
            continue
        key = (str(code), str(project))
        if key in seen:
            continue
        seen.add(key)
        projects.append({"project_code": code, "project": project})
    return projects


def copy_row(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for c in range(1, ws.max_column + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value


def copy_row_between_sheets(source_ws, source_row: int, target_ws, target_row: int) -> None:
    target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
    for c in range(1, source_ws.max_column + 1):
        src = source_ws.cell(source_row, c)
        dst = target_ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value


def parse_int_list(value: Any) -> list[int]:
    if value is None:
        return []
    return [int(item) for item in re.findall(r"\d+", str(value))]


def column_letters_to_number(value: str) -> int:
    from openpyxl.utils.cell import column_index_from_string

    return column_index_from_string(value)


def column_from_it_sumifs(value: Any) -> int | None:
    if not isinstance(value, str) or not value.startswith("="):
        return None
    match = re.search(r"'Timesheet IT'!\$([A-Z]+):\$\\1", value)
    if not match:
        match = re.search(r"'Timesheet IT'!\$([A-Z]+):\$[A-Z]+", value)
    if not match:
        return None
    letters = match.group(1)
    try:
        from openpyxl.utils.cell import column_index_from_string
        return column_index_from_string(letters)
    except Exception:
        return None


def find_cost_row_by_project_and_sum_col(cost, project: Any, sum_col: int) -> int | None:
    for r in range(3, cost.max_row + 1):
        if clean(cost.cell(r, 2).value) != clean(project):
            continue
        for c in range(1, cost.max_column + 1):
            if column_from_it_sumifs(cost.cell(r, c).value) == sum_col:
                return r
    return None


def find_cost_project_sample_row(cost, project: Any) -> int | None:
    for r in range(3, cost.max_row + 1):
        if clean(cost.cell(r, 2).value) == clean(project):
            return r
    return None


def find_cost_employee_template_row(cost, sum_col: int) -> int | None:
    for r in range(3, cost.max_row + 1):
        for c in range(1, cost.max_column + 1):
            if column_from_it_sumifs(cost.cell(r, c).value) == sum_col:
                return r
    return None


def find_cost_percent_col(cost, month: int) -> int | None:
    for c in range(1, cost.max_column + 1):
        if month_number(cost.cell(1, c).value) == month and norm(cost.cell(2, c).value) == "cong viec":
            return c
    return None


def fill_cost_month_formulas(cost, row: int, month: int, sum_col: int) -> str:
    percent_col = find_cost_percent_col(cost, month)
    if percent_col is None:
        return f"failed: no % công việc column for month {month}"

    sum_letter = get_column_letter(sum_col)
    percent_letter = get_column_letter(percent_col)
    salary_letter = get_column_letter(percent_col + 1)
    cost_letter = get_column_letter(percent_col + 2)
    changes = []

    percent_formula = (
        f"=SUMIFS('Timesheet IT'!${sum_letter}:${sum_letter},"
        f"'Timesheet IT'!$B:$B,'Chi phí nhân sự IT'!$B{row},"
        f"'Timesheet IT'!$E:$E,'Chi phí nhân sự IT'!{percent_letter}$1)"
    )
    salary_formula = (
        f"=SUMIFS('Lương nhân viên full time'!$AF:$AF,"
        f"'Lương nhân viên full time'!$E:$E,'Chi phí nhân sự IT'!$G{row},"
        f"'Lương nhân viên full time'!$A:$A,'Chi phí nhân sự IT'!{percent_letter}$1)"
    )
    cost_formula = f"={percent_letter}{row}*{salary_letter}{row}"

    targets = [
        (percent_col, percent_formula, percent_letter),
        (percent_col + 1, salary_formula, salary_letter),
        (percent_col + 2, cost_formula, cost_letter),
    ]
    for col, formula, letter in targets:
        cell = cost.cell(row, col)
        if cell.value in (None, ""):
            cell.value = formula
            changes.append(letter)

    return f"applied: filled {', '.join(changes)} row {row}" if changes else f"skipped: month {month} formulas already exist in row {row}"


def apply_clean_cost_project_text(cost, approval: dict[str, Any]) -> str:
    rows = parse_int_list(approval.get("cost_rows_after_clean")) or parse_int_list(approval.get("cost_rows"))
    if not rows:
        return "skipped: no cost row listed"
    changed = []
    for r in rows:
        current = cost.cell(r, 2).value
        cleaned = clean(current)
        if current != cleaned:
            cost.cell(r, 2).value = cleaned
            changed.append(r)
    return f"applied: cleaned rows {changed}" if changed else f"skipped: already clean rows {rows}"


def apply_add_cost_row(cost, approval: dict[str, Any]) -> str:
    employee_col = approval.get("employee_col")
    if not employee_col:
        return "failed: missing Employee column"
    sum_col = column_letters_to_number(str(employee_col))
    project = approval.get("project")
    month = approval.get("month")

    existing = find_cost_row_by_project_and_sum_col(cost, project, sum_col)
    if existing:
        if month:
            return fill_cost_month_formulas(cost, existing, month, sum_col)
        return f"skipped: existing cost row {existing}"

    project_row = find_cost_project_sample_row(cost, project)
    if project_row is None:
        return "failed: no existing project row to copy project metadata"

    template_row = find_cost_employee_template_row(cost, sum_col)
    if template_row is None:
        return "failed: no employee template row for Timesheet IT column"

    target_row = cost.max_row + 1
    copy_row(cost, template_row, target_row)
    for c in range(1, 7):
        cost.cell(target_row, c).value = clean(cost.cell(project_row, c).value)
    cost.cell(target_row, 2).value = clean(project)
    fill_status = fill_cost_month_formulas(cost, target_row, month, sum_col) if month else "skipped: no month"
    return f"applied: added cost row {target_row} from employee template row {template_row} and project row {project_row}; {fill_status}"


def apply_fill_cost_month_formula(cost, approval: dict[str, Any]) -> str:
    employee_col = approval.get("employee_col")
    month = approval.get("month")
    project = approval.get("project")
    if not employee_col or not month:
        return "failed: missing Employee column or Month"
    sum_col = column_letters_to_number(str(employee_col))
    existing = find_cost_row_by_project_and_sum_col(cost, project, sum_col)
    if not existing:
        return "failed: no existing cost row"
    return fill_cost_month_formulas(cost, existing, month, sum_col)


def write_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "IT_Approval_Result")
    headers = ["Approval row", "Action", "Project", "Employee", "Employee column", "Status"]
    append_table(ws, 1, "IT Approval Result", headers, results)
    for col, width in {"A": 14, "B": 28, "C": 60, "D": 24, "E": 16, "F": 80}.items():
        ws.column_dimensions[col].width = width


def write_media_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Media_Approval_Result")
    headers = ["Approval row", "Section", "Action", "Data media row", "Project", "Employee", "MNV", "Status"]
    append_table(ws, 1, "Media Approval Result", headers, results)
    for col, width in {"A": 14, "B": 16, "C": 28, "D": 16, "E": 42, "F": 24, "G": 16, "H": 80}.items():
        ws.column_dimensions[col].width = width


def build_media_checkpoint_data(
    wb,
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], list[list[Any]]]:
    data_ws = find_sheet(wb, SHEET_MEDIA)
    ts_ws = find_sheet(wb, SHEET_TIMESHEET_MEDIA)
    fulltime_ws = find_sheet(wb, SHEET_SALARY_FULLTIME)
    parttime_ws = find_sheet(wb, SHEET_SALARY_PARTTIME)
    if data_ws is None or ts_ws is None or fulltime_ws is None:
        summary_rows = [[month, 0, 0, 0] for month in range(1, 13)]
        placeholder = [[None] * 15]
        return summary_rows, placeholder, placeholder, []

    salary_lookup: dict[tuple[int, str], Any] = {}
    for r in range(3, (fulltime_ws.max_row or 0) + 1):
        month = month_number(fulltime_ws.cell(r, 1).value)
        employee_code = clean(fulltime_ws.cell(r, 5).value)
        n_hours = parse_hours(fulltime_ws.cell(r, 14).value) or 0
        t_hours = parse_hours(fulltime_ws.cell(r, 20).value) or 0
        standard_hours = n_hours * 8 + t_hours
        if month is None or not employee_code:
            continue
        salary_lookup[(month, norm(employee_code))] = standard_hours

    if parttime_ws is not None:
        for r in range(3, (parttime_ws.max_row or 0) + 1):
            month = month_number(parttime_ws.cell(r, 1).value)
            employee_code = clean(parttime_ws.cell(r, 3).value)
            standard_hours = parse_hours(parttime_ws.cell(r, 27).value)
            if month is None or not employee_code or not standard_hours:
                continue
            salary_lookup.setdefault((month, norm(employee_code)), standard_hours)

    ts_exact_map: dict[tuple[str, str, str, str], list[int]] = defaultdict(list)
    ts_clean_map: dict[tuple[str, str, str, str], list[int]] = defaultdict(list)
    ts_employee_rows: dict[str, list[int]] = defaultdict(list)
    for r in range(4, (ts_ws.max_row or 0) + 1):
        bu = clean(ts_ws.cell(r, 2).value)
        project = clean(ts_ws.cell(r, 4).value)
        employee = clean(ts_ws.cell(r, 7).value)
        mnv = clean(ts_ws.cell(r, 6).value)
        if not any([bu, project, employee, mnv]):
            continue
        exact_key = (str(bu or ""), str(project or ""), str(employee or ""), str(mnv or ""))
        clean_key = (norm(bu), norm(project), norm(employee), norm(mnv))
        ts_exact_map[exact_key].append(r)
        ts_clean_map[clean_key].append(r)
        if mnv:
            ts_employee_rows[norm(mnv)].append(r)

    monthly_groups: dict[tuple[int, str, str, str, str], dict[str, Any]] = {}
    source_issue_rows: list[list[Any]] = []
    for r in range(MEDIA_TEMPLATE_ROW, (data_ws.max_row or 0) + 1):
        bu = clean(data_ws.cell(r, 1).value)
        project = clean(data_ws.cell(r, 2).value)
        task = clean(data_ws.cell(r, 3).value)
        hours = parse_hours(data_ws.cell(r, 4).value)
        month = month_number(data_ws.cell(r, 5).value) or parse_month_value(data_ws.cell(r, 5).value)
        employee = clean(data_ws.cell(r, 6).value)
        mnv = clean(data_ws.cell(r, 7).value)
        if not any([bu, project, task, hours, month, employee, mnv]):
            continue
        if month is None:
            continue

        standard_hours = salary_lookup.get((month, norm(mnv or "")))
        if not isinstance(standard_hours, (int, float)) or not standard_hours:
            source_issue_rows.append(
                [
                    r,
                    project,
                    month,
                    employee,
                    mnv,
                    hours,
                    f"=D{r}/H{r}",
                    "missing salary standard hours for media employee/month",
                    "FIX_MEDIA_WEIGHT_OR_MONTH",
                    "Y",
                    "Y",
                ]
            )

        group_key = (month, norm(bu), norm(project), norm(employee), norm(mnv))
        group = monthly_groups.setdefault(
            group_key,
            {
                "month": month,
                "bu": bu,
                "project": project,
                "employee": employee,
                "mnv": mnv,
                "data_rows": [],
                "hours_total": 0.0,
                "standard_hours": standard_hours,
            },
        )
        group["data_rows"].append(r)
        if isinstance(hours, (int, float)):
            group["hours_total"] += float(hours)
        if not isinstance(group.get("standard_hours"), (int, float)) and isinstance(standard_hours, (int, float)):
            group["standard_hours"] = standard_hours

    summary_rows: list[list[Any]] = []
    for month in range(1, 13):
        row_num = month + 2
        summary_rows.append(
            [
                month,
                f"=SUMIF('{SHEET_MEDIA}'!$E:$E,A{row_num},'{SHEET_MEDIA}'!$I:$I)",
                f"=SUM(INDEX('{SHEET_TIMESHEET_MEDIA}'!$I:$S,0,A{row_num}))",
                f"=C{row_num}-B{row_num}",
            ]
        )

    detail_rows: list[list[Any]] = []
    over100_rows: list[list[Any]] = []
    for group in monthly_groups.values():
        exact_key = (str(group["bu"] or ""), str(group["project"] or ""), str(group["employee"] or ""), str(group["mnv"] or ""))
        clean_key = (norm(group["bu"]), norm(group["project"]), norm(group["employee"]), norm(group["mnv"]))
        exact_rows = ts_exact_map.get(exact_key, [])
        clean_rows = ts_clean_map.get(clean_key, [])
        matched_exact = bool(exact_rows)
        matched_after_clean = bool(clean_rows)
        standard_hours = group.get("standard_hours")
        input_value = float(group["hours_total"]) / float(standard_hours) if isinstance(standard_hours, (int, float)) and standard_hours else 0

        if not matched_after_clean:
            detail_rows.append(
                [
                    group["month"],
                    group["project"],
                    group["employee"],
                    group["mnv"],
                    ", ".join(str(r) for r in group["data_rows"]),
                    input_value,
                    0,
                    -input_value,
                    ", ".join(str(r) for r in exact_rows) if exact_rows else None,
                    ", ".join(str(r) for r in clean_rows) if clean_rows else None,
                    matched_after_clean,
                    "missing row/formula in Timesheet Media",
                    "ADD_TIMESHEET_MEDIA_ROW",
                    "Y",
                    "Y",
                ]
            )
        elif not matched_exact:
            detail_rows.append(
                [
                    group["month"],
                    group["project"],
                    group["employee"],
                    group["mnv"],
                    ", ".join(str(r) for r in group["data_rows"]),
                    input_value,
                    0,
                    -input_value,
                    ", ".join(str(r) for r in exact_rows) if exact_rows else None,
                    ", ".join(str(r) for r in clean_rows) if clean_rows else None,
                    matched_after_clean,
                    "criteria text mismatch, likely whitespace",
                    "CLEAN_TIMESHEET_MEDIA_TEXT",
                    "Y",
                    "Y",
                ]
            )

        if isinstance(input_value, (int, float)) and input_value > 1:
            ts_rows = ts_employee_rows.get(norm(group["mnv"] or ""), [])
            over100_rows.append(
                [
                    group["month"],
                    group["mnv"],
                    group["employee"],
                    input_value,
                    input_value - 1,
                    ", ".join(str(r) for r in ts_rows) if ts_rows else None,
                    "employee month over 100%",
                    "REDUCE_MEDIA_ALLOCATION_OR_FIX_PAYROLL_HOURS",
                    "Y",
                    "Y",
                ]
            )

    if not detail_rows:
        detail_rows.append([None] * 15)
    if not source_issue_rows:
        source_issue_rows.append([None] * 11)

    return summary_rows, detail_rows, source_issue_rows, over100_rows


def write_media_checkpoint_sheet(
    wb,
    summary_rows: list[list[Any]],
    detail_rows: list[list[Any]],
    source_issue_rows: list[list[Any]],
    over100_rows: list[list[Any]],
) -> None:
    ws = reset_sheet(wb, "Check_Media_Timesheet")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Month", "Input Data media", "Processed Timesheet Media", "Diff"],
        summary_rows,
    )
    next_row = append_table(
        ws,
        next_row + 1,
        "Details",
        [
            "Month",
            "Project",
            "Employee",
            "MNV",
            "Data media rows",
            "Input value",
            "Processed value",
            "Diff",
            "Timesheet Media rows matched exact",
            "Timesheet Media rows matched after clean",
            "Matched after clean",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        detail_rows,
    )
    next_row = append_table(
        ws,
        next_row + 1,
        "Source Issues",
        [
            "Data media row",
            "Project",
            "Month",
            "Employee",
            "MNV",
            "Hours",
            "Weight value/formula",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        source_issue_rows,
    )
    append_table(
        ws,
        next_row + 1,
        "Employee Month Over 100%",
        [
            "Month",
            "MNV",
            "Employee",
            "Total % work",
            "Over by",
            "Timesheet Media rows",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        over100_rows,
    )
    ws.freeze_panes = "A17"
    for col, width in {
        "A": 10,
        "B": 36,
        "C": 24,
        "D": 16,
        "E": 18,
        "F": 16,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 24,
        "K": 16,
        "L": 34,
        "M": 30,
        "N": 12,
        "O": 18,
    }.items():
        ws.column_dimensions[col].width = width


def write_downstream_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Downstream_Approval_Result")
    headers = ["Approval row", "Action", "Project code", "Project name", "Status"]
    append_table(ws, 1, "Downstream Approval Result", headers, results)
    for col, width in {"A": 14, "B": 34, "C": 18, "D": 60, "E": 90}.items():
        ws.column_dimensions[col].width = width


def write_project_master_approval_result_sheet(wb, results: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Project_Master_Approval_Result")
    headers = ["Approval row", "Action", "Project code", "Project name", "Status"]
    append_table(ws, 1, "Project Master Approval Result", headers, results)
    for col, width in {"A": 14, "B": 30, "C": 18, "D": 60, "E": 110}.items():
        ws.column_dimensions[col].width = width


def first_empty_row(ws, key_cols: list[int], start_row: int = 2) -> int:
    for r in range(start_row, ws.max_row + 1):
        if all(clean(ws.cell(r, c).value) is None for c in key_cols):
            return r
    return ws.max_row + 1


def bool_value(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = norm(value)
    if text in {"yes", "y", "true", "1", "co", "x"}:
        return True
    if text in {"no", "n", "false", "0", "khong"}:
        return False
    return default


def approval_value_is_yes(value: Any) -> bool:
    return bool_value(value, default=False)


def apply_add_to_project_catalog(catalog, approval: dict[str, Any]) -> str:
    code = clean(approval.get("project_code"))
    project = clean(approval.get("project_name"))
    if not code or not project:
        return "failed: missing project code or project name"
    if project_exists_in_column(catalog, code, 2, 2):
        return f"skipped: project code {code} already exists in catalog"

    target_row = first_empty_row(catalog, [2, 3], start_row=2)
    ensure_row_style(catalog, 2, target_row, catalog.max_column)
    catalog.cell(target_row, 1).value = approval.get("catalog_year") or 2026
    catalog.cell(target_row, 2).value = code
    catalog.cell(target_row, 3).value = project
    catalog.cell(target_row, 4).value = approval.get("system")
    catalog.cell(target_row, 5).value = approval.get("catalog_bu") or "SAPP"
    catalog.cell(target_row, 6).value = approval.get("catalog_classification")
    catalog.cell(target_row, 7).value = approval.get("catalog_start")
    catalog.cell(target_row, 8).value = approval.get("catalog_end")
    catalog.cell(target_row, 12).value = bool_value(approval.get("catalog_capitalization"), default=True)
    return f"applied: added catalog row {target_row}"


def find_capitalization_template_row(capitalization) -> int | None:
    for r in range(3, capitalization.max_row + 1):
        code = clean(capitalization.cell(r, 2).value)
        if code and str(code).startswith("IT"):
            return r
    return 3 if capitalization.max_row >= 3 else None


def apply_add_to_capitalization(capitalization, approval: dict[str, Any]) -> str:
    code = clean(approval.get("project_code"))
    project = clean(approval.get("project_name"))
    if not code or not project:
        return "failed: missing project code or project name"
    if project_exists_in_column(capitalization, code, 2, 3):
        return f"skipped: project code {code} already exists in 3.Von hoa"

    template_row = find_capitalization_template_row(capitalization)
    if template_row is None:
        return "failed: no capitalization template row found"

    target_row = first_empty_row(capitalization, [2, 3], start_row=3)
    copy_row(capitalization, template_row, target_row)
    capitalization.cell(target_row, 1).value = approval.get("catalog_year") or 2026
    capitalization.cell(target_row, 2).value = code
    capitalization.cell(target_row, 3).value = project
    capitalization.cell(target_row, 4).value = approval.get("catalog_bu") or "HO"
    capitalization.cell(target_row, 5).value = approval.get("catalog_classification")
    capitalization.cell(target_row, 6).value = approval.get("catalog_start")
    capitalization.cell(target_row, 7).value = approval.get("catalog_end")
    capitalization.cell(target_row, 8).value = None
    return f"applied: added 3.Von hoa row {target_row} from template row {template_row}"


def source_entries_for_project(ts, project: Any) -> dict[int, dict[str, Any]]:
    entries: dict[int, dict[str, Any]] = {}
    target = clean(project)
    for r in range(IT_HEADER_ROW + 2, ts.max_row + 1):
        if clean(ts.cell(r, 2).value) != target:
            continue
        month = month_number(ts.cell(r, 5).value)
        if month is None:
            continue
        for c in range(6, min(ts.max_column, 26) + 1):
            value = numeric(ts.cell(r, c).value)
            if not value:
                continue
            item = entries.setdefault(c, {"months": set(), "rows": [], "total": 0.0})
            item["months"].add(month)
            item["rows"].append(r)
            item["total"] += value
    return entries


def _project_code_number(value: Any) -> int | None:
    code = clean(value)
    if not code:
        return None
    compact = re.sub(r"\s+", "", str(code)).upper()
    match = re.fullmatch(r"IT0*(\d+)", compact)
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def next_it_project_code_number(wb) -> int:
    max_code = 0
    for sheet_name, col in [
        (SHEET_PROJECT_CATALOG, 2),
        (SHEET_CAPITALIZATION, 2),
        (SHEET_IT_COST, 1),
    ]:
        ws = find_sheet(wb, sheet_name)
        if ws is None:
            continue
        for r in range(2, ws.max_row + 1):
            num = _project_code_number(ws.cell(r, col).value)
            if num is not None:
                max_code = max(max_code, num)
    return max_code + 1 if max_code else 2601


def collect_timesheet_it_projects_missing_from_catalog(ts, catalog_projects: set[str]) -> dict[str, dict[str, Any]]:
    projects: dict[str, dict[str, Any]] = {}
    for r in range(IT_HEADER_ROW + 2, ts.max_row + 1):
        project_name = clean(ts.cell(r, 2).value)
        project_clean = norm(project_name)
        if not project_clean or project_clean in catalog_projects:
            continue

        item = projects.setdefault(
            project_clean,
            {
                "project_name": project_name,
                "system": clean(ts.cell(r, 3).value),
                "bu": clean(ts.cell(r, 4).value),
                "ts_rows": [],
                "months": set(),
            },
        )
        if item.get("system") is None:
            item["system"] = clean(ts.cell(r, 3).value)
        if item.get("bu") is None:
            item["bu"] = clean(ts.cell(r, 4).value)
        month = month_number(ts.cell(r, 5).value)
        if month is not None:
            item["months"].add(month)
        item["ts_rows"].append(r)
    return projects


def project_master_to_catalog_approval(approval: dict[str, Any]) -> dict[str, Any]:
    return {
        "project_code": approval.get("project_code"),
        "project_name": approval.get("project"),
        "system": approval.get("system"),
        "catalog_year": 2026,
        "catalog_bu": approval.get("bu") or "SAPP",
        "catalog_classification": None,
        "catalog_start": None,
        "catalog_end": None,
        "catalog_capitalization": True,
    }


def project_master_to_capitalization_approval(approval: dict[str, Any]) -> dict[str, Any]:
    start_month = approval.get("start_month")
    end_month = approval.get("end_month")
    start_date = datetime(2026, int(start_month), 1) if start_month else None
    end_date = datetime(2026, int(end_month), 1) if end_month else None
    return {
        "project_code": approval.get("project_code"),
        "project_name": approval.get("project"),
        "catalog_year": 2026,
        "catalog_bu": "HO",
        "catalog_classification": approval.get("bu") or None,
        "catalog_start": start_date,
        "catalog_end": end_date,
        "catalog_capitalization": True,
    }


def apply_project_master_approval(wb, approval: dict[str, Any]) -> str:
    action = clean(approval.get("action"))
    if action != "ADD_PROJECT_MASTER_FIRST":
        return f"skipped: unsupported action {action}"

    ts = find_sheet(wb, SHEET_IT)
    cost = find_sheet(wb, SHEET_IT_COST)
    catalog = find_sheet(wb, SHEET_PROJECT_CATALOG)
    capitalization = find_sheet(wb, SHEET_CAPITALIZATION)
    if ts is None or cost is None or catalog is None or capitalization is None:
        return "failed: missing Timesheet IT, Chi phí nhân sự IT, or 1.Danh mục dự án"

    project = clean(approval.get("project"))
    code = clean(approval.get("project_code"))
    if not project or not code:
        return "failed: missing project name or project code"

    statuses = [apply_add_to_project_catalog(catalog, project_master_to_catalog_approval(approval))]
    source_entries = source_entries_for_project(ts, project)
    if not source_entries:
        statuses.append("failed: no Timesheet IT source entries found")
        return "; ".join(statuses)

    added_rows = []
    skipped_rows = []
    for employee_col, source in sorted(source_entries.items()):
        existing = find_cost_row_by_project_and_sum_col(cost, project, employee_col)
        if existing:
            for month in sorted(source["months"]):
                fill_cost_month_formulas(cost, existing, month, employee_col)
            skipped_rows.append(str(existing))
            continue

        template_row = find_cost_employee_template_row(cost, employee_col)
        if template_row is None:
            statuses.append(f"failed: no employee template for Timesheet IT column {get_column_letter(employee_col)}")
            continue

        target_row = cost.max_row + 1
        copy_row(cost, template_row, target_row)
        cost.cell(target_row, 1).value = code
        cost.cell(target_row, 2).value = project
        cost.cell(target_row, 3).value = approval.get("system")
        cost.cell(target_row, 4).value = None
        cost.cell(target_row, 5).value = None
        cost.cell(target_row, 6).value = approval.get("bu") or "SAPP"
        for month in sorted(source["months"]):
            fill_cost_month_formulas(cost, target_row, month, employee_col)
        added_rows.append(str(target_row))

    if added_rows:
        statuses.append(f"added cost rows {', '.join(added_rows)}")
    if skipped_rows:
        statuses.append(f"updated existing cost rows {', '.join(skipped_rows)}")
    statuses.append(apply_add_to_capitalization(capitalization, project_master_to_capitalization_approval(approval)))
    return "; ".join(statuses)


def apply_it_mapping_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    cost = find_sheet(wb, SHEET_IT_COST)
    if cost is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_IT_COST}.")

    approvals = read_it_mapping_approvals(approval_file)
    results: list[list[Any]] = []
    for approval in approvals:
        action = clean(approval.get("action"))
        try:
            if action == "CLEAN_COST_PROJECT_TEXT":
                status = apply_clean_cost_project_text(cost, approval)
            elif action == "ADD_COST_ROW":
                status = apply_add_cost_row(cost, approval)
            elif action == "FILL_COST_MONTH_FORMULA":
                status = apply_fill_cost_month_formula(cost, approval)
            elif action == "ADD_PROJECT_MASTER_FIRST":
                status = "skipped: approve this action in IT_New_Project_Master"
            else:
                status = f"skipped: unsupported action {action}"
        except Exception as exc:
            status = f"failed: {exc}"
        results.append(
            [
                approval.get("approval_row"),
                action,
                approval.get("project"),
                approval.get("employee"),
                approval.get("employee_col"),
                status,
            ]
        )

    write_approval_result_sheet(wb, results)


def find_row_by_clean_value(ws, col: int, value: Any, start_row: int = 1) -> int | None:
    target = clean(value)
    if target is None:
        return None
    for r in range(start_row, ws.max_row + 1):
        if clean(ws.cell(r, col).value) == target:
            return r
    return None


def project_exists_in_column(ws, value: Any, col: int, start_row: int = 1) -> bool:
    return find_row_by_clean_value(ws, col, value, start_row=start_row) is not None


def carry_forward_project_master_results(wb, approval_file: Path | None) -> list[list[Any]]:
    if approval_file is None:
        return []

    projects = read_project_master_result_projects(approval_file)
    if not projects:
        return []

    source_wb = load_workbook(approval_file, data_only=False)
    source_cost = find_sheet(source_wb, SHEET_IT_COST)
    source_catalog = find_sheet(source_wb, SHEET_PROJECT_CATALOG)
    target_cost = find_sheet(wb, SHEET_IT_COST)
    target_catalog = find_sheet(wb, SHEET_PROJECT_CATALOG)
    if source_cost is None or source_catalog is None or target_cost is None or target_catalog is None:
        return [[None, "CARRY_FORWARD_PROJECT_MASTER", None, None, "failed: missing source/target catalog or cost sheet"]]

    results: list[list[Any]] = []
    for project in projects:
        code = project["project_code"]
        project_name = project["project"]
        statuses = []

        if find_row_by_clean_value(target_catalog, 2, code, start_row=2):
            statuses.append(f"skipped catalog: project code {code} already exists")
        else:
            source_catalog_row = find_row_by_clean_value(source_catalog, 2, code, start_row=2)
            if source_catalog_row is None:
                statuses.append("failed catalog: approved project not found in approval file catalog")
            else:
                target_catalog_row = first_empty_row(target_catalog, [2, 3], start_row=2)
                copy_row_between_sheets(source_catalog, source_catalog_row, target_catalog, target_catalog_row)
                statuses.append(f"carried catalog row {target_catalog_row}")

        existing_cost_rows = [
            r
            for r in range(3, target_cost.max_row + 1)
            if clean(target_cost.cell(r, 1).value) == clean(code) and clean(target_cost.cell(r, 2).value) == clean(project_name)
        ]
        if existing_cost_rows:
            statuses.append(f"skipped cost: rows already exist {', '.join(str(r) for r in existing_cost_rows)}")
        else:
            source_cost_rows = [
                r
                for r in range(3, source_cost.max_row + 1)
                if clean(source_cost.cell(r, 1).value) == clean(code) and clean(source_cost.cell(r, 2).value) == clean(project_name)
            ]
            if not source_cost_rows:
                statuses.append("failed cost: approved project rows not found in approval file cost sheet")
            else:
                target_rows = []
                for source_row in source_cost_rows:
                    target_row = target_cost.max_row + 1
                    copy_row_between_sheets(source_cost, source_row, target_cost, target_row)
                    target_rows.append(target_row)
                statuses.append(f"carried cost rows {', '.join(str(r) for r in target_rows)}")

        results.append([None, "CARRY_FORWARD_PROJECT_MASTER", code, project_name, "; ".join(statuses)])
    return results


def apply_project_master_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return

    approvals = read_project_master_approvals(approval_file)
    results: list[list[Any]] = carry_forward_project_master_results(wb, approval_file)
    for approval in approvals:
        action = clean(approval.get("action"))
        try:
            status = apply_project_master_approval(wb, approval)
        except Exception as exc:
            status = f"failed: {exc}"
        results.append(
            [
                approval.get("approval_row"),
                action,
                approval.get("project_code"),
                approval.get("project"),
                status,
            ]
        )

    if results:
        write_project_master_approval_result_sheet(wb, results)


def apply_fix_media_weight_or_month(data_ws, approval: dict[str, Any]) -> str:
    row_value = approval.get("data_row")
    if row_value is None:
        return "failed: missing Data media row"
    row = int(row_value)
    if row < MEDIA_TEMPLATE_ROW or row > data_ws.max_row:
        return f"failed: Data media row {row} is outside sheet range"

    changes = []
    month = month_number(data_ws.cell(row, 5).value)
    if month is None:
        month = parse_month_value(data_ws.cell(row, 3).value)
        if month is not None:
            data_ws.cell(row, 5).value = month
            changes.append("month")

    data_ws.cell(row, 8).value = media_formula_h(row)
    data_ws.cell(row, 9).value = f"=D{row}/H{row}"
    changes.extend(["H formula", "I formula"])

    if month is None:
        return f"partial: filled formulas row {row}, but month could not be inferred"
    return f"applied: filled {', '.join(changes)} row {row}"


def apply_media_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    data_ws = find_sheet(wb, SHEET_MEDIA)
    if data_ws is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_MEDIA}.")

    approvals = read_media_approvals(approval_file)
    results: list[list[Any]] = []
    for approval in approvals:
        action = clean(approval.get("action"))
        try:
            if action == "FIX_MEDIA_WEIGHT_OR_MONTH":
                status = apply_fix_media_weight_or_month(data_ws, approval)
            else:
                status = f"skipped: unsupported action {action}"
        except Exception as exc:
            status = f"failed: {exc}"
        results.append(
            [
                approval.get("approval_row"),
                approval.get("section"),
                action,
                approval.get("data_row"),
                approval.get("project"),
                approval.get("employee"),
                approval.get("mnv"),
                status,
            ]
        )

    if results:
        write_media_approval_result_sheet(wb, results)


def apply_downstream_approvals(wb, approval_file: Path | None) -> None:
    if approval_file is None:
        return
    catalog = find_sheet(wb, SHEET_PROJECT_CATALOG)
    if catalog is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_PROJECT_CATALOG}.")
    capitalization = find_sheet(wb, SHEET_CAPITALIZATION)
    if capitalization is None:
        raise SystemExit(f"Output workbook does not contain sheet {SHEET_CAPITALIZATION}.")

    approvals = read_downstream_approvals(approval_file)
    results: list[list[Any]] = []
    for approval in approvals:
        actions = [clean(action) for action in str(approval.get("action") or "").split(";")]
        statuses = []
        for action in actions:
            if not action:
                continue
            try:
                if action == "ADD_TO_PROJECT_CATALOG":
                    statuses.append(apply_add_to_project_catalog(catalog, approval))
                elif action == "ADD_TO_CAPITALIZATION":
                    statuses.append(apply_add_to_capitalization(capitalization, approval))
                elif action == "ADD_MNV_TO_IT_CHECKING":
                    statuses.append("skipped: ADD_MNV_TO_IT_CHECKING not enabled; Checking Von hoa IT salary base needs manual source validation")
                else:
                    statuses.append(f"skipped: unsupported action {action}")
            except Exception as exc:
                statuses.append(f"failed {action}: {exc}")
        if statuses:
            results.append(
                [
                    approval.get("approval_row"),
                    approval.get("action"),
                    approval.get("project_code"),
                    approval.get("project_name"),
                    "; ".join(statuses),
                ]
            )

    if results:
        write_downstream_approval_result_sheet(wb, results)


def build_it_downstream_checkpoint_data(wb) -> tuple[list[list[Any]], list[list[Any]]]:
    cost_ws = find_sheet(wb, SHEET_IT_COST)
    catalog_ws = find_sheet(wb, SHEET_PROJECT_CATALOG)
    capital_ws = find_sheet(wb, SHEET_CAPITALIZATION)
    checking_ws = find_sheet(wb, SHEET_IT_CHECKING)
    if cost_ws is None or catalog_ws is None or capital_ws is None or checking_ws is None:
        return (
            [
                ["Projects in Chi phí nhân sự IT", 0],
                ["Missing in 1.Danh mục dự án", 0],
                ["Missing in 3.Vốn hóa", 0],
                ["Projects with MNV missing in Checking Vốn hóa IT", 0],
            ],
            [
                [
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "missing downstream sheet(s)",
                    "Restore missing downstream sheet(s) in template",
                    "NO",
                    None,
                ]
            ],
        )

    checking_mnvs = {
        clean(checking_ws.cell(r, 11).value)
        for r in range(4, checking_ws.max_row + 1)
        if clean(checking_ws.cell(r, 11).value)
    }

    projects: dict[tuple[str, str], dict[str, Any]] = {}
    for r in range(3, cost_ws.max_row + 1):
        code = clean(cost_ws.cell(r, 1).value)
        project = clean(cost_ws.cell(r, 2).value)
        if not code or not project:
            continue
        key = (norm(code), norm(project))
        item = projects.setdefault(
            key,
            {
                "project_code": code,
                "project_name": project,
                "system": clean(cost_ws.cell(r, 3).value),
                "bu": clean(cost_ws.cell(r, 6).value),
                "cost_rows": [],
                "mnvs": [],
                "missing_mnvs": [],
            },
        )
        item["cost_rows"].append(r)
        mnv = clean(cost_ws.cell(r, 7).value)
        if mnv:
            item["mnvs"].append(mnv)
            if mnv not in checking_mnvs:
                item["missing_mnvs"].append(mnv)

    summary: list[list[Any]] = []
    detail_rows: list[list[Any]] = []
    missing_catalog = 0
    missing_capital = 0
    missing_checking = 0

    for project in projects.values():
        code = clean(project["project_code"])
        name = clean(project["project_name"])
        catalog_row = find_row_by_clean_value(catalog_ws, 2, code, start_row=2)
        capital_row = find_row_by_clean_value(capital_ws, 2, code, start_row=3)

        catalog_exists = catalog_row is not None
        capital_exists = capital_row is not None
        if not catalog_exists:
            missing_catalog += 1
        if not capital_exists:
            missing_capital += 1
        if project["missing_mnvs"]:
            missing_checking += 1

        if catalog_row is not None:
            catalog_year = catalog_ws.cell(catalog_row, 1).value
            catalog_bu = catalog_ws.cell(catalog_row, 5).value
            catalog_classification = catalog_ws.cell(catalog_row, 6).value
            catalog_start = catalog_ws.cell(catalog_row, 7).value
            catalog_end = catalog_ws.cell(catalog_row, 8).value
            catalog_capitalization = catalog_ws.cell(catalog_row, 12).value
        else:
            catalog_year = 2026
            catalog_bu = project["bu"]
            catalog_classification = None
            catalog_start = None
            catalog_end = None
            catalog_capitalization = None

        issue_parts = []
        recommended_action = []
        if not catalog_exists:
            issue_parts.append("missing in 1.Danh mục dự án")
            recommended_action.append("ADD_TO_PROJECT_CATALOG")
        if not capital_exists:
            issue_parts.append("missing in 3.Vốn hóa")
            recommended_action.append("ADD_TO_CAPITALIZATION")
        if project["missing_mnvs"]:
            issue_parts.append("missing Checking MNV")
            recommended_action.append("ADD_MNV_TO_IT_CHECKING")

        detail_rows.append(
            [
                code,
                name,
                project["system"],
                project["bu"],
                catalog_year,
                catalog_bu,
                catalog_classification,
                catalog_start,
                catalog_end,
                catalog_capitalization,
                ", ".join(str(r) for r in project["cost_rows"]),
                ", ".join(sorted({str(m) for m in project["mnvs"]})),
                catalog_exists,
                capital_exists,
                len(project["missing_mnvs"]),
                ", ".join(project["missing_mnvs"]) if project["missing_mnvs"] else None,
                "; ".join(issue_parts) if issue_parts else None,
                "; ".join(recommended_action) if recommended_action else None,
                "NO" if issue_parts else None,
                None,
            ]
        )

    summary.extend([
        ["Projects in Chi phí nhân sự IT", len(projects)],
        ["Missing in 1.Danh mục dự án", missing_catalog],
        ["Missing in 3.Vốn hóa", missing_capital],
        ["Projects with MNV missing in Checking Vốn hóa IT", missing_checking],
    ])
    return summary, detail_rows


def build_it_cpns_checkpoint_data(wb) -> tuple[list[list[Any]], list[list[Any]]]:
    ts = find_sheet(wb, SHEET_IT)
    cost = find_sheet(wb, SHEET_IT_COST)
    catalog = find_sheet(wb, SHEET_PROJECT_CATALOG)
    if ts is None or cost is None or catalog is None:
        return (
            [[month, None, None, None] for month in range(1, 13)],
            [[None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]],
        )

    catalog_projects = {norm(clean(catalog.cell(r, 3).value)) for r in range(2, catalog.max_row + 1) if clean(catalog.cell(r, 3).value)}

    # Map month number -> cost month input column from row 1.
    cost_month_cols: dict[int, int] = {}
    for c in range(1, cost.max_column + 1):
        v = cost.cell(1, c).value
        if isinstance(v, int) and 1 <= v <= 12:
            cost_month_cols[int(v)] = c

    # Collect timesheet month rows and the corresponding employee columns from row 14.
    ts_employee_cols: dict[str, int] = {}
    for c in range(6, ts.max_column + 1):
        employee_name = clean(ts.cell(14, c).value)
        if employee_name:
            ts_employee_cols[norm(employee_name)] = c

    ts_month_rows: dict[int, list[int]] = {month: [] for month in range(1, 13)}
    for r in range(15, ts.max_row + 1):
        month = month_number(ts.cell(r, 5).value)
        if month in ts_month_rows:
            ts_month_rows[month].append(r)

    summary_rows: list[list[Any]] = []
    for month in range(1, 13):
        ts_rows = ts_month_rows.get(month, [])
        ts_ranges = [f"F{r}:Z{r}" for r in ts_rows]
        ts_formula = f"=SUM({','.join(ts_ranges)})" if ts_ranges else "=0"
        cost_col = cost_month_cols.get(month)
        cost_formula = f"=SUM({get_column_letter(cost_col)}3:{get_column_letter(cost_col)}{cost.max_row})" if cost_col else "=0"
        summary_rows.append([month, ts_formula, cost_formula, f"=C{month + 2}-B{month + 2}"])

    project_issue_rows: dict[str, dict[str, Any]] = {}
    for r in range(3, cost.max_row + 1):
        project_code = clean(cost.cell(r, 1).value)
        raw_project_name = cost.cell(r, 2).value
        project_name = clean(raw_project_name)
        system = clean(cost.cell(r, 3).value)
        bu = clean(cost.cell(r, 6).value)
        employee_code = clean(cost.cell(r, 7).value)
        employee = clean(cost.cell(r, 8).value)
        position = clean(cost.cell(r, 9).value)
        cost_project_sample = clean(cost.cell(r, 2).value)
        if not project_name:
            continue

        project_clean = norm(project_name)
        project_mismatch = isinstance(raw_project_name, str) and clean(raw_project_name) != raw_project_name
        issue_parts: list[str] = []
        recommended_actions: list[str] = []
        if project_mismatch:
            issue_parts.append("criteria text mismatch, likely whitespace")
            recommended_actions.append("CLEAN_COST_PROJECT_TEXT")

        if not issue_parts:
            continue

        record = project_issue_rows.setdefault(
            project_clean,
            {
                "project_code": project_code,
                "project_name": project_name,
                "system": system,
                "bu": bu,
                "employees": [],
                "employee_cols": [],
                "ts_rows": [],
                "input_formulas": [],
                "processed_formulas": [],
                "cost_rows": [],
                "positions": [],
                "month_found": None,
                "issue_parts": [],
                "recommended_actions": [],
                "matched_after_clean": True,
                "cost_project_sample": cost_project_sample,
            },
        )
        record["cost_rows"].append(r)
        if employee:
            record["employees"].append(employee)
        if position:
            record["positions"].append(position)
        if employee_code:
            employee_col_found = ts_employee_cols.get(norm(employee)) if employee else None
        else:
            employee_col_found = None
        ts_row_found = None
        input_formula = None
        processed_formula = None
        month_found = None
        if employee_col_found is not None:
            for month, cost_col in cost_month_cols.items():
                for ts_row in ts_month_rows.get(month, []):
                    if norm(ts.cell(ts_row, 2).value) == project_clean:
                        month_found = month
                        ts_row_found = ts_row
                        input_formula = f"='Timesheet IT'!{get_column_letter(employee_col_found)}{ts_row_found}"
                        processed_formula = f"='Chi phí nhân sự IT'!{get_column_letter(cost_col)}{r}"
                        break
                if month_found is not None:
                    break
        if month_found is None:
            month_found = month_number(cost.cell(1, 14).value) or 1
        if input_formula is None and employee_col_found is not None and ts_month_rows.get(month_found):
            ts_row_found = ts_month_rows[month_found][0]
            input_formula = f"='Timesheet IT'!{get_column_letter(employee_col_found)}{ts_row_found}"
        if processed_formula is None and cost_month_cols.get(month_found):
            processed_formula = f"='Chi phí nhân sự IT'!{get_column_letter(cost_month_cols[month_found])}{r}"
        if employee_col_found is not None:
            record["employee_cols"].append(get_column_letter(employee_col_found))
        if ts_row_found is not None:
            record["ts_rows"].append(ts_row_found)
        if input_formula is not None:
            record["input_formulas"].append(input_formula)
        if processed_formula is not None:
            record["processed_formulas"].append(processed_formula)
        record["issue_parts"].extend(issue_parts)
        record["recommended_actions"].extend(recommended_actions)
        record["month_found"] = month_found if record["month_found"] is None else record["month_found"]
        record["matched_after_clean"] = record["matched_after_clean"] and not project_mismatch

    detail_rows: list[list[Any]] = []
    for record in project_issue_rows.values():
        issue_text = "; ".join(dict.fromkeys(record["issue_parts"]))
        recommended_text = "; ".join(dict.fromkeys(record["recommended_actions"]))
        detail_rows.append(
            [
                record["month_found"],
                record["project_name"],
                ", ".join(dict.fromkeys(record["employees"])) or None,
                ", ".join(dict.fromkeys(record["employee_cols"])) or None,
                ", ".join(str(x) for x in dict.fromkeys(record["ts_rows"])) or None,
                ", ".join(str(x) for x in dict.fromkeys(record["cost_rows"])) or None,
                record["input_formulas"][0] if record["input_formulas"] else None,
                record["processed_formulas"][0] if record["processed_formulas"] else None,
                f"=G{len(detail_rows)+19}-H{len(detail_rows)+19}",
                record["cost_project_sample"],
                record["matched_after_clean"],
                issue_text,
                recommended_text,
                "YES",
                None,
            ]
        )

    # If there are no issues, keep a single informative row so the sheet is not empty.
    if not detail_rows:
        detail_rows.append([None, None, None, None, None, None, None, None, None, None, None, None, None, "NO", None])

    return summary_rows, detail_rows


def build_it_new_project_master_data(wb) -> tuple[list[list[Any]], list[list[Any]]]:
    ts = find_sheet(wb, SHEET_IT)
    cost = find_sheet(wb, SHEET_IT_COST)
    catalog = find_sheet(wb, SHEET_PROJECT_CATALOG)
    if ts is None or cost is None or catalog is None:
        return (
            [["Projects requiring new project master", 0]],
            [[None, None, None, None, None, None, None, None, None, None, None]],
        )

    catalog_projects = {norm(clean(catalog.cell(r, 3).value)) for r in range(2, catalog.max_row + 1) if clean(catalog.cell(r, 3).value)}
    timesheet_projects = collect_timesheet_it_projects_missing_from_catalog(ts, catalog_projects)
    project_rows: dict[str, dict[str, Any]] = {}

    for r in range(3, cost.max_row + 1):
        project_code = clean(cost.cell(r, 1).value)
        project_name = clean(cost.cell(r, 2).value)
        system = clean(cost.cell(r, 3).value)
        bu = clean(cost.cell(r, 6).value)
        if not project_name:
            continue

        project_clean = norm(project_name)
        if project_clean in catalog_projects:
            continue

        item = project_rows.setdefault(
            project_clean,
            {
                "project_code": project_code,
                "project_name": project_name,
                "system": system,
                "bu": bu,
                "cost_rows": [],
                "ts_rows": [],
                "months": set(),
            },
        )
        item["cost_rows"].append(r)

        ts_matches = source_entries_for_project(ts, project_name)
        for source in ts_matches.values():
            item["ts_rows"].extend(source["rows"])
            item["months"].update(source["months"])

    for project_clean, item in timesheet_projects.items():
        record = project_rows.setdefault(
            project_clean,
            {
                "project_code": None,
                "project_name": item.get("project_name"),
                "system": item.get("system"),
                "bu": item.get("bu"),
                "cost_rows": [],
                "ts_rows": [],
                "months": set(),
            },
        )
        if not record.get("project_name"):
            record["project_name"] = item.get("project_name")
        if not record.get("system"):
            record["system"] = item.get("system")
        if not record.get("bu"):
            record["bu"] = item.get("bu")
        record["ts_rows"].extend(item.get("ts_rows", []))
        record["months"].update(item.get("months", set()))

    next_code_num = next_it_project_code_number(wb)
    for record in project_rows.values():
        if not record.get("project_code"):
            record["project_code"] = f"IT{next_code_num:04d}"
            next_code_num += 1

    summary_rows = [["Projects requiring new project master", len(project_rows)]]
    detail_rows: list[list[Any]] = []
    for record in project_rows.values():
        months = sorted(record["months"])
        start_month = months[0] if months else None
        end_month = months[-1] if months else None
        detail_rows.append(
            [
                record["project_name"],
                record["project_code"],
                record["bu"],
                record["system"],
                start_month,
                end_month,
                ", ".join(str(r) for r in dict.fromkeys(record["ts_rows"])) or None,
                ", ".join(str(r) for r in dict.fromkeys(record["cost_rows"])) or None,
                "ADD_PROJECT_MASTER_FIRST",
                "NO",
                None,
            ]
        )

    if not detail_rows:
        detail_rows.append([None, None, None, None, None, None, None, None, None, "NO", None])

    return summary_rows, detail_rows


def write_it_new_project_master_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "IT_New_Project_Master")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Count"],
        summary_rows,
    )
    append_table(
        ws,
        next_row + 1,
        "New Project Master Required",
        [
            "Project name",
            "Suggested project code",
            "Suggested BU",
            "System",
            "Start month",
            "End month",
            "Timesheet rows",
            "Cost rows",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        detail_rows,
    )
    ws.freeze_panes = "A5"
    for col, width in {
        "A": 48,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 12,
        "F": 12,
        "G": 20,
        "H": 18,
        "I": 28,
        "J": 10,
        "K": 18,
    }.items():
        ws.column_dimensions[col].width = width


def write_it_cpns_checkpoint_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Check_IT_CPNS")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Month", "Input Timesheet IT", "Processed Chi phí nhân sự IT", "Diff"],
        summary_rows,
    )
    append_table(
        ws,
        next_row + 1,
        "Details",
        [
            "Month",
            "Project",
            "Employee",
            "Employee column",
            "Timesheet rows",
            "Cost rows matched after clean",
            "Input value",
            "Processed value",
            "Diff",
            "Cost project sample",
            "Matched after clean",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        detail_rows,
    )
    ws.freeze_panes = "A18"
    for col, width in {
        "A": 10,
        "B": 48,
        "C": 24,
        "D": 16,
        "E": 16,
        "F": 18,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 50,
        "L": 18,
        "M": 28,
        "N": 28,
        "O": 12,
        "P": 18,
    }.items():
        ws.column_dimensions[col].width = width


def write_it_downstream_checkpoint_sheet(wb, summary_rows: list[list[Any]], detail_rows: list[list[Any]]) -> None:
    ws = reset_sheet(wb, "Check_IT_Downstream")
    next_row = append_table(
        ws,
        1,
        "Summary",
        ["Metric", "Count"],
        summary_rows,
    )
    append_table(
        ws,
        next_row + 1,
        "Details",
        [
            "Project code",
            "Project name",
            "System",
            "BU",
            "Catalog year",
            "Catalog BU",
            "Catalog classification",
            "Catalog start date",
            "Catalog end date",
            "Catalog capitalization flag",
            "Chi phí nhân sự IT rows",
            "MNVs in cost",
            "Exists in 1.Danh mục dự án",
            "Exists in 3.Vốn hóa",
            "Missing Checking MNV count",
            "Missing Checking MNVs",
            "Issue",
            "Recommended action",
            "Apply?",
            "Approval notes",
        ],
        detail_rows,
    )
    for col, width in {
        "A": 16,
        "B": 52,
        "C": 22,
        "D": 18,
        "E": 14,
        "F": 18,
        "G": 22,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 24,
        "L": 36,
        "M": 22,
        "N": 18,
        "O": 22,
        "P": 18,
        "Q": 26,
        "R": 28,
        "S": 12,
        "T": 18,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A9"


def run(
    download: bool,
    open_after: bool = True,
    it_url: str | None = None,
    media_url: str | None = None,
    approval_file: Path | None = None,
    sx_year: int = 2026,
    sx_month: int = 4,
) -> Path:
    setup_dirs()
    sources = load_source_config()
    if it_url:
        sources["IT"]["url"] = it_url
    if media_url:
        sources["MEDIA"]["url"] = media_url
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    approval_source_wb = None

    print("=" * 60)
    print(f"Von hoa chi phi nhan su 2026 - {datetime.now():%d/%m/%Y %H:%M}")
    print("=" * 60)

    if download:
        print("[1/4] Downloading source files...")
        download_all(sources)
    else:
        print("[1/4] Using local files in data/input/raw")

    print(f"[1.5/4] Building SX staging for {sx_year}-{sx_month:02d}...")
    sx_config = sx_merge.load_source_config()
    sx_paths = sx_merge.refresh_source_files_strict(sx_config)
    sx_rows_raw, sx_warnings = sx_merge.collect_rows_for_month(sx_paths, sx_year, sx_month)
    sx_output = sx_merge.output_path_for_month(sx_year, sx_month)
    sx_output = sx_merge.write_output_workbook(sx_rows_raw, sx_output)
    sx_rows = [normalize_sx_row(row) for row in sx_rows_raw]
    print(f"  SX staging <- {sx_output}")
    print(f"  SX staging rows <- {len(sx_rows)}")
    for warning in sx_warnings:
        print(f"  SX warning: {warning}")

    missing = [k for k, meta in sources.items() if not (DIRS["raw"] / meta["file"]).exists()]
    if missing:
        raise SystemExit(
            f"Missing source files: {missing}\n"
            "Run with --download or place the XLSX files in data/input/raw."
        )

    print("[2/4] Reading sources...")
    employee_lookup = {}

    template = find_template()
    if template is None:
        raise SystemExit(
            "Template file not found.\n"
            f"Put the workbook in: {DIRS['template']}\n"
            f"Expected one of: {', '.join(TEMPLATE_CANDIDATES)}"
        )

    template_wb = load_workbook(template)
    if approval_file is not None and approval_file.exists():
        try:
            approval_source_wb = load_workbook(approval_file, data_only=False)
        except Exception as exc:
            print(f"  approval workbook could not be opened for SX lookup: {exc}")
    approval_file_for_apply = prepare_synced_approval_file(approval_file)

    employee_lookup_source = approval_source_wb if approval_source_wb and SHEET_EMPLOYEE in approval_source_wb.sheetnames else template_wb
    employee_lookup = build_employee_lookup(employee_lookup_source)

    it_source = DIRS["raw"] / sources["IT"]["file"]
    media_source = DIRS["raw"] / sources["MEDIA"]["file"]

    it_sheet_name, it_rows = read_it_sheet(it_source)
    print(f"  IT -> sheet '{it_sheet_name}', {len(it_rows)} rows")

    media_rows = extract_media_rows(media_source, employee_lookup)
    media_sheets = sorted({row["source_sheet"] for row in media_rows})
    print(f"  MEDIA -> {len(media_rows)} rows from sheets: {', '.join(media_sheets) if media_sheets else '(none)'}")

    print("[3/4] Preparing output workbook...")
    output_path = DIRS["final"] / f"von_hoa_{ts}.xlsx"
    shutil.copy2(template, output_path)

    wb = load_workbook(output_path)
    capital_values_wb = load_workbook(output_path, data_only=True)
    restore_carry_forward_sheets_from_approval_file(wb, approval_file_for_apply)
    sync_common_processing_hub_approvals(wb, approval_file_for_apply)
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    if approval_source_wb is not None and SHEET_EMPLOYEE in approval_source_wb.sheetnames and SHEET_EMPLOYEE in wb.sheetnames:
        copy_sheet_content(approval_source_wb[SHEET_EMPLOYEE], wb[SHEET_EMPLOYEE])
    written_it = False
    written_media = False
    payroll_rows = sync_payroll_from_onedrive(wb)
    payroll_ok = all(row[-1] == "OK" for row in payroll_rows)
    print(f"  payroll sync <- {PAYROLL_SOURCE_PATH} ({len(payroll_rows)} sheets, ok={payroll_ok})")

    if SHEET_IT in wb.sheetnames:
        write_it_sheet(wb[SHEET_IT], it_rows, it_source)
        written_it = True
        print(f"  wrote {SHEET_IT}")
    else:
        print(f"  missing sheet: {SHEET_IT}")

    if SHEET_MEDIA in wb.sheetnames:
        write_media_sheet(wb[SHEET_MEDIA], media_rows)
        written_media = True
        print(f"  wrote {SHEET_MEDIA}")
    else:
        print(f"  missing sheet: {SHEET_MEDIA}")

    media_by_bu = {
        "CFA": [row for row in media_rows if norm(row.get("bu")) == "cfa"],
        "CMA": [row for row in media_rows if norm(row.get("bu")) == "cma"],
        "ACCA": [row for row in media_rows if norm(row.get("bu")) == "acca"],
    }
    for bu, rows in media_by_bu.items():
        write_media_audit_sheet(wb, f"Media_{bu}", rows)
        print(f"  audit sheet Media_{bu} <- {len(rows)} rows")

    if approval_file is not None:
        apply_it_mapping_approvals(wb, approval_file_for_apply)
        apply_project_master_approvals(wb, approval_file_for_apply)
        apply_downstream_approvals(wb, approval_file_for_apply)
        apply_media_approvals(wb, approval_file_for_apply)
        apply_it_cost_month_approvals(wb, approval_file_for_apply)
        print(f"  applied approvals from {approval_file_for_apply}")

    if SHEET_SX_TARGET in wb.sheetnames:
        sx_append_rows, sx_skipped_rows = append_sx_to_template(wb, sx_rows, employee_lookup, sx_year, sx_month)
        print(f"  appended SX rows to {SHEET_SX_TARGET}: {sx_append_rows} kept, {sx_skipped_rows} skipped")
    else:
        print(f"  missing sheet: {SHEET_SX_TARGET}")

    apply_sx_approvals(wb, approval_file_for_apply, employee_lookup)

    if SHEET_EMPLOYEE in wb.sheetnames:
        employee_lookup = build_employee_lookup(wb)

    if SHEET_SX_TARGET in wb.sheetnames:
        sx_formula_fixes = sanitize_sx_raw_formulas(wb)
        if sx_formula_fixes:
            print(f"  sanitized {sx_formula_fixes} SX raw formula cell(s)")
        allocation_rows = rebuild_sx_allocation_sheet(wb, employee_lookup, sx_year)
        print(f"  rebuilt {SHEET_SX_ALLOCATION} from raw SX rows: {allocation_rows} rows")
        rewritten_cells = rewrite_timesheet_sx_from_allocation(wb)
        print(f"  rewrote Timesheet SX formulas from {SHEET_SX_ALLOCATION}: {rewritten_cells} cells")
        if remove_worksheet_if_exists(wb, "Data SX consol"):
            print("  removed obsolete sheet: Data SX consol")

    if SHEET_CAPITALIZATION in wb.sheetnames:
        cap_ws = wb[SHEET_CAPITALIZATION]
        cap_summary, cap_details = build_capitalization_month_checkpoint(cap_ws, sx_month)
        write_capitalization_month_checkpoint_sheet(wb, cap_summary, cap_details)
        apply_capitalization_month_approvals(wb, approval_file_for_apply)
        print("  wrote checkpoint sheet: Check_Vonhoa_Month_Block")

    if SHEET_SX_DOWNSTREAM_CHECKPOINT in wb.sheetnames:
        ws = wb[SHEET_SX_DOWNSTREAM_CHECKPOINT]
        ws.delete_rows(1, ws.max_row)
    sx_downstream_summary, sx_downstream_details = build_sx_downstream_checkpoint_data(
        wb,
        sx_rows,
        sx_year,
        sx_month,
        employee_lookup,
        capital_values_wb=capital_values_wb,
    )
    write_sx_downstream_checkpoint_sheet(wb, sx_downstream_summary, sx_downstream_details)
    apply_sx_downstream_approvals(wb, approval_file_for_apply)

    restored_it_checkpoints = restore_missing_checkpoint_sheets_from_latest_output(
        wb,
        output_path,
        [
            "Check_IT_CPNS",
            "Check_IT_Downstream",
            "Check_Media_Timesheet",
        ],
    )
    if restored_it_checkpoints:
        print(f"  restored missing checkpoint sheets from latest output: {', '.join(restored_it_checkpoints)}")

    if "IT_New_Project_Master" not in wb.sheetnames and SHEET_IT in wb.sheetnames and SHEET_IT_COST in wb.sheetnames and SHEET_PROJECT_CATALOG in wb.sheetnames:
        it_new_summary, it_new_details = build_it_new_project_master_data(wb)
        write_it_new_project_master_sheet(wb, it_new_summary, it_new_details)
        print("  rebuilt checkpoint sheet: IT_New_Project_Master")

    if SHEET_IT_COST in wb.sheetnames:
        it_cost_summary, it_cost_details = build_it_cost_month_checkpoint(wb[SHEET_IT_COST], sx_month)
        write_it_cost_month_checkpoint_sheet(wb, it_cost_summary, it_cost_details)
        print("  wrote checkpoint sheet: Check_IT_Cost_Month_Block")

    if SHEET_MEDIA in wb.sheetnames and SHEET_TIMESHEET_MEDIA in wb.sheetnames and SHEET_SALARY_FULLTIME in wb.sheetnames:
        media_summary, media_details, media_source_issues, media_over100 = build_media_checkpoint_data(wb)
        write_media_checkpoint_sheet(wb, media_summary, media_details, media_source_issues, media_over100)
        print("  rebuilt checkpoint sheet: Check_Media_Timesheet")

    if SHEET_IT in wb.sheetnames and SHEET_IT_COST in wb.sheetnames and SHEET_PROJECT_CATALOG in wb.sheetnames:
        it_cpns_summary, it_cpns_details = build_it_cpns_checkpoint_data(wb)
        write_it_cpns_checkpoint_sheet(wb, it_cpns_summary, it_cpns_details)
        print("  rebuilt checkpoint sheet: Check_IT_CPNS")

    if SHEET_IT_COST in wb.sheetnames and SHEET_PROJECT_CATALOG in wb.sheetnames and SHEET_CAPITALIZATION in wb.sheetnames and SHEET_IT_CHECKING in wb.sheetnames:
        it_downstream_summary, it_downstream_details = build_it_downstream_checkpoint_data(wb)
        write_it_downstream_checkpoint_sheet(wb, it_downstream_summary, it_downstream_details)
        print("  rebuilt checkpoint sheet: Check_IT_Downstream")

    write_approval_guide_sheet(wb, output_path, approval_file)
    write_common_processing_hub(wb)
    print("  wrote checkpoint sheets: Check_Payroll, Check_IT_CPNS, IT_New_Project_Master, Check_IT_Cost_Month_Block, Check_IT_Downstream, Check_Media_Timesheet, checkpoint data SX, Check_SX_Downstream, Check_Vonhoa_Month_Block")

    arrange_workbook_sections(wb)

    wb.save(output_path)
    patch_pivot_refresh_flags(output_path)
    if recalculate_workbook_with_excel(output_path):
        print("  recalculated workbook with Excel")

    print("[4/4] Writing staging files...")
    try:
        csv_path = DIRS["staging"] / f"media_rows_{ts}.csv"
        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=MEDIA_AUDIT_FIELDS)
            writer.writeheader()
            writer.writerows(media_rows)

        xlsx_path = DIRS["staging"] / f"media_rows_{ts}.xlsx"
        staging_wb = Workbook()
        staging_ws = staging_wb.active
        staging_ws.title = "media_rows"
        staging_ws.append(MEDIA_AUDIT_FIELDS)
        for row in media_rows:
            staging_ws.append([row.get(field) for field in MEDIA_AUDIT_FIELDS])
        staging_ws.freeze_panes = "A2"
        staging_ws.auto_filter.ref = staging_ws.dimensions
        staging_wb.save(xlsx_path)
    except Exception:
        pass

    print("=" * 60)
    print("DONE")
    print(f"Output: {output_path}")
    print(f"IT written: {written_it}")
    print(f"Media written: {written_media}")
    print(f"Payroll sync: {payroll_ok}")
    print("=" * 60)

    if open_after:
        import subprocess

        subprocess.Popen(["explorer", str(output_path)])

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Von hoa chi phi nhan su 2026 automation")
    parser.add_argument("--download", action="store_true", help="Download the latest source XLSX files first")
    parser.add_argument("--open", action="store_true", dest="open_after", default=True, help="Open the result workbook after finishing")
    parser.add_argument("--no-open", action="store_false", dest="open_after", help="Do not open the result workbook after finishing")
    parser.add_argument("--it-url", dest="it_url", help="Override the IT source URL for this run")
    parser.add_argument("--media-url", dest="media_url", help="Override the media source URL for this run")
    parser.add_argument("--sx-year", type=int, default=2026, help="Year to export for SX staging, default is 2026")
    parser.add_argument("--sx-month", type=int, default=4, help="Month to export for SX staging, default is 4")
    parser.add_argument(
        "--approval-file",
        type=Path,
        help="Workbook containing approval rows. Put YES in Apply? for rows to automate. For SX approvals, pass the matching --sx-year and --sx-month explicitly.",
    )
    args = parser.parse_args()
    run(
        download=args.download,
        open_after=args.open_after,
        it_url=args.it_url,
        media_url=args.media_url,
        approval_file=args.approval_file,
        sx_year=args.sx_year,
        sx_month=args.sx_month,
    )


if __name__ == "__main__":
    main()

