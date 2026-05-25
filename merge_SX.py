from __future__ import annotations

import argparse
import shutil
import json
import re
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter


WORK_DIR = Path(__file__).parent

DIRS = {
    "raw": WORK_DIR / "data" / "input" / "raw",
    "staging": WORK_DIR / "data" / "output" / "staging",
    "logs": WORK_DIR / "logs",
}

DEFAULT_SOURCE_FILES = {
    "ACCA": "ACCA.xlsx",
    "CMA": "CMA.xlsx",
    "CFA": "CFA.xlsx",
}

DEFAULT_SOURCE_URLS = {
    "ACCA": "https://docs.google.com/spreadsheets/d/1ppb8FD9HDR9QJORO13V6FJJPHQmzzNL4Wp8YtpPeyDA/export?format=xlsx",
    "CMA": "https://docs.google.com/spreadsheets/d/16O-oOfPgqt2dmE3_TPEZvSh8zc12hYfik5Onda0cUiU/export?format=xlsx",
    "CFA": "https://docs.google.com/spreadsheets/d/1i2nq5OvJ31B8e5b1xRcUnCldAEbNWvg6lkHBlEz2IWE/export?format=xlsx",
}

CFA_IGNORED_SHEETS = {"drop list", "jan 26", "jan 2026"}

OUTPUT_SHEET_NAME = "Data SX ACCA + CMA + CFA"
OUTPUT_FILE_NAME = "gop_thong_tin_SX_thang_4.xlsx"

MERGE_SX_HEADERS = [
    "Năm",
    "Tháng",
    "Chương trình",
    "Vị trí",
    "Tên nhân viên",
    "Tên sản phẩm",
    "Sản phẩm mới/sản phẩm cũ",
    "Tên dự án",
    "Bộ môn",
    "Đặc tính sản phẩm",
    "Sản phẩm bàn giao",
    "Đơn vị tính",
    "Số lượng actual",
    "KPI standard (h)/1 ĐV quy đổi",
]

MONTH_NAME_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value).strip()
        return text or None
    return value


def norm(value: Any) -> str:
    value = clean(value)
    if value is None:
        return ""
    text = str(value).lower()
    text = "".join(
        ch for ch in re.sub(r"[^\w\s]+", " ", text, flags=re.UNICODE)
        if True
    )
    text = re.sub(r"\s+", " ", text).strip()
    return text


def setup_dirs() -> None:
    for path in DIRS.values():
        path.mkdir(parents=True, exist_ok=True)


def load_source_config() -> dict[str, dict[str, Any]]:
    cfg_path = WORK_DIR / "config" / "sources.json"
    if not cfg_path.exists():
        return {}

    with cfg_path.open("r", encoding="utf-8") as f:
        loaded = json.load(f)

    if not isinstance(loaded, dict):
        return {}

    result: dict[str, dict[str, Any]] = {}
    for key, value in loaded.items():
        if isinstance(value, dict):
            result[key] = value
    return result


def source_url(source_name: str, config: dict[str, dict[str, Any]]) -> str:
    meta = config.get(source_name, {})
    url = clean(meta.get("url"))
    if url:
        return str(url)
    return DEFAULT_SOURCE_URLS[source_name]


def source_file_path(source_name: str, config: dict[str, dict[str, Any]]) -> Path:
    meta = config.get(source_name, {})
    chosen = clean(meta.get("file")) or DEFAULT_SOURCE_FILES[source_name]
    return DIRS["raw"] / str(chosen)


def build_export_url(url: str) -> str:
    if "/export?" in url:
        return url
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        return url
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def download_source(url: str, out_path: Path) -> None:
    export_url = build_export_url(url)
    tmp = out_path.with_name(f".{out_path.stem}.download.xlsx")
    req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as resp, tmp.open("wb") as f:
        shutil.copyfileobj(resp, f)
    try:
        tmp.replace(out_path)
    except PermissionError:
        shutil.copy2(tmp, out_path)
        tmp.unlink(missing_ok=True)


def refresh_source_files(config: dict[str, dict[str, Any]]) -> dict[str, Path]:
    paths: dict[str, Path] = {}
    for source_name in DEFAULT_SOURCE_FILES:
        path = source_file_path(source_name, config)
        url = source_url(source_name, config)
        try:
            download_source(url, path)
        except Exception as exc:
            if not path.exists():
                raise SystemExit(f"Không tải được nguồn {source_name}: {exc}") from exc
            print(f"Warning: failed to refresh {source_name}, using cached file: {exc}")
        paths[source_name] = path
    return paths


def refresh_source_files_strict(config: dict[str, dict[str, Any]]) -> dict[str, Path]:
    paths: dict[str, Path] = {}
    for source_name in DEFAULT_SOURCE_FILES:
        path = source_file_path(source_name, config)
        url = source_url(source_name, config)
        try:
            download_source(url, path)
        except urllib.error.HTTPError as exc:
            if exc.code in {401, 403}:
                raise SystemExit(
                    f"Không thể tải {source_name} từ Google Sheets ({exc.code}). "
                    "Nguồn này cần quyền export/public access."
                ) from exc
            if not path.exists():
                raise SystemExit(f"Không tải được nguồn {source_name}: {exc}") from exc
            print(f"Warning: failed to refresh {source_name}, using cached file: {exc}")
        except Exception as exc:
            if not path.exists():
                raise SystemExit(f"Không tải được nguồn {source_name}: {exc}") from exc
            print(f"Warning: failed to refresh {source_name}, using cached file: {exc}")
        paths[source_name] = path
    return paths


def check_source_access(config: dict[str, dict[str, Any]]) -> list[tuple[str, bool, str]]:
    results: list[tuple[str, bool, str]] = []
    for source_name in DEFAULT_SOURCE_FILES:
        path = source_file_path(source_name, config)
        url = source_url(source_name, config)
        try:
            download_source(url, path)
            results.append((source_name, True, f"OK -> {path.name}"))
        except urllib.error.HTTPError as exc:
            results.append((source_name, False, f"HTTP {exc.code}: {exc.reason}"))
        except Exception as exc:
            results.append((source_name, False, str(exc)))
    return results


def parse_month_from_sheet_title(title: str) -> tuple[int, int] | None:
    text = clean(title)
    if not text:
        return None

    raw = text.lower().replace(".", " ").replace("-", " ").replace("_", " ")
    raw = re.sub(r"\s+", " ", raw).strip()

    m = re.fullmatch(r"([a-z]{3,9})\s+(\d{2,4})", raw)
    if m:
        month_name = m.group(1)
        year_text = m.group(2)
        month = MONTH_NAME_MAP.get(month_name)
        if month is None:
            return None
        year = int(year_text)
        if year < 100:
            year += 2000
        return year, month

    m = re.fullmatch(r"(\d{1,2})[\/\-](\d{2,4})", raw)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return year, month

    m = re.fullmatch(r"(\d{2})(\d{2})", raw)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        year += 2000
        if 1 <= month <= 12:
            return year, month

    m = re.search(r"(\d{1,2})[\/\-](\d{2,4})", raw)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        if 1 <= month <= 12:
            return year, month

    return None


def sheet_uses_program_column(ws) -> bool:
    header = [clean(ws.cell(18, c).value) for c in range(1, min(ws.max_column, 16) + 1)]
    header_text = " ".join(str(v) for v in header if v)
    return "Chương trình".lower() in header_text.lower()


def is_target_sheet(source_name: str, ws) -> tuple[int, int] | None:
    """
    Return (year, month) if the worksheet is a monthly SX sheet we should parse.
    """

    if source_name == "CFA" and str(ws.title).strip().lower() in CFA_IGNORED_SHEETS:
        return None

    parsed = parse_month_from_sheet_title(str(ws.title))
    if parsed is None:
        return None

    header = [clean(ws.cell(18, c).value) for c in range(1, min(ws.max_column, 16) + 1)]
    header_text = " ".join(str(v) for v in header if v)
    required_terms = ["Vị trí", "Tên nhân viên", "Tên sản phẩm"]
    if sheet_uses_program_column(ws):
        required_terms.insert(0, "Chương trình")
    if all(term.lower() in header_text.lower() for term in required_terms):
        return parsed
    return None


def formula_text(cell) -> str | None:
    value = cell.value
    if value is None:
        return None
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def cell_value(data_ws, formula_ws, row: int, col: int) -> Any:
    data_cell = data_ws.cell(row, col).value
    if data_cell is not None:
        return clean(data_cell)
    formula = formula_text(formula_ws.cell(row, col))
    if formula is not None:
        return formula
    return None


def row_has_data(data_ws, row: int, start_col: int = 1, end_col: int = 16) -> bool:
    return any(clean(data_ws.cell(row, c).value) is not None for c in range(start_col, end_col + 1))


def find_header_col(ws, header_name: str) -> int | None:
    target = norm(header_name)
    if not target:
        return None
    for col in range(1, min(ws.max_column, 16) + 1):
        if norm(ws.cell(18, col).value) == target:
            return col
    return None


def is_cma_sheet(ws) -> bool:
    title = str(ws.title).strip()
    return re.match(r"^\d+\.\s+.+\s+-\s+(Sup|FT|PT)$", title, flags=re.IGNORECASE) is not None


def row_month_year(data_ws, row: int) -> tuple[int, int] | None:
    year = cell_value(data_ws, data_ws, row, 1)
    month = cell_value(data_ws, data_ws, row, 2)
    if isinstance(year, (int, float)) and isinstance(month, (int, float)):
        year_i = int(year)
        month_i = int(month)
        if 1 <= month_i <= 12:
            return year_i, month_i
    return None


def extract_rows_from_sheet(source_name: str, source_file: str, data_ws, formula_ws, year: int, month: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    uses_program_column = sheet_uses_program_column(data_ws)
    total_kpi_col = find_header_col(data_ws, "Total KPI")

    for r in range(21, data_ws.max_row + 1):
        if not row_has_data(data_ws, r):
            continue

        if uses_program_column:
            program = cell_value(data_ws, formula_ws, r, 1)
            position = cell_value(data_ws, formula_ws, r, 2)
            employee = cell_value(data_ws, formula_ws, r, 3)
            product = cell_value(data_ws, formula_ws, r, 4)
            product_type_col = 6
            project_col = 7
            department_col = 8
            feature_col = 9
            deliverable_col = 10
            unit_col = 13
            actual_col = 14
            standard_col = 15
        else:
            program = "CFA"
            position = cell_value(data_ws, formula_ws, r, 1)
            employee = cell_value(data_ws, formula_ws, r, 2)
            product = cell_value(data_ws, formula_ws, r, 3)
            product_type_col = 5
            project_col = 6
            department_col = 7
            feature_col = 8
            deliverable_col = 9
            unit_col = 12
            actual_col = 13
            standard_col = 14

        if uses_program_column and norm(program) != norm(source_name):
            continue

        actual = cell_value(data_ws, formula_ws, r, actual_col)
        standard = cell_value(data_ws, formula_ws, r, standard_col)
        total_kpi = cell_value(data_ws, formula_ws, r, total_kpi_col) if total_kpi_col else None

        # Skip title/group rows like "ACCA" that only carry a program label.
        if not employee or not product:
            continue
        if not any([position, employee, product, actual, standard]):
            continue
        if total_kpi_col and total_kpi in {0, 0.0, "0", "0.0"}:
            continue

        row = {
            "Năm": year,
            "Tháng": month,
            "Chương trình": program,
            "Vị trí": position,
            "Tên nhân viên": employee,
            "Tên sản phẩm": product,
            "Sản phẩm mới/sản phẩm cũ": cell_value(data_ws, formula_ws, r, product_type_col),
            "Tên dự án": cell_value(data_ws, formula_ws, r, project_col),
            "Bộ môn": cell_value(data_ws, formula_ws, r, department_col),
            "Đặc tính sản phẩm": cell_value(data_ws, formula_ws, r, feature_col),
            "Sản phẩm bàn giao": cell_value(data_ws, formula_ws, r, deliverable_col),
            "Đơn vị tính": cell_value(data_ws, formula_ws, r, unit_col),
            "Số lượng actual": actual,
            "KPI standard (h)/1 ĐV quy đổi": standard,
            "_source_file": source_file,
            "_source_sheet": data_ws.title,
            "_source_row": r,
        }
        rows.append(row)

    return rows


def extract_rows_from_cma_sheet(source_file: str, data_ws, formula_ws, year: int, month: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    total_kpi_col = find_header_col(data_ws, "Tổng KPI") or find_header_col(data_ws, "Total KPI")

    for r in range(21, data_ws.max_row + 1):
        if not row_has_data(data_ws, r, 1, min(data_ws.max_column, 20)):
            continue

        row_year_month = row_month_year(data_ws, r)
        if row_year_month is None:
            continue
        row_year, row_month = row_year_month
        if row_year != year or row_month != month:
            continue

        program = cell_value(data_ws, formula_ws, r, 3)
        position = cell_value(data_ws, formula_ws, r, 4)
        employee = cell_value(data_ws, formula_ws, r, 5)
        product = cell_value(data_ws, formula_ws, r, 6)
        actual = cell_value(data_ws, formula_ws, r, 16)
        standard = cell_value(data_ws, formula_ws, r, 17)
        total_kpi = cell_value(data_ws, formula_ws, r, total_kpi_col) if total_kpi_col else None

        if norm(program) != "cma":
            continue
        if not employee or not product:
            continue
        if not any([position, employee, product, actual, standard]):
            continue
        if total_kpi_col and total_kpi in {0, 0.0, "0", "0.0"}:
            continue
        if any(isinstance(v, str) and norm(v) == "tong cong" for v in [cell_value(data_ws, formula_ws, r, 17), cell_value(data_ws, formula_ws, r, 18)]):
            continue

        rows.append(
            {
                "Năm": row_year,
                "Tháng": row_month,
                "Chương trình": program,
                "Vị trí": position,
                "Tên nhân viên": employee,
                "Tên sản phẩm": product,
                "Sản phẩm mới/sản phẩm cũ": cell_value(data_ws, formula_ws, r, 8),
                "Tên dự án": cell_value(data_ws, formula_ws, r, 9),
                "Bộ môn": cell_value(data_ws, formula_ws, r, 10),
                "Đặc tính sản phẩm": cell_value(data_ws, formula_ws, r, 11),
                "Sản phẩm bàn giao": cell_value(data_ws, formula_ws, r, 12),
                "Đơn vị tính": cell_value(data_ws, formula_ws, r, 15),
                "Số lượng actual": actual,
                "KPI standard (h)/1 ĐV quy đổi": standard,
                "_source_file": source_file,
                "_source_sheet": data_ws.title,
                "_source_row": r,
            }
        )

    return rows


def collect_rows_for_month(paths: dict[str, Path], year: int, month: int) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    for source_name, path in paths.items():
        data_wb = load_workbook(path, data_only=True)
        formula_wb = load_workbook(path, data_only=False)
        source_rows = 0

        if source_name == "CMA":
            for data_ws, formula_ws in zip(data_wb.worksheets, formula_wb.worksheets):
                if not is_cma_sheet(data_ws):
                    continue
                extracted = extract_rows_from_cma_sheet(path.name, data_ws, formula_ws, year, month)
                rows.extend(extracted)
                source_rows += len(extracted)
        else:
            for data_ws, formula_ws in zip(data_wb.worksheets, formula_wb.worksheets):
                parsed = is_target_sheet(source_name, data_ws)
                if parsed is None:
                    continue

                sheet_year, sheet_month = parsed
                if sheet_month != month or sheet_year != year:
                    continue

                extracted = extract_rows_from_sheet(source_name, path.name, data_ws, formula_ws, sheet_year, sheet_month)
                rows.extend(extracted)
                source_rows += len(extracted)

        if source_rows == 0:
            warnings.append(f"{source_name}: no rows extracted for {year}-{month:02d} from {path.name}")

    rows.sort(key=lambda r: (r.get("N?m"), r.get("Th?ng"), str(r.get("Ch??ng tr?nh") or ""), str(r.get("V? tr?") or ""), str(r.get("T?n nh?n vi?n") or ""), str(r.get("T?n s?n ph?m") or ""), int(r.get("_source_row") or 0)))
    return rows, warnings


def style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_output_workbook(rows: list[dict[str, Any]], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME
    ws.append(MERGE_SX_HEADERS)
    style_header(ws)

    for output_row_idx, row in enumerate(rows, start=2):
        values: list[Any] = []
        source_row = int(row.get("_source_row") or output_row_idx)
        for col_idx, header in enumerate(MERGE_SX_HEADERS, start=1):
            value = row.get(header)
            if isinstance(value, str) and value.startswith("=") and header in {
                "Đơn vị tính",
                "KPI standard (h)/1 ĐV quy đổi",
            }:
                source_col = "M" if header == "Đơn vị tính" else "O"
                dest_col = get_column_letter(col_idx)
                try:
                    value = Translator(value, origin=f"{source_col}{source_row}").translate_formula(
                        dest=f"{dest_col}{output_row_idx}"
                    )
                except Exception:
                    pass
            values.append(value)
        ws.append(values)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 10,
        "B": 8,
        "C": 14,
        "D": 12,
        "E": 20,
        "F": 24,
        "G": 18,
        "H": 20,
        "I": 15,
        "J": 18,
        "K": 18,
        "L": 16,
        "M": 16,
        "N": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(
            f"{output_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{output_path.suffix}"
        )
        wb.save(fallback)
        print(f"Warning: output locked, saved to {fallback}")
        return fallback


def output_path_for_month(year: int, month: int) -> Path:
    return DIRS["staging"] / f"gop_thong_tin_SX_{year}_thang_{month:02d}.xlsx"


def run(year: int = 2026, month: int = 4) -> Path:
    setup_dirs()

    config = load_source_config()
    paths = refresh_source_files_strict(config)
    if not paths:
        raise SystemExit(f"Không tìm thấy file nguồn nào trong {DIRS['raw']}.")

    rows, warnings = collect_rows_for_month(paths, year, month)
    output_path = output_path_for_month(year, month)
    output_path = write_output_workbook(rows, output_path)

    print(f"Output: {output_path}")
    print(f"Rows: {len(rows)}")
    for warning in warnings:
        print(f"Warning: {warning}")

    return output_path


def check_sources() -> int:
    setup_dirs()
    config = load_source_config()
    results = check_source_access(config)
    failed = 0
    for source_name, ok, message in results:
        status = "OK" if ok else "FAIL"
        if not ok:
            failed += 1
        print(f"{source_name}: {status} - {message}")
    return failed


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge SX source files into a staging workbook")
    parser.add_argument("--month", type=int, default=4, help="Month to export, default is 4")
    parser.add_argument("--year", type=int, default=2026, help="Year to export, default is 2026")
    parser.add_argument("--check-sources", action="store_true", help="Check each SX source independently and exit")
    args = parser.parse_args()
    if args.check_sources:
        raise SystemExit(check_sources())
    run(year=args.year, month=args.month)


if __name__ == "__main__":
    main()
